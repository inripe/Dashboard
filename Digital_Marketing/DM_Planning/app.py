"""
INRIPE DM PLANNING DASHBOARD — v4
Business intelligence view of P4. Channel Plan

v3  reads the workbook live from SharePoint instead of a copy committed to the
    repo, falling back to a local file when SharePoint is not configured.
v4  converts revenue to AED using the T5. FX Rates table. Revenue is entered in
    each market's own currency and budget in AED, so without this every total
    across markets was adding AED to SAR to QAR at face value. ROAS is
    recomputed from the converted revenue rather than read from the sheet,
    where it divided local-currency revenue by an AED budget.
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook
from pathlib import Path

import sharepoint_loader as SP

st.set_page_config(page_title="Inripe DM Planning 2026", page_icon="🌿", layout="wide")

st.markdown("""
<style>
.metric-card{background:#f8f9fa;border-radius:10px;padding:16px;border-left:4px solid #1A6B4A}
.metric-label{font-size:12px;color:#666;margin-bottom:4px}
.metric-val{font-size:24px;font-weight:600;color:#1A6B4A}
.metric-sub{font-size:12px;color:#888;margin-top:2px}
.insight-box{background:#E8F5EE;border-radius:8px;padding:12px 16px;border-left:4px solid #1A6B4A;margin:8px 0}
.warn-box{background:#FFF3CD;border-radius:8px;padding:12px 16px;border-left:4px solid #854F0B;margin:8px 0}
.red-box{background:#FFEBEE;border-radius:8px;padding:12px 16px;border-left:4px solid #A32D2D;margin:8px 0}
</style>
""", unsafe_allow_html=True)

EXCEL_FALLBACKS = [
    "Digital_Marketing/DM_Planning/DM_Planing_Tracking_2026_25JUL26_V1_21.xlsx",
    "DM_Planing_Tracking_2026_25JUL26_V1_21.xlsx",
]
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
MON_COL = {m: 3+i for i,m in enumerate(MONTHS)}
GREEN="#1A6B4A"; BLUE="#1B4F8A"; PURPLE="#534AB7"; AMBER="#854F0B"; RED="#A32D2D"

MARKETS = {
    "UAE":   {"rows":{"tgt_rev":7,"tgt_ord":11,"api_ord":16,"meta_ord":28,"tt_ord":44,
                      "tot_ord":58,"tot_bud":61,"roas":62,"cac":63,
                      "api_bud":22,"meta_bud":37,"tt_bud":53,
                      "api_cac":23,"api_roas":24,"meta_roas":40,
                      "api_cap":14,"gap":18,"bsk":9,"tgt_days":10,"tgt_units":6}},
    "KSA":   {"rows":{"tgt_rev":69,"tgt_ord":73,"api_ord":78,"meta_ord":90,"tt_ord":106,
                      "tot_ord":120,"tot_bud":123,"roas":124,"cac":125,
                      "api_bud":84,"meta_bud":99,"tt_bud":115,
                      "api_cac":85,"api_roas":86,"meta_roas":102,
                      "api_cap":76,"gap":80,"bsk":71,"tgt_days":72,"tgt_units":67}},
    "Qatar": {"rows":{"tgt_rev":131,"tgt_ord":135,"api_ord":140,"meta_ord":152,"tt_ord":168,
                      "tot_ord":182,"tot_bud":185,"roas":186,"cac":187,
                      "api_bud":146,"meta_bud":161,"tt_bud":177,
                      "api_cac":147,"api_roas":148,"meta_roas":164,
                      "api_cap":138,"gap":142,"bsk":133,"tgt_days":134,"tgt_units":129}},
    "Egypt": {"rows":{"tgt_rev":193,"tgt_ord":196,"api_ord":201,"meta_ord":214,"tt_ord":230,
                      "tot_ord":245,"tot_bud":248,"roas":249,"cac":250,
                      "api_bud":208,"meta_bud":223,"tt_bud":239,
                      "api_cac":209,"api_roas":210,"meta_roas":226,
                      "api_cap":199,"gap":203,"bsk":195,"tgt_days":196,"tgt_units":191}},
}


# ── CURRENCY ─────────────────────────────────────────────────────────
# Revenue is entered in each market's own currency; budget is entered in AED.
# Reporting currency is AED, so revenue converts and budget does not.
# Mirrors the same table and rule used by the tracking dashboard's engine.
FX_SHEET = "T5. FX Rates"


def read_fx(wb):
    """Read T5. FX Rates -> {(market, 'YYYY-MM' or None): rate}.

    A missing sheet is not an error - the app falls back to 1.0 and says so,
    rather than converting silently or refusing to load.
    """
    if FX_SHEET not in wb.sheetnames:
        return {}, "no FX table — revenue read as if already in AED"
    ws = wb[FX_SHEET]
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            hdr[str(v).strip().lower()] = c
    mc = hdr.get("market")
    rc = next((hdr[k] for k in hdr if "rate" in k), None)
    moc = hdr.get("month")
    if not mc or not rc:
        return {}, "FX table found but Market / Rate columns not recognised"
    out = {}
    for r in range(2, ws.max_row + 1):
        m = ws.cell(r, mc).value
        try:
            rate = float(ws.cell(r, rc).value)
        except (TypeError, ValueError):
            continue
        if not m or rate <= 0:
            continue
        key_month = None
        if moc:
            mv = ws.cell(r, moc).value
            if mv not in (None, ""):
                try:
                    key_month = pd.to_datetime(mv).strftime("%Y-%m")
                except Exception:
                    key_month = str(mv).strip()
        out[(str(m).strip(), key_month)] = rate
    if not out:
        return {}, "FX table is empty"
    return out, ", ".join(f"{k[0]} x{v:.4f}" for k, v in out.items() if k[1] is None)


def fx_for(fx, market, month_label):
    """Month-specific rate first, then the market default, then 1.0."""
    if not fx:
        return 1.0
    if month_label and month_label != "FY":
        try:
            key = pd.to_datetime(f"{month_label} 2026").strftime("%Y-%m")
            if (market, key) in fx:
                return fx[(market, key)]
        except Exception:
            pass
    return fx.get((market, None), 1.0)


# Rows holding money in the market's own currency. Budget rows are excluded
# on purpose - they are already AED.
REVENUE_KEYS = {"tgt_rev"}


@st.cache_data(ttl=60)
def load_data():
    """SharePoint first, local copy second. Returns (data, source_label, fx_note)."""
    if SP.is_configured():
        buf, meta = SP.fetch_workbook()
        label = f"SharePoint · {meta['name']}"
        if meta.get("modified"):
            label += f" · edited {meta['modified'][:16].replace('T',' ')}"
            if meta.get("modified_by"):
                label += f" by {meta['modified_by']}"
        wb = load_workbook(buf, data_only=True)
    else:
        path = next((Path(p) for p in EXCEL_FALLBACKS if Path(p).exists()), None)
        if path is None:
            return None, None
        wb = load_workbook(path, data_only=True)
        label = f"local file · {path}"

    fx, fx_note = read_fx(wb)
    ws = wb['P4. Channel Plan']
    data = {}
    for mkt, cfg in MARKETS.items():
        data[mkt] = {}
        for key, row in cfg["rows"].items():
            data[mkt][key] = {}
            for mo, col in MON_COL.items():
                raw = ws.cell(row, col).value
                try: data[mkt][key][mo] = float(raw) if raw not in (None,"","—") else 0
                except: data[mkt][key][mo] = 0
            fy = ws.cell(row, 15).value
            try: data[mkt][key]["FY"] = float(fy) if fy not in (None,"","—") else 0
            except: data[mkt][key]["FY"] = 0
            if data[mkt][key]["FY"] == 0:
                data[mkt][key]["FY"] = sum(data[mkt][key][m] for m in MONTHS)

    # Convert revenue to AED. Budget is untouched. ROAS is then recomputed from
    # the converted figure rather than read from the sheet, because the sheet's
    # ROAS divides local-currency revenue by an AED budget.
    for mkt in data:
        for key in REVENUE_KEYS:
            if key not in data[mkt]:
                continue
            for per in list(data[mkt][key]):
                data[mkt][key][per] *= fx_for(fx, mkt, per)
        if "roas" in data[mkt] and "tot_bud" in data[mkt]:
            for per in list(data[mkt]["roas"]):
                b = data[mkt]["tot_bud"].get(per, 0)
                data[mkt]["roas"][per] = (data[mkt]["tgt_rev"].get(per, 0) / b) if b else 0
    return data, label, fx_note


def g(data, mkt, key, mo):
    try: return data[mkt][key][mo] or 0
    except: return 0

def agg(data, mkts, key, mo):
    return sum(g(data,m,key,mo) for m in mkts)

def fmt(n, prefix="", suffix="", dec=0):
    if not n: return "—"
    if abs(n)>=1_000_000: return f"{prefix}{n/1_000_000:.1f}M{suffix}"
    if abs(n)>=1_000: return f"{prefix}{n/1_000:.0f}K{suffix}"
    return f"{prefix}{round(n,dec) if dec else int(round(n))}{suffix}"


try:
    data, SOURCE, FX_NOTE = load_data()
except Exception as e:
    st.error(f"Cannot load the workbook.\n\n{e}")
    if not SP.is_configured():
        st.caption("SharePoint is not configured. Missing settings: "
                   + ", ".join(SP.missing_keys()))
    st.stop()

if not data:
    st.error("Workbook not found. Configure SharePoint, or place the file at: "
             + EXCEL_FALLBACKS[0])
    st.stop()

if FX_NOTE.startswith("no FX") or "not recognised" in FX_NOTE or "empty" in FX_NOTE:
    st.markdown(
        f"<div class='warn-box'>⚠️ <b>Currency not converted</b> — {FX_NOTE}. "
        f"Revenue is entered in each market's own currency, so any total across "
        f"markets is mixing currencies. Add a '{FX_SHEET}' sheet with Market, "
        f"Currency and Rate to AED columns.</div>", unsafe_allow_html=True)

# ── HEADER ───────────────────────────────────────────────────────────
st.markdown(f"""
<div style='background:{GREEN};padding:16px 24px;border-radius:10px;margin-bottom:20px;display:flex;justify-content:space-between;align-items:center'>
<div>
<div style='color:white;font-size:20px;font-weight:600'>🌿 DM Planning Intelligence · Inripe 2026</div>
<div style='color:#9FE1CB;font-size:13px;margin-top:2px'>Channel plan · Budget allocation · Market strategy · Capacity analysis</div>
</div>
<div style='color:#9FE1CB;font-size:12px;text-align:right'>P4. Channel Plan<br>Auto-refreshes every 60s</div>
</div>""", unsafe_allow_html=True)

# ── SELECTORS ────────────────────────────────────────────────────────
c1,c2,c3 = st.columns([1,1,3])
sel_mkt = c1.selectbox("Market", ["All Markets","UAE","KSA","Qatar","Egypt"])
active_months = [m for m in MONTHS if any(g(data,mkt,"tot_ord",m)>0 for mkt in MARKETS)]
sel_mo = c2.selectbox("View", ["Full Year"]+active_months)
mo = "FY" if sel_mo=="Full Year" else sel_mo
mkts = list(MARKETS.keys()) if sel_mkt=="All Markets" else [sel_mkt]

tot_ord=agg(data,mkts,"tot_ord",mo); tot_rev=agg(data,mkts,"tgt_rev",mo)
tot_bud=agg(data,mkts,"tot_bud",mo); tot_days=agg(data,mkts,"tgt_days",mo)
tot_units=agg(data,mkts,"tgt_units",mo)
roas=tot_rev/tot_bud if tot_bud>0 else 0
cac=tot_bud/tot_ord if tot_ord>0 else 0
daily=tot_ord/tot_days if tot_days>0 else 0
bsk=next((g(data,m,"bsk",mo) for m in mkts if g(data,m,"bsk",mo)>0), None) or next((g(data,m,"bsk","Jul") for m in mkts if g(data,m,"bsk","Jul")>0), 0)

# ── KPI STRIP ────────────────────────────────────────────────────────
k = st.columns(6)
for col, label, val, sub in zip(k, [
    "Target Orders","Target Revenue","Total Budget","Blended ROAS","Blended CAC","Daily Target"],[
    fmt(tot_ord),fmt(tot_rev,"AED "),fmt(tot_bud,"AED "),
    f"{roas:.1f}x" if roas else "—",fmt(cac,"AED ",dec=1),
    f"{daily:.0f}/day" if daily else "—"],[
    "Units planned","All markets combined","Across all channels",
    "Revenue per AED spent","Budget per order","Orders per working day"]):
    col.markdown(f"""<div class='metric-card'>
    <div class='metric-label'>{label}</div>
    <div class='metric-val'>{val}</div>
    <div class='metric-sub'>{sub}</div></div>""", unsafe_allow_html=True)

st.markdown("---")

# ── INSIGHT ALERTS ───────────────────────────────────────────────────
api_dep = agg(data,mkts,"api_ord",mo)/tot_ord*100 if tot_ord else 0
meta_dep = agg(data,mkts,"meta_ord",mo)/tot_ord*100 if tot_ord else 0

if api_dep > 70:
    st.markdown(f"<div class='insight-box'>✅ <b>API-led plan</b> — {api_dep:.0f}% of orders from organic API capacity. Low paid channel dependency. Cost-efficient.</div>", unsafe_allow_html=True)
if meta_dep > 60:
    st.markdown(f"<div class='warn-box'>⚠️ <b>High Meta dependency</b> — {meta_dep:.0f}% of orders depend on paid Meta. Budget risk if CPM increases.</div>", unsafe_allow_html=True)
if roas < 5:
    st.markdown(f"<div class='red-box'>🔴 <b>Low ROAS alert</b> — {roas:.1f}x blended ROAS. Review channel budget allocation.</div>", unsafe_allow_html=True)
if roas >= 15:
    st.markdown(f"<div class='insight-box'>✅ <b>Strong ROAS</b> — {roas:.1f}x blended. Plan is budget-efficient.</div>", unsafe_allow_html=True)

# ── ROW 1: CHANNEL MIX + MARKET REVENUE ──────────────────────────────
col1, col2 = st.columns(2)

with col1:
    st.subheader("Channel order mix")
    api_o=agg(data,mkts,"api_ord",mo); meta_o=agg(data,mkts,"meta_ord",mo); tt_o=agg(data,mkts,"tt_ord",mo)
    fig = go.Figure(go.Pie(
        labels=["API","Meta","TikTok"],
        values=[api_o,meta_o,tt_o],
        hole=0.5,
        marker_colors=[GREEN,PURPLE,AMBER],
        textinfo="label+percent",
        textfont_size=13,
    ))
    fig.add_annotation(text=f"<b>{fmt(tot_ord)}</b><br>orders",x=0.5,y=0.5,
                       font_size=14,showarrow=False,align="center")
    fig.update_layout(height=320,margin=dict(t=10,b=10,l=10,r=10),showlegend=False)
    st.plotly_chart(fig, use_container_width=True)

with col2:
    st.subheader("Revenue by market")
    mkt_rev = {m: g(data,m,"tgt_rev",mo) for m in mkts if g(data,m,"tgt_rev",mo)>0}
    if mkt_rev:
        fig2 = go.Figure(go.Bar(
            x=list(mkt_rev.keys()), y=list(mkt_rev.values()),
            marker_color=[GREEN,BLUE,PURPLE,AMBER][:len(mkt_rev)],
            text=[fmt(v,"AED ") for v in mkt_rev.values()],
            textposition="outside", textfont_size=12,
        ))
        fig2.update_layout(height=320,margin=dict(t=10,b=10,l=10,r=10),
                           yaxis_title="AED",xaxis_title="",
                           plot_bgcolor="white",yaxis_gridcolor="#f0f0f0")
        st.plotly_chart(fig2, use_container_width=True)

# ── ROW 2: BUDGET EFFICIENCY + GAP ANALYSIS ──────────────────────────
col3, col4 = st.columns(2)

with col3:
    st.subheader("ROAS by market")
    roas_data = [(m, g(data,m,"roas",mo), g(data,m,"tot_bud",mo))
                 for m in mkts if g(data,m,"tot_ord",mo)>0]
    if roas_data:
        fig3 = go.Figure()
        colors=[GREEN,BLUE,PURPLE,AMBER]
        for i,(m,r,b) in enumerate(roas_data):
            fig3.add_trace(go.Bar(name=m,x=[m],y=[r],
                marker_color=colors[i%4],
                text=[f"{r:.1f}x"],textposition="outside",textfont_size=13))
        fig3.add_hline(y=10,line_dash="dash",line_color="gray",
                       annotation_text="10x target",annotation_position="right")
        fig3.update_layout(height=320,margin=dict(t=10,b=10,l=10,r=10),
                           yaxis_title="ROAS",showlegend=False,
                           plot_bgcolor="white",yaxis_gridcolor="#f0f0f0")
        st.plotly_chart(fig3, use_container_width=True)

with col4:
    st.subheader("API capacity vs paid gap")
    gap_data = [(m, g(data,m,"api_cap",mo), g(data,m,"gap",mo))
                for m in mkts if g(data,m,"tgt_ord",mo)>0]
    if gap_data:
        fig4 = go.Figure()
        fig4.add_trace(go.Bar(name="API (organic)",
            x=[d[0] for d in gap_data], y=[d[1] for d in gap_data],
            marker_color=GREEN, text=[fmt(d[1]) for d in gap_data],
            textposition="inside",textfont_color="white"))
        fig4.add_trace(go.Bar(name="Gap (paid)",
            x=[d[0] for d in gap_data], y=[d[2] for d in gap_data],
            marker_color=AMBER, text=[fmt(d[2]) for d in gap_data],
            textposition="inside",textfont_color="white"))
        fig4.update_layout(barmode="stack",height=320,
                           margin=dict(t=10,b=10,l=10,r=10),
                           yaxis_title="Orders",plot_bgcolor="white",
                           yaxis_gridcolor="#f0f0f0",legend=dict(orientation="h",y=1.1))
        st.plotly_chart(fig4, use_container_width=True)

# ── ROW 3: MONTHLY TREND ─────────────────────────────────────────────
st.subheader("Monthly plan — orders and budget trend")
trend = []
for m in MONTHS:
    o = sum(g(data,mkt,"tot_ord",m) for mkt in mkts)
    b = sum(g(data,mkt,"tot_bud",m) for mkt in mkts)
    r = sum(g(data,mkt,"tgt_rev",m) for mkt in mkts)
    if o>0: trend.append({"Month":m,"Orders":o,"Budget":b,"Revenue":r})

if trend:
    df = pd.DataFrame(trend)
    fig5 = make_subplots(specs=[[{"secondary_y":True}]])
    fig5.add_trace(go.Bar(x=df["Month"],y=df["Orders"],name="Orders",
                          marker_color=GREEN,opacity=0.85), secondary_y=False)
    fig5.add_trace(go.Scatter(x=df["Month"],y=df["Budget"],name="Budget (AED)",
                              mode="lines+markers",line_color=AMBER,
                              marker_size=8,line_width=2), secondary_y=True)
    fig5.update_layout(height=300,margin=dict(t=10,b=10,l=10,r=10),
                       plot_bgcolor="white",yaxis_gridcolor="#f0f0f0",
                       legend=dict(orientation="h",y=1.1))
    fig5.update_yaxes(title_text="Orders",secondary_y=False)
    fig5.update_yaxes(title_text="Budget (AED)",secondary_y=True)
    st.plotly_chart(fig5, use_container_width=True)

# ── ROW 4: FULL PLAN TABLE ───────────────────────────────────────────
st.subheader("Full plan detail")
rows=[]
for m in mkts:
    for mo2 in MONTHS:
        o=g(data,m,"tot_ord",mo2); b=g(data,m,"tot_bud",mo2); r=g(data,m,"tgt_rev",mo2)
        if o>0:
            rows.append({"Market":m,"Month":mo2,
                "Target Orders":int(o),"Target Revenue":int(r),"Budget (AED)":int(b),
                "API Orders":int(g(data,m,"api_ord",mo2)),
                "Meta Orders":int(g(data,m,"meta_ord",mo2)),
                "TikTok Orders":int(g(data,m,"tt_ord",mo2)),
                "ROAS":f"{g(data,m,'roas',mo2):.1f}x",
                "CAC":f"{g(data,m,'cac',mo2):.1f}",
                "Daily Target":f"{o/g(data,m,'tgt_days',mo2):.0f}" if g(data,m,"tgt_days",mo2)>0 else "—"})
if rows:
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

st.caption(f"Source: {SOURCE} · P4. Channel Plan · revenue converted to AED "
           f"({FX_NOTE}) · budget already AED · auto-refreshes every 60s")
