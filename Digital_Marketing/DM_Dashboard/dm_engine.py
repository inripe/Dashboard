"""
INRIPE DM ENGINE V2

Pure calculation against DM_Model_2026_V2. No Streamlit import, so every figure
can be verified without launching the app: run audit.py.

The workbook is long-format and register-driven — a market or channel is a row,
never a hardcoded block — so nothing here names a market or channel either.
Adding either to the workbook makes it appear in the dashboard with no code
change.

Rules carried from the workbook
-------------------------------
- Missing is not zero. None propagates; it is never scored as 0%.
- Every percentage states the basis it was measured against.
- Paced plan = month plan x days elapsed / days in month. Ratios are never paced.
- Polarity is declared per metric. Spend and budget carry no verdict: spending
  more is neither good nor bad on its own, CPA and ROAS judge whether it bought
  anything.
- Revenue is entered in the market's own currency, spend in AED except Egypt.
  Everything reports in AED.
- Plan budget is a ceiling on actual spend.
- ROAS is revenue / spend. Always. No AOV anywhere.
"""

from __future__ import annotations

import calendar
import datetime as dt
from dataclasses import dataclass, field
from typing import Optional

import numpy as np
import pandas as pd

# ── sheet and column names, in one place ─────────────────────────────
SH_TARGETS = "P1. Targets"
SH_CAPACITY = "P2. Capacity"
SH_CHANNEL = "P3. Channel plan"
SH_SUMMARY = "S1. Summary"
SH_ACTUALS = "A1. Actuals"
SH_SETUP = "R1. Setup"

# Setup carries four tables on one sheet. Each is found by its header text
# rather than a fixed cell, so moving a block down does not break the loader.
SETUP_BLOCKS = {"markets": ("Markets", "Market"),
                "channels": ("Channels", "Channel"),
                "fx": ("FX overrides", "Market")}

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_NO = {m: i + 1 for i, m in enumerate(MONTHS)}

M_ORDERS, M_UNITS = "Orders", "Units"
M_REVENUE, M_SPEND = "Revenue", "Spend"
M_MSG_CUST, M_MSG_LEAD = "Messages to Customers", "Messages to Leads"
M_MSG_RECV = "Messages Received"
# Outbound messages are two metrics in the sheet, never one: P1 plans customer
# and lead volume against different lists and cycles. These names are lists so
# that stays true everywhere downstream.
M_MSG_SENT = [M_MSG_CUST, M_MSG_LEAD]
M_MSG_ALL = [M_MSG_CUST, M_MSG_LEAD, M_MSG_RECV]

# Higher is better / lower is better / no verdict at all.
POLARITY = {
    "orders": "up", "units": "up", "revenue": "up", "roas": "up", "cr": "up",
    "cpa": "down",
    "spend": "neutral", "budget": "neutral", "messages": "neutral",
}

TH_UP = (0.90, 0.70)        # >=90% good, >=70% watch
TH_DOWN = (1.05, 1.20)      # <=105% good, <=120% watch
COVERAGE_MIN = 0.90
IMPLAUSIBLE = 300.0

GREEN, AMBER, RED, GREY, BLUE = "#0F6E56", "#854F0B", "#A32D2D", "#8A8A8A", "#185FA5"


# ─────────────────────────────────────────────────────────────────────
# FORMATTING
# ─────────────────────────────────────────────────────────────────────
def fmt(n, prefix="", suffix="", dec=0) -> str:
    """None renders n/a. Zero renders 0. One decimal on K and M so two
    different figures never collapse to the same string."""
    if n is None or (isinstance(n, float) and (np.isnan(n) or np.isinf(n))):
        return "n/a"
    if abs(n) >= 1_000_000:
        return f"{prefix}{n/1_000_000:.2f}M{suffix}"
    if abs(n) >= 10_000:
        return f"{prefix}{n/1_000:.1f}K{suffix}"
    if abs(n) >= 1_000:
        return f"{prefix}{n:,.0f}{suffix}"
    return f"{prefix}{n:,.{dec}f}{suffix}"


def pct(a, b) -> Optional[float]:
    if a is None or b in (None, 0) or (isinstance(b, float) and np.isnan(b)):
        return None
    return a / b * 100.0


def fmt_pct(a, b, dec=0) -> str:
    v = pct(a, b)
    return "n/a" if v is None else f"{v:.{dec}f}%"


def delta_pct(a, b) -> Optional[float]:
    """True change. Identical inputs give 0, not 100."""
    if a is None or b in (None, 0):
        return None
    return (a - b) / b * 100.0


def div(a, b):
    if a is None or b in (None, 0):
        return None
    return a / b


def nsum(*vals):
    """Sum that stays None when every input is missing."""
    present = [v for v in vals if v is not None and not (isinstance(v, float) and np.isnan(v))]
    return None if not present else float(sum(present))


# ─────────────────────────────────────────────────────────────────────
# VERDICTS
# ─────────────────────────────────────────────────────────────────────
@dataclass
class Verdict:
    label: str
    color: str
    ratio: Optional[float]
    basis: str
    scored: bool = True
    note: str = ""


def plausible(actual, ratio, key) -> tuple[bool, str]:
    """Catch a data entry fault before it renders as a verdict. A misplaced
    decimal is the most common real error in a hand-kept workbook, and without
    this it shows up green."""
    if ratio is not None and ratio > IMPLAUSIBLE:
        return False, f"{ratio:,.0f}% of plan is not a real result"
    if key == "cpa" and actual is not None and actual <= 0:
        return False, "acquisition cost cannot be zero"
    if key == "cr" and actual is not None and actual > 100:
        return False, "conversion rate cannot exceed 100%"
    if key == "roas" and actual is not None and actual > 500:
        return False, f"{actual:,.0f}x return is not a real result"
    return True, ""


def rag(actual, basis_value, direction="up", basis="", gated=False,
        gate_note="", key="") -> Verdict:
    if gated:
        return Verdict("n/a", GREY, None, basis, False, gate_note)
    r = pct(actual, basis_value)
    if r is None:
        return Verdict("n/a", GREY, None, basis, False, "no basis to compare against")

    ok, why = plausible(actual, r, key)
    if not ok:
        return Verdict("check data", RED, r, basis, False, why)

    if direction == "neutral":
        return Verdict(f"{r:.0f}% of {basis or 'plan'}", GREY, r, basis, False)
    if direction == "up":
        g, a = TH_UP[0] * 100, TH_UP[1] * 100
        if r >= g:
            return Verdict(f"{r:.0f}% on plan", GREEN, r, basis)
        if r >= a:
            return Verdict(f"{r:.0f}% watch", AMBER, r, basis)
        return Verdict(f"{r:.0f}% behind", RED, r, basis)
    if direction == "down":
        g, a = TH_DOWN[0] * 100, TH_DOWN[1] * 100
        if r <= g:
            return Verdict(f"{r:.0f}% on plan", GREEN, r, basis)
        if r <= a:
            return Verdict(f"{r:.0f}% watch", AMBER, r, basis)
        return Verdict(f"{r:.0f}% over", RED, r, basis)
    raise ValueError(f"unknown direction {direction!r}")


def corr_band(r) -> str:
    if r is None or (isinstance(r, float) and np.isnan(r)):
        return "not enough data"
    a = abs(r)
    return "weak" if a < 0.30 else "moderate" if a < 0.60 else "strong"


# ─────────────────────────────────────────────────────────────────────
# LOAD
# ─────────────────────────────────────────────────────────────────────
@dataclass
class Model:
    targets: pd.DataFrame
    capacity: pd.DataFrame
    channel: pd.DataFrame
    summary: pd.DataFrame
    actuals: pd.DataFrame
    vs: pd.DataFrame
    markets: pd.DataFrame
    channels: pd.DataFrame
    fx: dict
    fx_note: str

    def market_list(self) -> list:
        """Every registered market, plus any that appears in the data. An Active
        column is honoured if present and its absence is not an error — V3's
        Markets table does not carry one."""
        d = self.markets
        if "Active" in d.columns:
            reg = d[d["Active"].astype(str).str.lower() == "yes"]["Market"].tolist()
        else:
            reg = d["Market"].dropna().tolist()
        seen = set(self.actuals["Market"]) | set(self.targets["Market"])
        return sorted({str(x).strip() for x in reg} | seen)

    def display_channels(self, split_meta: bool) -> list:
        """What the user picks from.

        Combined, a channel with children is shown as the parent and its
        children are hidden — selecting Meta then means Meta API + Meta Ecom.
        Split, the children appear and the parent does not, because selecting a
        parent alongside its children would double-count.
        """
        par = self.parent_of()
        kids = set(par)
        parents = set(par.values())
        seen = set(self.actuals["Channel"])
        reg = (self.channels["Channel"].dropna().tolist()
               if "Active" not in self.channels.columns else
               self.channels[self.channels["Active"].astype(str).str.lower() == "yes"]
               ["Channel"].tolist())
        all_ch = {str(x).strip() for x in reg} | seen
        if split_meta:
            return sorted(all_ch - parents)
        return sorted((all_ch - kids) | parents)

    def expand(self, channels, split_meta: bool) -> list:
        """Turn a display selection into the channels the actuals are stored
        against. Combined, a parent expands to its children."""
        if not channels:
            return None
        if split_meta:
            return list(channels)
        out = []
        for c in channels:
            out += self.children_of(c)
        return sorted(set(out))

    def channel_list(self) -> list:
        """Channels a user picks from: the ones actuals are reported against.
        A planning-only parent like Meta is excluded, because selecting it would
        return nothing — its data lives under its children."""
        d = self.channels
        if "Active" in d.columns:
            reg = d[d["Active"].astype(str).str.lower() == "yes"]["Channel"].tolist()
        else:
            reg = d["Channel"].dropna().tolist()
        parents = set(self.parent_of().values())
        seen = set(self.actuals["Channel"])
        return sorted(({str(x).strip() for x in reg} | seen) - (parents - seen))

    def parent_of(self) -> dict:
        """{tracking channel: planning channel}. Meta is planned once and
        reported as Meta API and Meta Ecom, so a plan written against Meta has
        to be compared with the sum of its children."""
        d = self.channels
        if "Rolls up to" not in d.columns:
            return {}
        out = {}
        for _, r in d.iterrows():
            p = r.get("Rolls up to")
            if pd.notna(p) and str(p).strip():
                out[str(r["Channel"]).strip()] = str(p).strip()
        return out

    def children_of(self, channel) -> list:
        """Tracking channels that roll up to this one, or itself if none do."""
        kids = [c for c, p in self.parent_of().items() if p == channel]
        return kids or [channel]

    def plan_channel(self, channel) -> str:
        """The channel a plan would be written against for this tracking channel."""
        return self.parent_of().get(channel, channel)

    def months_in(self, year) -> list:
        """Months with a plan or actuals in this year, in calendar order."""
        p = {mo for y, mo in zip(self.targets["Year"], self.targets["Month"])
             if int(y) == int(year)}
        a = {MONTHS[d.month - 1] for d in self.actuals["Date"] if d.year == int(year)}
        return sorted(p | a, key=lambda mo: MONTH_NO.get(mo, 0))

    def years(self) -> list:
        p = {int(y) for y in self.targets["Year"].dropna()}
        a = {int(d.year) for d in self.actuals["Date"]}
        return sorted(p | a)

    def periods(self) -> list:
        """(year, month) pairs present in the plan or the actuals, newest last."""
        p = {(int(y), m) for y, m in zip(self.targets["Year"], self.targets["Month"])}
        a = {(d.year, MONTHS[d.month - 1]) for d in self.actuals["Date"]}
        return sorted(p | a, key=lambda t: (t[0], MONTH_NO.get(t[1], 0)))


def _read(xl, sheet, header_row=3) -> pd.DataFrame:
    d = pd.read_excel(xl, sheet_name=sheet, header=header_row)
    d = d.dropna(how="all")
    d.columns = [str(c).strip() for c in d.columns]
    return d


def _find_block(ws, title, header) -> tuple[int, int]:
    """Locate a table on the shared Setup sheet by its title and header row.

    Returns (header_row, first_column). Searching by text rather than by a fixed
    address means a block can move without silently reading the wrong cells.
    """
    for r in range(1, 80):
        for c in range(1, 20):
            if str(ws.cell(r, c).value or "").strip() == title:
                for rr in range(r, r + 4):
                    if str(ws.cell(rr, c).value or "").strip() == header:
                        return rr, c
    raise KeyError(f"{title!r} not found on {ws.title}")


def _read_block(ws, hdr_row, col0) -> pd.DataFrame:
    cols = []
    c = col0
    while ws.cell(hdr_row, c).value:
        cols.append(str(ws.cell(hdr_row, c).value).strip())
        c += 1
    rows = []
    r = hdr_row + 1
    while ws.cell(r, col0).value:
        rows.append([ws.cell(r, col0 + k).value for k in range(len(cols))])
        r += 1
    return pd.DataFrame(rows, columns=cols)


def load_model(path) -> Model:
    xl = pd.ExcelFile(path)

    def sheet(name):
        d = pd.read_excel(xl, sheet_name=name, header=3)
        d = d.dropna(how="all")
        d.columns = [str(c).strip() for c in d.columns]
        return d

    tg = sheet(SH_TARGETS)
    tg = tg[tg["Market"].notna()].copy()
    tg["Year"] = pd.to_numeric(tg["Year"], errors="coerce").astype("Int64")

    cap = sheet(SH_CAPACITY)
    cap = cap[cap["Market"].notna()].copy()
    cap["Year"] = pd.to_numeric(cap["Year"], errors="coerce").astype("Int64")

    ch = sheet(SH_CHANNEL)
    ch = ch[ch["Market"].notna()].copy()
    ch["Year"] = pd.to_numeric(ch["Year"], errors="coerce").astype("Int64")

    sm = sheet(SH_SUMMARY)
    sm = sm[sm["Market"].notna() & (sm["Market"] != "")].copy()
    sm["Year"] = pd.to_numeric(sm["Year"], errors="coerce").astype("Int64")

    ac = sheet(SH_ACTUALS)
    ac = ac[ac["Market"].notna()].copy()
    ac["Date"] = pd.to_datetime(ac["Date"], errors="coerce")
    ac["Value"] = pd.to_numeric(ac["Value"], errors="coerce")
    ac = ac.dropna(subset=["Date", "Value"])
    ac["Year"] = ac["Date"].dt.year
    ac["Month"] = ac["Date"].dt.month.map(lambda m: MONTHS[m - 1])
    ac["Day"] = ac["Date"].dt.date

    # Setup: four tables on one sheet, each found by its header text
    import openpyxl
    wbv = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wbv[SH_SETUP]
    mk = _read_block(ws, *_find_block(ws, "Markets", "Market"))
    cn = _read_block(ws, *_find_block(ws, "Channels", "Channel"))
    try:
        fxo = _read_block(ws, *_find_block(ws, "FX overrides", "Market"))
    except KeyError:
        fxo = pd.DataFrame(columns=["Market", "Month", "Rate to AED"])

    fx, note = {}, "ok"
    for _, r in mk.iterrows():
        rate = pd.to_numeric(r.get("Rate to AED"), errors="coerce")
        if pd.notna(rate) and rate > 0:
            fx[(str(r["Market"]).strip(), None)] = float(rate)
    for _, r in fxo.iterrows():
        rate = pd.to_numeric(r.get("Rate to AED"), errors="coerce")
        if pd.isna(rate) or rate <= 0 or pd.isna(r.get("Market")):
            continue
        key = None
        if pd.notna(r.get("Month")) and str(r.get("Month")).strip():
            try:
                key = pd.to_datetime(r["Month"]).strftime("%Y-%m")
            except Exception:
                key = str(r["Month"]).strip()
        fx[(str(r["Market"]).strip(), key)] = float(rate)
    if not fx:
        note = "no FX rates — revenue read as if already in AED"

    # Spend is entered in AED except where the register says otherwise, so the
    # conversion is driven by the register rather than by naming a market here.
    spend_local = {str(r["Market"]).strip() for _, r in mk.iterrows()
                   if str(r.get("Spend ccy", "AED")).strip().upper() != "AED"}

    if fx:
        rates = np.array([fx_rate(fx, m, MONTH_NO.get(mo, 1), y)
                          for m, mo, y in zip(ac["Market"], ac["Month"], ac["Year"])])
        conv = ac["Metric"].eq(M_REVENUE) | (
            ac["Metric"].eq(M_SPEND) & ac["Market"].isin(spend_local))
        ac.loc[conv, "Value"] = ac.loc[conv, "Value"].to_numpy() * rates[conv.to_numpy()]
        col = "Target revenue (local ccy)"
        if col in tg.columns:
            trates = np.array([fx_rate(fx, m, MONTH_NO.get(mo, 1), int(y))
                               for m, mo, y in zip(tg["Market"], tg["Month"], tg["Year"])])
            tg[col] = pd.to_numeric(tg[col], errors="coerce") * trates

    return Model(tg, cap, ch, sm, ac, sm, mk, cn, fx, note)


def fx_rate(fx: dict, market, month: int, year: int) -> float:
    """A dated row wins over the standing row. A pegged currency never needs one."""
    if not fx:
        return 1.0
    key = f"{int(year):04d}-{int(month):02d}"
    if (market, key) in fx:
        return fx[(market, key)]
    return fx.get((market, None), 1.0)


# ─────────────────────────────────────────────────────────────────────
# SCOPING
# ─────────────────────────────────────────────────────────────────────
def scope(df, markets=None, channels=None, year=None, month=None,
          date_from=None, date_to=None) -> pd.DataFrame:
    d = df
    if markets:
        d = d[d["Market"].isin(markets)]
    if channels and "Channel" in d.columns:
        d = d[d["Channel"].isin(channels)]
    if year is not None and "Year" in d.columns:
        years = [year] if isinstance(year, (int, np.integer)) else list(year)
        d = d[d["Year"].isin(years)]
    if month is not None and "Month" in d.columns:
        months = [month] if isinstance(month, str) else list(month)
        d = d[d["Month"].isin(months)]
    if date_from is not None and "Date" in d.columns:
        d = d[d["Date"] >= pd.Timestamp(date_from)]
    if date_to is not None and "Date" in d.columns:
        d = d[d["Date"] <= pd.Timestamp(date_to)]
    return d


def actual(m: Model, metric, **kw) -> Optional[float]:
    d = scope(m.actuals, **kw)
    d = d[d["Metric"] == metric] if isinstance(metric, str) else d[d["Metric"].isin(metric)]
    return None if d.empty else float(d["Value"].sum())


def _plan_channels(m: Model, channels):
    """Map a tracking selection to the channels the plan is written against, and
    drop duplicates: selecting Meta API and Meta Ecom must not count the single
    Meta plan twice."""
    if not channels:
        return None
    return sorted({m.plan_channel(c) for c in channels})


def _plan_field(m: Model, field, markets, channels, year, month):
    d = scope(m.channel, markets, _plan_channels(m, channels), year, month)
    if d.empty or field not in d:
        return None
    return float(pd.to_numeric(d[field], errors="coerce").sum())


def plan_orders(m: Model, markets=None, channels=None, year=None, month=None):
    return _plan_field(m, "Plan orders", markets, channels, year, month)


def plan_budget(m: Model, markets=None, channels=None, year=None, month=None):
    return _plan_field(m, "Plan budget (AED)", markets, channels, year, month)


def target_orders(m: Model, markets=None, year=None, month=None):
    d = scope(m.targets, markets, None, year, month)
    c = "Target orders"
    return None if d.empty or c not in d else float(pd.to_numeric(d[c], errors="coerce").sum())


def target_revenue(m: Model, markets=None, year=None, month=None):
    d = scope(m.targets, markets, None, year, month)
    c = "Target revenue (local ccy)"
    return None if d.empty or c not in d else float(pd.to_numeric(d[c], errors="coerce").sum())


def target_units(m: Model, markets=None, year=None, month=None):
    d = scope(m.targets, markets, None, year, month)
    c = "Target units"
    return None if d.empty or c not in d else float(pd.to_numeric(d[c], errors="coerce").sum())


# ─────────────────────────────────────────────────────────────────────
# COVERAGE AND PACING
# ─────────────────────────────────────────────────────────────────────
@dataclass
class Coverage:
    year: int
    month: str
    days_in_month: int
    days_elapsed: int
    days_remaining: int
    per_market: dict = field(default_factory=dict)

    def gate(self, market) -> tuple[bool, str]:
        """Thin data greys out rather than turning red — a missing feed is not
        a performance signal."""
        if not market:
            return False, ""
        rep = self.per_market.get(market, (0, 0))[0]
        if not self.days_elapsed:
            return True, "no data"
        if rep / self.days_elapsed < COVERAGE_MIN:
            return True, f"only {rep}/{self.days_elapsed} days reported"
        return False, ""


def coverage(m: Model, year: int, month) -> Coverage:
    """Days elapsed and days in the period, across however many months are
    selected. A completed month contributes all its days, so pacing a mix of
    finished and running months still lands on the right number."""
    months = [month] if isinstance(month, str) else list(month)
    d = scope(m.actuals, year=year, month=months)
    dim = sum(calendar.monthrange(year, MONTH_NO[mo])[1] for mo in months)

    elapsed = 0
    for mo in months:
        dm = d[d["Month"] == mo]
        days_in = calendar.monthrange(year, MONTH_NO[mo])[1]
        seen = int(dm["Day"].nunique()) if not dm.empty else 0
        # A month with no actuals has not started; one with data counts the
        # days it reported.
        elapsed += min(seen, days_in)

    per = {}
    for mk in sorted(d["Market"].unique()):
        dmk = d[d["Market"] == mk]
        rep = int(dmk["Day"].nunique())
        od = dmk[dmk["Metric"] == M_ORDERS].groupby("Day")["Value"].sum()
        per[mk] = (rep, int((od > 0).sum()))
    label = months[0] if len(months) == 1 else f"{months[0]}–{months[-1]}"
    return Coverage(year, label, dim, elapsed, max(dim - elapsed, 0), per)


def paced(plan, cov: Coverage):
    """Month plan pro-rated to the days elapsed in the period.

    Across several months this is the ratio of days elapsed to days in the whole
    period, so a finished month contributes its full plan and a running one
    contributes its share.
    """
    if plan is None or not cov.days_in_month:
        return None
    return plan * cov.days_elapsed / cov.days_in_month


def eom(act, cov: Coverage):
    """Straight run rate, not a forecast: it assumes no change in trajectory."""
    if act is None or not cov.days_elapsed:
        return None
    return act / cov.days_elapsed * cov.days_in_month


# ─────────────────────────────────────────────────────────────────────
# MOMENTUM
# ─────────────────────────────────────────────────────────────────────
@dataclass
class Momentum:
    recent: Optional[float]
    prior: Optional[float]
    window: int
    label: str
    arrow: str


def momentum(series: pd.Series, window=7) -> Momentum:
    """Last n days against the n before. Deliberately not first half vs second
    half: on a 28-day month that compares a fortnight-old average against a
    three-week-old one and calls it current."""
    v = list(series.values)
    if len(v) < 4:
        return Momentum(None, None, window, "not enough data", "flat")
    w = min(window, len(v) // 2)
    recent, prior = float(np.mean(v[-w:])), float(np.mean(v[-2 * w:-w]))
    if not prior:
        return Momentum(recent, prior, w, "not enough data", "flat")
    if recent > prior * 1.05:
        return Momentum(recent, prior, w, "accelerating", "up")
    if recent < prior * 0.95:
        return Momentum(recent, prior, w, "slowing", "down")
    return Momentum(recent, prior, w, "steady", "flat")


def daily_series(m: Model, metric, markets=None, channels=None,
                 year=None, month=None) -> pd.Series:
    d = scope(m.actuals, markets, channels, year, month)
    d = d[d["Metric"] == metric] if isinstance(metric, str) else d[d["Metric"].isin(metric)]
    if d.empty:
        return pd.Series(dtype=float)
    return d.groupby("Date")["Value"].sum().sort_index()


# ─────────────────────────────────────────────────────────────────────
# OVERVIEW
# ─────────────────────────────────────────────────────────────────────
@dataclass
class Card:
    key: str
    label: str
    actual: Optional[float]
    paced: Optional[float]
    plan: Optional[float]
    verdict: Verdict
    prefix: str = ""
    suffix: str = ""
    dec: int = 0
    trend: str = "flat"


def overview_cards(m: Model, markets, channels, year, month, cov: Coverage) -> list:
    """Four numbers. Each carries paced AND month plan, because 'am I on track
    today' and 'what am I chasing' are different questions and one dashboard
    serves both audiences."""
    kw = dict(markets=markets, channels=channels, year=year, month=month)
    o = actual(m, M_ORDERS, **kw)
    rev = actual(m, M_REVENUE, **kw)
    sp = actual(m, M_SPEND, **kw)
    t_ord = target_orders(m, markets, year, month)
    t_rev = target_revenue(m, markets, year, month)
    ceiling = plan_budget(m, markets, channels, year, month)
    cpa = div(sp, o)
    plan_cpa = div(ceiling, plan_orders(m, markets, channels, year, month))

    def trend_of(metric):
        return momentum(daily_series(m, metric, markets, channels, year, month)).arrow

    return [
        Card("orders", "Orders", o, paced(t_ord, cov), t_ord,
             rag(o, paced(t_ord, cov), "up", "paced plan", key="orders"),
             trend=trend_of(M_ORDERS)),
        Card("revenue", "Revenue", rev, paced(t_rev, cov), t_rev,
             rag(rev, paced(t_rev, cov), "up", "paced plan", key="revenue"),
             prefix="AED ", trend=trend_of(M_REVENUE)),
        Card("spend", "Spend", sp, paced(ceiling, cov), ceiling,
             rag(sp, paced(ceiling, cov), "neutral", "paced ceiling", key="spend"),
             prefix="AED ", trend=trend_of(M_SPEND)),
        Card("cpa", "CPA", cpa, None, plan_cpa,
             rag(cpa, plan_cpa, "down", "plan", key="cpa"),
             prefix="AED ", dec=2),
    ]


def gap_table(m: Model, markets, channels, year, month, cov: Coverage,
              split_meta: bool = False) -> pd.DataFrame:
    """Market x channel ranked by share of the shortfall. The plan tells you the
    month is behind; this says where, which is the first question anyone asks."""
    rows = []
    sel = channels or m.display_channels(split_meta)
    # Combined, group by the channel the plan is written against, so Meta's
    # single plan is compared with Meta API + Meta Ecom rather than once per
    # platform. Split, each platform stands on its own and shares the parent's
    # plan proportionally to what it delivered — a plan written for Meta as a
    # whole cannot be attributed to one platform without inventing a split.
    groups = {}
    if split_meta:
        for c in sel:
            groups.setdefault(c, []).append(c)
    else:
        for c in sel:
            groups.setdefault(m.plan_channel(c), []).extend(m.children_of(c))
    for mk in (markets or m.market_list()):
        for ch, kids in groups.items():
            act = actual(m, M_ORDERS, markets=[mk], channels=kids,
                         year=year, month=month) or 0.0
            sp = actual(m, M_SPEND, markets=[mk], channels=kids,
                        year=year, month=month) or 0.0
            po = plan_orders(m, [mk], kids, year, month)
            if split_meta and m.parent_of().get(ch):
                po = None          # the plan belongs to the parent, not this platform
            if po is None and not act and not sp:
                continue
            pc = paced(po, cov)
            gap = None if pc is None else pc - act
            rows.append({
                "Market": mk, "Channel": ch,
                "Actual": act, "Paced plan": pc,
                "Behind by": gap, "vs paced": pct(act, pc),
                "Spend": sp, "_gap": gap if (gap and gap > 0) else 0.0,
                "_noplan": po is None,
            })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    tot = df["_gap"].sum()
    df["Share of gap"] = df["_gap"].apply(lambda g: (g / tot * 100) if tot else 0.0)
    return df.sort_values("_gap", ascending=False).reset_index(drop=True)


def allocation_table(m: Model, markets, channels, year, month, cov: Coverage,
                     split_meta: bool = False) -> pd.DataFrame:
    """Ranked by what an order actually costs, cheapest first, so the table reads
    in the order money should flow. Budget used separates a channel that stopped
    working from one that simply has not spent — identical on pace alone."""
    rows = []
    sel = channels or m.display_channels(split_meta)
    groups = {}
    if split_meta:
        for c in sel:
            groups.setdefault(c, []).append(c)
    else:
        for c in sel:
            groups.setdefault(m.plan_channel(c), []).extend(m.children_of(c))
    for mk in (markets or m.market_list()):
        for ch, kids in groups.items():
            o = actual(m, M_ORDERS, markets=[mk], channels=kids,
                       year=year, month=month) or 0.0
            sp = actual(m, M_SPEND, markets=[mk], channels=kids,
                        year=year, month=month) or 0.0
            rev = actual(m, M_REVENUE, markets=[mk], channels=kids,
                         year=year, month=month) or 0.0
            po = plan_orders(m, [mk], kids, year, month)
            pb = plan_budget(m, [mk], kids, year, month)
            if split_meta and m.parent_of().get(ch):
                po = pb = None     # the plan belongs to the parent, not this platform
            if po is None and not o and not sp:
                continue
            cpa, plan_cpa = div(sp, o), div(pb, po)
            rows.append({
                "Market": mk, "Channel": ch, "Orders": o, "Spend": sp,
                "CPA": cpa, "Plan CPA": plan_cpa,
                "Cost vs plan": div(cpa, plan_cpa),
                "ROAS": div(rev, sp),
                "Budget used": pct(sp, paced(pb, cov)),
                "Unspent": (paced(pb, cov) - sp) if pb is not None else None,
                "vs paced": pct(o, paced(po, cov)),
                "Read": _read_row(pct(o, paced(po, cov)), pct(sp, paced(pb, cov)),
                                  div(cpa, plan_cpa),
                                  (paced(pb, cov) - sp) if pb is not None else None),
            })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    # A cell with orders but no recorded spend shows a CPA of zero; it must not
    # take the top of a cost ranking.
    df["_sort"] = [c if (c and c > 0) else float("inf") for c in df["CPA"]]
    df = df.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)
    priced = df[df["CPA"].fillna(0) > 0]
    if len(priced) > 1:
        df.loc[priced.index[0], "Read"] = "Cheapest orders here. " + df.loc[priced.index[0], "Read"]
    return df


def _read_row(vs_pace, used, cost_idx, headroom) -> str:
    """One sentence per row, written from that row's own figures. A verdict word
    needs a legend and a rule the reader has to hold in their head; a sentence
    carries its own justification."""
    if cost_idx is None or vs_pace is None:
        return "Reporting actuals with no plan to compare against."
    if cost_idx < 1:
        cost = f"costing {(1-cost_idx)*100:.0f}% less than planned"
    elif cost_idx < 1.1:
        cost = "costing about what was planned"
    elif cost_idx < 2:
        cost = f"costing {(cost_idx-1)*100:.0f}% more than planned"
    else:
        cost = f"costing {cost_idx:.1f}x what was planned"

    if used is not None and used < 70 and vs_pace < 90 and cost_idx < 1.3 \
            and headroom and headroom > 0:
        return (f"Behind because AED {headroom:,.0f} never went out, not because it "
                f"stopped working — {cost}.")
    if used is not None and used < 70 and cost_idx <= 1:
        return f"{cost.capitalize()}, and only {used:.0f}% of its budget has gone out."
    if cost_idx >= 1.5 and vs_pace < 90:
        return f"Spent {used:.0f}% of budget {cost} and returned {vs_pace:.0f}% of paced orders."
    if used is not None and used > 150:
        return f"{vs_pace:.0f}% of paced orders, but {used:.0f}% of budget spent, {cost}."
    if vs_pace < 90:
        return f"{vs_pace:.0f}% of paced orders on {used:.0f}% of budget, {cost}."
    return f"On pace at {vs_pace:.0f}%, {cost}."


def headline(m: Model, markets, channels, year, month, cov: Coverage,
             split_meta: bool = False) -> list:
    """Returns [(severity, sentence)]. Every figure is read from the data, never
    retyped, so the sentence cannot drift from the table beneath it."""
    kw = dict(markets=markets, channels=channels, year=year, month=month)
    o = actual(m, M_ORDERS, **kw)
    t = target_orders(m, markets, year, month)
    out = []
    if o is None or t is None:
        return [("info", f"No plan or no actuals for {month} {year}.")]

    pc = paced(t, cov)
    r = pct(o, pc)
    sev = "good" if r >= 90 else "warn" if r >= 70 else "risk"
    state = "on track" if sev == "good" else "behind pace" if sev == "warn" else "significantly behind"
    e = eom(o, cov)
    out.append((sev, f"{'All markets' if not markets or len(markets) > 1 else markets[0]} "
                     f"{'is' if markets and len(markets) == 1 else 'are'} {state} in {month}. "
                     f"Day {cov.days_elapsed} of {cov.days_in_month}: {o:,.0f} orders, "
                     f"{r:.0f}% of paced plan. Run rate lands at {e:,.0f}, "
                     f"{pct(e, t):.0f}% of the month."))

    if cov.days_remaining > 0:
        need = (t - o) / cov.days_remaining
        rate = t / cov.days_in_month
        if need > 0:
            out.append(("risk" if need > rate * 1.5 else "warn",
                        f"Closing the gap needs {need:,.0f} orders/day for the "
                        f"{cov.days_remaining} days left, {(need/rate-1)*100:.0f}% above the "
                        f"{rate:,.0f}/day plan rate."))

    mo = momentum(daily_series(m, M_ORDERS, markets, channels, year, month))
    if mo.label in ("accelerating", "slowing"):
        out.append(("good" if mo.label == "accelerating" else "warn",
                    f"Momentum {mo.label}: last {mo.window} days average {mo.recent:,.0f}/day "
                    f"against {mo.prior:,.0f}/day in the {mo.window} before."))

    g = gap_table(m, markets, channels, year, month, cov, split_meta)
    behind = g[g["Share of gap"] > 0] if not g.empty else g
    if len(behind):
        top = behind.iloc[0]
        if top["Share of gap"] > 20:
            out.append(("risk", f"{top['Market']} {top['Channel']} is "
                                f"{top['Share of gap']:.0f}% of the shortfall: "
                                f"{top['Actual']:,.0f} orders against "
                                f"{top['Paced plan']:,.0f} paced."))

    sp = actual(m, M_SPEND, **kw)
    ceiling = plan_budget(m, markets, channels, year, month)
    if sp and ceiling and sp > ceiling:
        out.append(("risk", f"Spend has passed the budget ceiling by "
                            f"AED {sp-ceiling:,.0f}."))
    return out


# ─────────────────────────────────────────────────────────────────────
# COMPARISON
# ─────────────────────────────────────────────────────────────────────
CMP_KEYS = ["orders", "units", "revenue", "spend", "cpa", "roas", "cr", "daily"]
CMP_LABEL = {"orders": "Orders", "units": "Units", "revenue": "Revenue",
             "spend": "Spend", "cpa": "CPA", "roas": "ROAS",
             "cr": "CR%", "daily": "Orders/day"}
CMP_FMT = {"orders": ("", "", 0), "units": ("", "", 0), "revenue": ("AED ", "", 0),
           "spend": ("AED ", "", 0), "cpa": ("AED ", "", 2), "roas": ("", "x", 1),
           "cr": ("", "%", 2), "daily": ("", "", 1)}


def cmp_block(m: Model, start, end, markets=None, channels=None) -> dict:
    kw = dict(markets=markets, channels=channels, date_from=start, date_to=end)
    o = actual(m, M_ORDERS, **kw) or 0.0
    u = actual(m, M_UNITS, **kw) or 0.0
    rev = actual(m, M_REVENUE, **kw) or 0.0
    sp = actual(m, M_SPEND, **kw) or 0.0
    msg = nsum(actual(m, M_MSG_SENT, **kw), actual(m, M_MSG_RECV, **kw)) or 0.0
    days = max((pd.Timestamp(end) - pd.Timestamp(start)).days + 1, 1)
    d = scope(m.actuals, **kw)
    return {"orders": o, "units": u, "revenue": rev, "spend": sp, "messages": msg,
            "cpa": div(sp, o), "roas": div(rev, sp), "cr": pct(o, msg),
            "daily": div(o, days), "days": days,
            "reported": int(d["Day"].nunique()) if not d.empty else 0}


def cmp_change(a, b, key) -> dict:
    """Delta plus how to read it. Spend never reads better or worse — it is an
    input, and CPA and ROAS already judge what it bought."""
    va, vb = a.get(key), b.get(key)
    pol = POLARITY.get(key, "up")
    if va is None and vb is None:
        return {"delta": None, "pct": None, "read": "n/a"}
    va = va or 0.0
    if not vb:
        return {"delta": va, "pct": None, "read": "new" if va else "n/a"}
    d = va - vb
    p = d / vb * 100
    if pol == "neutral":
        read = "higher" if p > 1 else "lower" if p < -1 else "flat"
    elif abs(p) < 1:
        read = "flat"
    else:
        read = "better" if ((p > 0) == (pol != "down")) else "worse"
    return {"delta": d, "pct": p, "read": read}


def cmp_headline(m: Model, ar, br, markets=None, channels=None) -> pd.DataFrame:
    A, B = cmp_block(m, *ar, markets, channels), cmp_block(m, *br, markets, channels)
    rows = []
    for k in CMP_KEYS:
        p, s, dc = CMP_FMT[k]
        c = cmp_change(A, B, k)
        rows.append({"Metric": CMP_LABEL[k],
                     "Period A": fmt(A[k], p, s, dc),
                     "Period B": fmt(B[k], p, s, dc),
                     "Change": ("n/a" if c["delta"] is None else
                                fmt(c["delta"], p, s, dc) if c["pct"] is not None else "new"),
                     "Δ%": "n/a" if c["pct"] is None else f"{c['pct']:+.1f}%",
                     "Direction": c["read"]})
    return pd.DataFrame(rows)


def cmp_hierarchy(m: Model, ar, br, markets=None, channels=None,
                  split_meta: bool = False) -> pd.DataFrame:
    """Group, then market, then channel. Markets ordered by how much they moved,
    so whatever drove the change is at the top rather than wherever the alphabet
    puts it."""
    mkts = markets or m.market_list()
    chans = channels or m.display_channels(split_meta)

    def row(label, level, mks, chs):
        A, B = cmp_block(m, *ar, mks, chs), cmp_block(m, *br, mks, chs)
        c = cmp_change(A, B, "orders")
        return {"_level": level, "Scope": label,
                "A orders": A["orders"], "B orders": B["orders"],
                "Δ orders": c["delta"], "Δ%": c["pct"],
                "A revenue": A["revenue"], "B revenue": B["revenue"],
                "A spend": A["spend"], "B spend": B["spend"],
                "A CPA": A["cpa"], "B CPA": B["cpa"],
                "A ROAS": A["roas"], "B ROAS": B["roas"],
                "_abs": abs(c["delta"] or 0)}

    flat = m.expand(chans, split_meta)
    out = [row("All markets", 0, mkts, flat)]
    mrows = sorted([row(mk, 1, [mk], flat) for mk in mkts], key=lambda r: -r["_abs"])
    for mr in mrows:
        out.append(mr)
        for ch in chans:
            cr = row(ch, 2, [mr["Scope"]], m.expand([ch], split_meta))
            if cr["A orders"] or cr["B orders"] or cr["A spend"] or cr["B spend"]:
                out.append(cr)
    df = pd.DataFrame(out)
    tot = df[df["_level"] == 1]["_abs"].sum()
    df["Share of change"] = df.apply(
        lambda r: (r["_abs"] / tot * 100) if tot and r["_level"] == 1 else None, axis=1)
    return df


def cmp_daily(m: Model, ar, br, markets=None, channels=None) -> pd.DataFrame:
    """Aligned by position in each range, not by calendar date, so day 1 of A
    sits against day 1 of B and two windows of different length stay readable."""
    def ser(rng):
        d = scope(m.actuals, markets, channels, date_from=rng[0], date_to=rng[1])
        d = d[d["Metric"] == M_ORDERS]
        s = d.groupby("Date")["Value"].sum() if not d.empty else pd.Series(dtype=float)
        idx = pd.date_range(rng[0], rng[1], freq="D")
        return s.reindex(idx, fill_value=0.0)

    sa, sb = ser(ar), ser(br)
    n = max(len(sa), len(sb))
    return pd.DataFrame({
        "Day": [f"d{i+1}" for i in range(n)],
        "Period A": [float(sa.iloc[i]) if i < len(sa) else None for i in range(n)],
        "Period B": [float(sb.iloc[i]) if i < len(sb) else None for i in range(n)],
        "A date": [sa.index[i].strftime("%d %b") if i < len(sa) else "" for i in range(n)],
        "B date": [sb.index[i].strftime("%d %b") if i < len(sb) else "" for i in range(n)],
    })


def cmp_summary(m: Model, ar, br, markets=None, channels=None) -> list:
    A, B = cmp_block(m, *ar, markets, channels), cmp_block(m, *br, markets, channels)
    out = []
    o, s = cmp_change(A, B, "orders"), cmp_change(A, B, "spend")
    if o["pct"] is None:
        return [("info", "Period B has no orders, so there is nothing to compare against.")]

    sev = "good" if o["pct"] > 1 else "risk" if o["pct"] < -1 else "info"
    sp_txt = (f" on {abs(s['pct']):.0f}% {'more' if s['pct'] > 0 else 'less'} spend"
              if s["pct"] is not None and abs(s["pct"]) >= 1 else " on flat spend")
    if abs(o["pct"]) < 1:
        out.append((sev, f"Orders held flat at {A['orders']:,.0f}{sp_txt}."))
    else:
        out.append((sev, f"Orders {'rose' if o['pct'] > 0 else 'fell'} "
                         f"{abs(o['pct']):.0f}% to {A['orders']:,.0f}{sp_txt}."))

    cpa, roas = cmp_change(A, B, "cpa"), cmp_change(A, B, "roas")
    if cpa["pct"] is not None and roas["pct"] is not None:
        if abs(cpa["pct"]) < 1 and abs(roas["pct"]) < 1:
            out.append(("info", f"Efficiency unchanged: CPA {fmt(A['cpa'],'AED ',dec=2)}, "
                                f"ROAS {A['roas']:.1f}x."))
        else:
            out.append(("good" if cpa["read"] == "better" else "warn",
                        f"CPA {'fell' if cpa['pct'] < 0 else 'rose'} from "
                        f"{fmt(B['cpa'],'AED ',dec=2)} to {fmt(A['cpa'],'AED ',dec=2)} "
                        f"and ROAS moved {B['roas']:.1f}x to {A['roas']:.1f}x."))

    mkts = markets or m.market_list()
    if len(mkts) > 1:
        moves = [(mk, (cmp_block(m, *ar, [mk], channels)["orders"] or 0)
                  - (cmp_block(m, *br, [mk], channels)["orders"] or 0)) for mk in mkts]
        tot = sum(abs(x[1]) for x in moves)
        moves.sort(key=lambda x: -abs(x[1]))
        if tot and abs(moves[0][1]) > 0:
            mk, d = moves[0]
            out.append(("info", f"{mk} accounts for {abs(d)/tot*100:.0f}% of the movement, "
                                f"{'up' if d > 0 else 'down'} {abs(d):,.0f} orders."))

    chans = channels or m.channel_list()
    for ch in chans:
        a_, b_ = cmp_block(m, *ar, mkts, [ch]), cmp_block(m, *br, mkts, [ch])
        c = cmp_change(a_, b_, "orders")
        if c["pct"] is not None and (c["pct"] < 0) != (o["pct"] < 0) and abs(c["pct"]) > 10:
            out.append(("warn", f"{ch} moved the other way: {c['pct']:+.0f}% to "
                                f"{a_['orders']:,.0f} orders, CPA "
                                f"{fmt(b_['cpa'],'AED ',dec=2)} to {fmt(a_['cpa'],'AED ',dec=2)}."))

    for lbl, blk in (("A", A), ("B", B)):
        if blk["reported"] < blk["days"]:
            out.append(("warn", f"Period {lbl} has {blk['reported']} days of data across a "
                                f"{blk['days']}-day window — not like for like."))
    return out


def cmp_presets(dates: list) -> dict:
    if not dates:
        return {}
    ds = sorted(dates)
    last = ds[-1]
    out = {}
    if len(ds) >= 8:
        a_s = last - dt.timedelta(days=6)
        b_e = a_s - dt.timedelta(days=1)
        out["Last 7 days vs prior 7"] = (a_s, last, b_e - dt.timedelta(days=6), b_e)
    if len(ds) >= 28:
        a_s = last - dt.timedelta(days=13)
        b_e = a_s - dt.timedelta(days=1)
        out["Last 14 days vs prior 14"] = (a_s, last, b_e - dt.timedelta(days=13), b_e)
    m_start = last.replace(day=1)
    prev_end = m_start - dt.timedelta(days=1)
    if prev_end >= ds[0]:
        out["This month vs last month"] = (m_start, last, prev_end.replace(day=1), prev_end)
        try:
            same = prev_end.replace(day=min(last.day, prev_end.day))
            out["Same period last month"] = (m_start, last, prev_end.replace(day=1), same)
        except ValueError:
            pass
    return out


# ─────────────────────────────────────────────────────────────────────
# YEAR TO DATE
# ─────────────────────────────────────────────────────────────────────
def closed_months(m: Model, year: int, upto: str = None) -> list:
    """Months that have finished reporting, in calendar order.

    A month counts as closed when it reported on its last day. Year-to-date is
    built from these alone: blending a finished month with a running one gives a
    percentage that measures neither.
    """
    out = []
    for mo in m.months_in(year):
        if upto and MONTH_NO[mo] > MONTH_NO[upto]:
            continue
        d = scope(m.actuals, year=year, month=mo)
        if d.empty:
            continue
        dim = calendar.monthrange(year, MONTH_NO[mo])[1]
        if int(d["Day"].nunique()) >= dim:
            out.append(mo)
    return out


def current_month(m: Model, year: int) -> Optional[str]:
    """The latest month carrying actuals, closed or not."""
    d = scope(m.actuals, year=year)
    if d.empty:
        return None
    return MONTHS[int(d["Date"].dt.month.max()) - 1]


# ─────────────────────────────────────────────────────────────────────
# SPEND TRAJECTORY
# ─────────────────────────────────────────────────────────────────────
@dataclass
class SpendPath:
    daily: pd.Series
    spent: float
    ceiling: Optional[float]
    eom: Optional[float]
    recent_rate: Optional[float]
    prior_rate: Optional[float]
    direction: str
    landing_pct: Optional[float]
    note: str


def spend_path(m: Model, markets, channels, year, month, cov: Coverage,
               window=7) -> SpendPath:
    """Where spend lands if the current rate holds.

    A monthly ceiling only warns after the money is gone. Buyers move bids
    daily, so the rate that matters is the recent one, not the month average —
    projecting from the month average would understate a channel that has just
    scaled up.
    """
    s = daily_series(m, M_SPEND, markets, channels, year, month)
    spent = float(s.sum()) if len(s) else 0.0
    ceiling = plan_budget(m, markets, channels, year, month)

    if not len(s) or not cov.days_remaining:
        eom = spent if len(s) else None
        return SpendPath(s, spent, ceiling, eom, None, None, "flat",
                         pct(eom, ceiling), "" if len(s) else "no spend recorded")

    w = min(window, max(len(s) // 2, 1))
    recent = float(s.iloc[-w:].mean())
    prior = float(s.iloc[-2 * w:-w].mean()) if len(s) >= 2 * w else None

    if prior is None or not prior:
        direction = "flat"
    elif recent > prior * 1.10:
        direction = "rising"
    elif recent < prior * 0.90:
        direction = "falling"
    else:
        direction = "steady"

    eom = spent + recent * cov.days_remaining
    note = (f"last {w} days average AED {recent:,.0f}/day"
            + (f" against AED {prior:,.0f} in the {w} before" if prior else ""))
    return SpendPath(s, spent, ceiling, eom, recent, prior, direction,
                     pct(eom, ceiling), note)


# ─────────────────────────────────────────────────────────────────────
# CAPACITY MODEL vs DELIVERED
# ─────────────────────────────────────────────────────────────────────
def capacity_check(m: Model, markets, year, month) -> pd.DataFrame:
    """Did the WhatsApp model hold?

    Every paid budget is sized from the gap this model leaves, so an assumption
    that is quietly wrong makes every downstream figure wrong. Separating the
    two causes matters: a market can miss because its list converted worse than
    assumed, or because the messages never went out. Those need opposite
    responses and look identical in the order count.
    """
    rows = []
    cap = m.capacity
    for mk in (markets or m.market_list()):
        r = cap[(cap["Market"] == mk) & (cap["Month"] == month)]
        if r.empty:
            continue
        r = r.iloc[0]
        modelled = pd.to_numeric(r.get("NET CAPACITY"), errors="coerce")
        msg_plan = pd.to_numeric(r.get("Total messages"), errors="coerce")
        cr_plan = pd.to_numeric(r.get("Blended CR%"), errors="coerce")
        # Capacity is messages x CR% x uptime. Delivery carries no such haircut,
        # so a market that sends every planned message at exactly the assumed
        # conversion lands at 105% of capacity, not 100%. Without uptime the
        # decomposition does not reconstruct the hit.
        uptime = pd.to_numeric(r.get("Uptime"), errors="coerce")
        if pd.isna(uptime) or uptime <= 0:
            uptime = 1.0
        if pd.isna(modelled):
            continue
        api = [c for c in m.channel_list() if c == "API"] or ["API"]
        # A market with no actuals for this month has not reported, which is not
        # the same as delivering nothing. Scoring it gives "0% of capacity".
        if scope(m.actuals, [mk], None, year, month).empty:
            continue
        delivered = actual(m, M_ORDERS, markets=[mk], channels=api,
                           year=year, month=month) or 0.0
        msg_act = nsum(actual(m, M_MSG_CUST, markets=[mk], channels=api,
                              year=year, month=month),
                       actual(m, M_MSG_LEAD, markets=[mk], channels=api,
                              year=year, month=month)) or 0.0
        cr_act = div(delivered, msg_act)
        rows.append({
            "Market": mk,
            "Modelled capacity": float(modelled),
            "Delivered": delivered,
            "Hit": pct(delivered, modelled),
            "Messages planned": float(msg_plan) if pd.notna(msg_plan) else None,
            "Messages sent": msg_act,
            "Messages %": pct(msg_act, msg_plan),
            "CR% assumed": float(cr_plan) * 100 if pd.notna(cr_plan) else None,
            "CR% actual": cr_act * 100 if cr_act is not None else None,
            "Uptime": float(uptime),
            # Capacity before the uptime haircut — the figure delivery is
            # actually comparable with.
            "Reachable": float(modelled) / float(uptime),
            "Hit before uptime": pct(delivered, float(modelled) / float(uptime)),
            "Read": _capacity_read(pct(delivered, float(modelled) / float(uptime)),
                                   pct(msg_act, msg_plan),
                                   float(cr_plan) * 100 if pd.notna(cr_plan) else None,
                                   cr_act * 100 if cr_act is not None else None,
                                   float(uptime)),
        })
    return pd.DataFrame(rows)


def _capacity_read(hit, msg_pct, cr_plan, cr_act, uptime=1.0) -> str:
    """Attribute the miss arithmetically rather than by threshold.

    Capacity = messages x CR%, so a shortfall is either fewer messages than
    planned, a list converting worse than assumed, or both — and the two need
    opposite responses. Guessing from thresholds mislabels a case where one
    factor moved 6% and the other held exactly.
    """
    if hit is None:
        return "No capacity modelled for this month."
    msg_eff = (msg_pct - 100) if msg_pct is not None else None
    cr_eff = ((cr_act - cr_plan) / cr_plan * 100) if (cr_plan and cr_act) else None
    if msg_eff is None and cr_eff is None:
        return f"{hit:.0f}% of the modelled capacity. Not enough detail to attribute."

    def phrase(v, what):
        return (f"{abs(v):.0f}% {'more' if v > 0 else 'fewer'} {what}"
                if what == "messages sent" else
                f"converted {abs(v):.0f}% {'better' if v > 0 else 'worse'} than assumed")

    parts = []
    if msg_eff is not None and abs(msg_eff) >= 2:
        parts.append(phrase(msg_eff, "messages sent"))
    if cr_eff is not None and abs(cr_eff) >= 2:
        parts.append(phrase(cr_eff, "cr"))

    tail = ""
    if uptime and uptime < 0.999:
        # The haircut is the model's own, and it is worth stating once: without
        # it a market on plan for both factors reads as over-delivering.
        tail = (f" The model also discounts capacity {(1-uptime)*100:.0f}% for "
                f"downtime, which delivery does not carry.")
    if not parts:
        return f"Model held on both send volume and conversion.{tail}"
    if len(parts) == 1:
        driver = parts[0]
        held = ("the CR% assumption held" if "messages" in driver
                else "send volume was on plan")
        return f"{hit:.0f}% of reachable capacity: {driver}, and {held}.{tail}"
    return f"{hit:.0f}% of reachable capacity: {parts[0]}, and {parts[1]}.{tail}"


# ─────────────────────────────────────────────────────────────────────
# CPA IN CONTEXT
# ─────────────────────────────────────────────────────────────────────
def cpa_context(m: Model, markets, channels, year, month,
                prev_month=None) -> pd.DataFrame:
    """CPA against the cheapest channel in the same market, and against last
    month.

    Deliberately not primarily against plan CPA: the budget is orders x that
    number, so it is an assumption the user typed. A lenient assumption turns an
    expensive channel green. The cheapest alternative in the same market is an
    opportunity cost and cannot be set; last month is movement and cannot be
    set either.
    """
    rows = []
    for mk in (markets or m.market_list()):
        cells = []
        for ch in (channels or m.display_channels(False)):
            kids = m.children_of(ch)
            o = actual(m, M_ORDERS, markets=[mk], channels=kids,
                       year=year, month=month) or 0.0
            sp = actual(m, M_SPEND, markets=[mk], channels=kids,
                        year=year, month=month) or 0.0
            rev = actual(m, M_REVENUE, markets=[mk], channels=kids,
                         year=year, month=month) or 0.0
            cpa = div(sp, o)
            if not cpa:
                continue
            prev = None
            if prev_month:
                po = actual(m, M_ORDERS, markets=[mk], channels=kids,
                            year=year, month=prev_month)
                ps = actual(m, M_SPEND, markets=[mk], channels=kids,
                            year=year, month=prev_month)
                prev = div(ps, po)
            cells.append({"Market": mk, "Channel": ch, "Orders": o, "Spend": sp,
                          "CPA": cpa, "Last month": prev,
                          "vs last month": delta_pct(cpa, prev),
                          "ROAS": div(rev, sp),
                          "Plan CPA": div(plan_budget(m, [mk], kids, year, month),
                                          plan_orders(m, [mk], kids, year, month))})
        if not cells:
            continue
        best = min(c["CPA"] for c in cells)
        best_ch = next(c["Channel"] for c in cells if c["CPA"] == best)
        for c in cells:
            c["vs cheapest"] = c["CPA"] / best
            c["Cheapest here"] = best_ch
            c["Read"] = _cpa_read(c["Channel"], best_ch, c["CPA"] / best,
                                  c["vs last month"])
        rows += cells
    df = pd.DataFrame(rows)
    return df.sort_values(["Market", "CPA"]).reset_index(drop=True) if len(df) else df


def _cpa_read(ch, best_ch, ratio, vs_last) -> str:
    if ch == best_ch:
        base = "Cheapest order in this market."
    elif ratio < 1.3:
        base = f"Close to {best_ch} on cost."
    else:
        base = (f"{ratio:.1f} {best_ch} orders for the price of one."
                if ratio >= 2 else
                f"{(ratio-1)*100:.0f}% dearer than {best_ch}.")
    if vs_last is not None and abs(vs_last) >= 5:
        base += (f" {'Up' if vs_last > 0 else 'Down'} {abs(vs_last):.0f}% "
                 f"on last month.")
    return base


# ─────────────────────────────────────────────────────────────────────
# EXECUTIVE COMMENTARY
#
# Built for a review every three days. Three stacked alert bars read as a feed:
# a manager has to assemble the picture from colour-coded fragments, and the
# reasoning that connects them — revenue held BECAUSE the orders were larger,
# behind BUT ending stronger — cannot be expressed in separate bars at all.
#
# Every figure is read from the data, so the commentary cannot drift from the
# tables beneath it.
# ─────────────────────────────────────────────────────────────────────
@dataclass
class Window:
    label: str          # what question it answers
    a_from: object
    a_to: object
    b_from: object
    b_to: object
    dates: str          # shown on screen, so "last 3 days" is never ambiguous
    metrics: dict       # key -> (value, pct change, read)
    text: str


@dataclass
class Freshness:
    last_day: object
    today: object
    lag_days: int
    stale: bool
    markets_behind: list
    text: str


@dataclass
class Commentary:
    verdict: str            # "Behind plan" / "On plan" / "Ahead of plan"
    severity: str           # good / warn / risk
    headline: str
    freshness: Freshness
    windows: list           # [Window]
    month_text: str
    open_items: list        # [str]


def _fmt_range(a, b) -> str:
    a, b = pd.Timestamp(a), pd.Timestamp(b)
    if a.month == b.month:
        return f"{a.day}–{b.day} {b.strftime('%b')}"
    return f"{a.strftime('%-d %b')}–{b.strftime('%-d %b')}"


def freshness(m: Model, today=None) -> Freshness:
    """How far behind the data is, and which markets are the cause.

    On a three-day review rhythm a missing day reads as a collapse that never
    happened, so this sits above everything else rather than in a footer.
    """
    today = pd.Timestamp(today or dt.date.today()).date()
    days = sorted(m.actuals["Day"].unique())
    if not days:
        return Freshness(None, today, 999, True, [], "No actuals entered yet.")
    last = days[-1]
    lag = (today - last).days
    behind = []
    for mk in sorted(m.actuals["Market"].unique()):
        d = m.actuals[m.actuals["Market"] == mk]
        mlast = max(d["Day"])
        if mlast < last:
            behind.append(f"{mk} to {pd.Timestamp(mlast).strftime('%-d %b')}")
    stale = lag > 1 or bool(behind)
    if not stale:
        txt = (f"Data is current — complete through "
               f"{pd.Timestamp(last).strftime('%-d %b')}, "
               f"all {m.actuals['Market'].nunique()} markets reporting.")
    else:
        bits = [f"Data runs to {pd.Timestamp(last).strftime('%-d %b')}"]
        if lag > 0:
            bits.append(f"{lag} day{'s' if lag != 1 else ''} behind today")
        txt = " — ".join(bits) + "."
        if behind:
            txt += " Behind: " + ", ".join(behind) + "."
        txt += (" Every window below is measured on reported days, not calendar "
                "days, so nothing counts an unentered day as a zero.")
    return Freshness(last, today, lag, stale, behind, txt)


def _window(m: Model, days: list, n: int, label: str, markets, channels,
            question: str) -> Optional[Window]:
    """One comparison window over REPORTED days.

    Calendar days would show a fake collapse for any day not yet entered.
    """
    if len(days) < n * 2:
        return None          # not enough reported days in this period to compare
    a = (days[-n], days[-1])
    b = (days[-2 * n], days[-n - 1])
    A = cmp_block(m, *a, markets, channels)
    B = cmp_block(m, *b, markets, channels)
    mets = {}
    for k in ("orders", "spend", "cpa", "revenue"):
        c = cmp_change(A, B, k)
        mets[k] = (A[k], c["pct"], c["read"])
    return Window(label, a[0], a[1], b[0], b[1],
                  f"{_fmt_range(*a)} vs {_fmt_range(*b)}", mets, "")


def _movers(m: Model, w: Window, markets, channels, top=3) -> list:
    """Which market x channel moved most between the two halves of a window."""
    out = []
    for mk in (markets or m.market_list()):
        for ch in (channels or m.display_channels(False)):
            kids = m.children_of(ch)
            a = cmp_block(m, w.a_from, w.a_to, [mk], kids)["orders"] or 0
            b = cmp_block(m, w.b_from, w.b_to, [mk], kids)["orders"] or 0
            if a or b:
                out.append((mk, ch, a - b, a, b))
    out.sort(key=lambda x: -abs(x[2]))
    return [x for x in out[:top] if abs(x[2]) >= 1]


def _moved_text(m: Model, w: Window, markets, channels) -> str:
    o_v, o_p, _ = w.metrics["orders"]
    s_v, s_p, _ = w.metrics["spend"]
    c_v, c_p, _ = w.metrics["cpa"]
    r_v, r_p, _ = w.metrics["revenue"]
    parts = []

    if o_p is None:
        parts.append(f"Orders reached {o_v:,.0f}, with nothing to compare against.")
    elif abs(o_p) < 1:
        parts.append(f"Orders held flat at {o_v:,.0f}.")
    else:
        rose = o_p > 0
        if s_p is not None and abs(s_p) >= 1:
            # Orders up on more spend is not the same as orders up on less.
            if rose and s_p > o_p:
                parts.append(f"Orders rose {abs(o_p):.0f}% but cost {abs(s_p):.0f}% "
                             f"more to get, so CPA moved the wrong way.")
            elif rose and s_p < 0:
                parts.append(f"Orders rose {abs(o_p):.0f}% on {abs(s_p):.0f}% "
                             f"less spend.")
            elif rose:
                parts.append(f"Orders rose {abs(o_p):.0f}% on {abs(s_p):.0f}% "
                             f"more spend.")
            else:
                parts.append(f"Orders fell {abs(o_p):.0f}% on "
                             f"{abs(s_p):.0f}% {'more' if s_p > 0 else 'less'} spend.")
        else:
            parts.append(f"Orders {'rose' if rose else 'fell'} {abs(o_p):.0f}% "
                         f"to {o_v:,.0f} on flat spend.")

    # Orders up while revenue falls means smaller baskets — a shift no single
    # table shows, because the two figures never sit together.
    if o_p is not None and r_p is not None and o_p > 1 and r_p < -1:
        parts.append(f"Revenue fell {abs(r_p):.0f}% against a rising order count — "
                     f"the orders coming in are smaller.")
    elif o_p is not None and r_p is not None and o_p < -1 and r_p > 1:
        parts.append(f"Revenue rose {abs(r_p):.0f}% despite fewer orders — "
                     f"the orders coming in are larger.")

    mv = _movers(m, w, markets, channels)
    if mv:
        bits = []
        for mk, ch, d, a, b in mv:
            if b == 0 and a > 0:
                bits.append(f"{mk} {ch} switched on, adding {a:,.0f} orders "
                            f"from a standing start")
            elif d > 0:
                bits.append(f"{mk} {ch} added {d:,.0f}")
            else:
                bits.append(f"{mk} {ch} gave back {abs(d):,.0f}")
        parts.append(bits[0][0].upper() + bits[0][1:] +
                     ("" if len(bits) == 1 else ", and " + ", ".join(bits[1:])) + ".")
    return " ".join(parts)


def _trend_text(m: Model, w: Window, short: Window) -> str:
    o_v, o_p, _ = w.metrics["orders"]
    c_v, c_p, _ = w.metrics["cpa"]
    days_a = (pd.Timestamp(w.a_to) - pd.Timestamp(w.a_from)).days + 1
    A = cmp_block(m, w.a_from, w.a_to)
    B = cmp_block(m, w.b_from, w.b_to)
    rate_a, rate_b = A["daily"], B["daily"]

    # The two windows can disagree, and that disagreement is the useful part.
    opposed = (short and short.metrics["cpa"][1] is not None and c_p is not None
               and (short.metrics["cpa"][1] > 0) != (c_p > 0))
    lead = ("Over seven days the picture is the opposite: " if opposed
            else "Over seven days: ")
    txt = (f"{lead}{rate_a:,.0f} orders/day against {rate_b:,.0f}")
    if c_p is not None:
        txt += (f", with CPA {'falling' if c_p < 0 else 'rising'} from "
                f"{fmt(B['cpa'],'AED ',dec=2)} to {fmt(A['cpa'],'AED ',dec=2)}")
    txt += "."
    if opposed:
        txt += (" Three days is short enough that one weak day reads as a "
                "reversal — the seven-day view is the one to act on.")
    return txt


def _month_text(m: Model, markets, channels, year, month, cov: Coverage) -> str:
    kw = dict(markets=markets, channels=channels, year=year, month=month)
    o = actual(m, M_ORDERS, **kw) or 0
    t = target_orders(m, markets, year, month)
    rev = actual(m, M_REVENUE, **kw) or 0
    t_rev = target_revenue(m, markets, year, month)
    sp = actual(m, M_SPEND, **kw) or 0
    ceil = plan_budget(m, markets, channels, year, month)
    label = month if isinstance(month, str) else f"{month[0]}–{month[-1]}"

    parts = []
    if cov.days_elapsed == 0:
        parts.append(f"{label} has not started — no actuals entered yet."
                     + (f" It is planned for {t:,.0f} orders." if t else ""))
    elif t:
        parts.append(f"{label} delivered {o:,.0f} orders against {t:,.0f} planned.")
    else:
        parts.append(f"{label} delivered {o:,.0f} orders, with no plan set.")

    if t_rev and rev:
        r = pct(rev, t_rev)
        if 95 <= r <= 105 and t and pct(o, t) < 90:
            # Orders short but revenue on plan can only mean larger baskets.
            parts.append(f"Revenue held at {fmt(rev,'AED ')}, essentially on plan, "
                         f"because the orders that did land were larger than assumed.")
        else:
            parts.append(f"Revenue reached {fmt(rev,'AED ')}, {r:.0f}% of plan.")

    if ceil and cov.days_elapsed == 0:
        parts.append(f"A budget of {fmt(ceil,'AED ')} is set. Nothing spent yet.")
    elif ceil:
        used = pct(sp, ceil)
        if used > 100:
            parts.append(f"Spend closed at {fmt(sp,'AED ')} against a "
                         f"{fmt(ceil,'AED ')} ceiling — over by "
                         f"{fmt(sp-ceil,'AED ')}.")
        else:
            parts.append(f"Spend closed at {fmt(sp,'AED ')} of a {fmt(ceil,'AED ')} "
                         f"ceiling, so budget was never the constraint.")

    g = gap_table(m, markets, channels, year, month, cov, False)
    behind = g[g["Share of gap"] > 0] if not g.empty else g
    if len(behind):
        top = behind.iloc[0]
        if top["Share of gap"] > 20:
            parts.append(f"{top['Market']} {top['Channel']} accounts for "
                         f"{top['Share of gap']:.0f}% of the miss — "
                         f"{top['Actual']:,.0f} orders against "
                         f"{top['Paced plan']:,.0f} planned.")
    return " ".join(parts)


def _open_items(m: Model, markets, channels, year, month, cov: Coverage) -> list:
    """What the numbers say is unchanged. Not a task list — this is what stops a
    three-day review repeating itself."""
    out = []
    if cov.days_elapsed == 0:
        return out          # nothing can be open in a period that has not run
    A = allocation_table(m, markets, channels, year, month, cov, False)
    if A.empty:
        return out

    idle = A[A["Unspent"].fillna(0) > 0]
    if len(idle) and idle["Unspent"].sum() > 500:
        top = idle.loc[idle["Unspent"].idxmax()]
        out.append(f"**{fmt(idle['Unspent'].sum(),'AED ')} of paced budget has not "
                   f"gone out**, most of it in {top['Market']} {top['Channel']}. "
                   f"A constraint of execution, not of the channel.")

    ctx = cpa_context(m, markets, channels, year, month)
    if len(ctx):
        worst = ctx.loc[ctx["vs cheapest"].idxmax()]
        if worst["vs cheapest"] >= 2:
            best_cpa = worst["CPA"] / worst["vs cheapest"]
            would = worst["Spend"] / best_cpa
            out.append(
                f"**{worst['Market']} {worst['Channel']} still costs "
                f"{fmt(worst['CPA'],'AED ',dec=2)} per order** against "
                f"{worst['Market']} {worst['Cheapest here']}'s "
                f"{fmt(best_cpa,'AED ',dec=2)} — "
                f"{worst['vs cheapest']:.0f} of those orders for the price of one. "
                f"The same {fmt(worst['Spend'],'AED ')} would have bought roughly "
                f"{would:,.0f} orders instead of {worst['Orders']:,.0f}. "
                f"A ceiling, not a forecast: it assumes the cheaper channel absorbs "
                f"the budget without its cost rising.")

    cc = capacity_check(m, markets, year,
                        month if isinstance(month, str) else month[-1])
    if len(cc):
        short = cc[cc["Hit"].fillna(100) < 90]
        for _, r in short.iterrows():
            out.append(f"**{r['Market']} API delivered {r['Hit']:.0f}% of its "
                       f"modelled capacity.** {r['Read']}")
    return out


def commentary(m: Model, markets, channels, year, month, cov: Coverage,
               today=None) -> Commentary:
    kw_days = scope(m.actuals, markets, channels, year, month)
    days = sorted(kw_days["Day"].unique())
    all_days = sorted(m.actuals["Day"].unique())

    o = actual(m, M_ORDERS, markets=markets, channels=channels,
               year=year, month=month) or 0
    t = target_orders(m, markets, year, month)
    label0 = month if isinstance(month, str) else f"{month[0]}–{month[-1]}"

    # A month that has not started yet has nothing to measure. Scoring it
    # produces "0% of capacity" and "budget was never the constraint", which are
    # arithmetically true and completely wrong.
    if not days:
        ceil0 = plan_budget(m, markets, channels, year, month)
        if t:
            txt = (f"{label0} has not started — no actuals entered yet. "
                   f"It is planned for {t:,.0f} orders")
            txt += f" on a budget of {fmt(ceil0, 'AED ')}." if ceil0 else "."
            v, sv = "Not started", "info"
            head = f"{label0} has not started"
        else:
            txt = f"{label0} has neither a plan nor any actuals."
            v, sv = "Nothing to show", "info"
            head = f"{label0} has nothing to show"
        return Commentary(v, sv, head, freshness(m, today), [], txt, [])

    r = pct(o, paced(t, cov)) if t else None
    if r is None:
        verdict, sev = "No plan", "info"
    elif r >= 95:
        verdict, sev = ("Ahead of plan", "good") if r > 105 else ("On plan", "good")
    elif r >= 80:
        verdict, sev = "Slightly behind", "warn"
    else:
        verdict, sev = "Behind plan", "risk"

    label = month if isinstance(month, str) else f"{month[0]}–{month[-1]}"
    closed = cov.days_remaining == 0
    headline = (f"{label} {'closed at' if closed else 'is running at'} "
                f"{r:.0f}% of {'plan' if closed else 'paced plan'}"
                if r is not None else f"{label} has no plan to measure against")

    # Windows are drawn from days INSIDE the selected period. Using every day in
    # the workbook meant selecting August showed July's movement under an August
    # heading — the figures were right and the label made them a lie.
    wins = []
    w3 = _window(m, days, 3, "What moved", markets, channels, "what changed")
    w7 = _window(m, days, 7, "Is it a trend", markets, channels, "is it a trend")
    if w3:
        w3.text = _moved_text(m, w3, markets, channels)
        wins.append(w3)
    if w7:
        w7.text = _trend_text(m, w7, w3)
        wins.append(w7)

    return Commentary(verdict, sev, headline, freshness(m, today), wins,
                      _month_text(m, markets, channels, year, month, cov),
                      _open_items(m, markets, channels, year, month, cov))


# ─────────────────────────────────────────────────────────────────────
# CARD MODEL
#
# Four paragraphs of prose is a report, not a dashboard. Everything a visual can
# carry belongs on a card; prose is reserved for what a visual cannot say — a
# relationship between two figures, or a pattern that spans rows.
# ─────────────────────────────────────────────────────────────────────
@dataclass
class MetricCard:
    key: str
    label: str
    value: Optional[float]
    paced: Optional[float]
    plan: Optional[float]
    eom: Optional[float]
    pct_of: Optional[float]      # % of paced, or % of plan for a ratio
    basis: str                   # what pct_of is measured against
    colour: str
    spark: list
    foot: str
    prefix: str = ""
    suffix: str = ""
    dec: int = 0
    ratio: bool = False          # a ratio never paces


def _spark(series: pd.Series, n=31) -> list:
    if not len(series):
        return []
    return [float(v) for v in series.values][-n:]


def _card_colour(pctv, direction) -> str:
    if pctv is None:
        return GREY
    if direction == "neutral":
        return GREY
    if direction == "up":
        return GREEN if pctv >= 90 else AMBER if pctv >= 70 else RED
    return GREEN if pctv <= 105 else AMBER if pctv <= 120 else RED


def management_cards(m: Model, markets, channels, year, month,
                     cov: Coverage) -> list:
    """Orders, revenue, spend, ROAS, CPA, AOV.

    Volume metrics carry paced, plan and a landing figure. Ratios carry plan
    only: pacing a ratio is meaningless, since it does not accumulate.
    """
    kw = dict(markets=markets, channels=channels, year=year, month=month)
    o = actual(m, M_ORDERS, **kw)
    rev = actual(m, M_REVENUE, **kw)
    sp = actual(m, M_SPEND, **kw)
    t_o = target_orders(m, markets, year, month)
    t_rev = target_revenue(m, markets, year, month)
    ceil = plan_budget(m, markets, channels, year, month)
    p_o = plan_orders(m, markets, channels, year, month)

    d_o = daily_series(m, M_ORDERS, markets, channels, year, month)
    d_rev = daily_series(m, M_REVENUE, markets, channels, year, month)
    d_sp = daily_series(m, M_SPEND, markets, channels, year, month)
    # A day can carry orders and no spend row, or the reverse, so the series do
    # not share an index. Zipping them by position silently pairs the wrong days
    # — or fails outright when the lengths differ.
    def _ratio_series(num: pd.Series, den: pd.Series) -> pd.Series:
        if not len(num) or not len(den):
            return pd.Series(dtype=float)
        idx = num.index.union(den.index).sort_values()
        n = num.reindex(idx)
        d = den.reindex(idx)
        return pd.Series(
            [(a / b) if (pd.notna(a) and pd.notna(b) and b) else np.nan
             for a, b in zip(n.values, d.values)], index=idx)

    d_cpa = _ratio_series(d_sp, d_o)
    d_roas = _ratio_series(d_rev, d_sp)
    d_aov = _ratio_series(d_rev, d_o)

    closed = cov.days_remaining == 0
    basis = "plan" if closed else "paced"

    def vol(key, label, val, target, series, pfx="", dec=0, direction="up",
            foot_extra=""):
        pc = paced(target, cov)
        r = pct(val, pc)
        eom_v = None if closed else eom(val, cov)
        foot = (f"{'plan' if closed else 'paced'} {fmt(pc, pfx, dec=dec)}"
                + (f" · plan {fmt(target, pfx, dec=dec)}" if not closed else ""))
        if eom_v is not None and target:
            foot += f"\nlands at {fmt(eom_v, pfx, dec=dec)} — {pct(eom_v, target):.0f}%"
        if foot_extra:
            foot += f"\n{foot_extra}"
        return MetricCard(key, label, val, pc, target, eom_v, r, basis,
                          _card_colour(r, direction), _spark(series), foot,
                          pfx, "", dec)

    def ratio(key, label, val, plan_v, series, pfx="", sfx="", dec=1,
              direction="up", foot=""):
        r = pct(val, plan_v)
        return MetricCard(key, label, val, None, plan_v, None, r, "plan",
                          _card_colour(r, direction), _spark(series), foot,
                          pfx, sfx, dec, ratio=True)

    ceil_day = div(ceil, cov.days_in_month)
    sp_day = div(sp, cov.days_elapsed)

    return [
        vol("orders", "Orders", o, t_o, d_o),
        vol("revenue", "Revenue", rev, t_rev, d_rev, "AED "),
        vol("spend", "Spend", sp, ceil, d_sp, "AED ", direction="neutral",
            foot_extra=(f"{fmt(ceil_day,'AED ')}/day allowed · "
                        f"{fmt(sp_day,'AED ')}/day actual"
                        if ceil_day and sp_day else "")),
        ratio("roas", "ROAS", div(rev, sp), div(t_rev, ceil), d_roas,
              sfx="x", foot="revenue ÷ spend, always"),
        ratio("cpa", "CPA", div(sp, o), div(ceil, p_o), d_cpa, "AED ", dec=2,
              direction="down", foot="budget control only — see Why"),
        ratio("aov", "AOV", div(rev, o), div(t_rev, t_o), d_aov, "AED ", dec=0,
              foot="observed, never used to derive"),
    ]


def management_line(m: Model, markets, channels, year, month,
                    cov: Coverage) -> Optional[str]:
    """The one sentence a card cannot carry: a relationship between two of them,
    and which cell is driving the gap."""
    kw = dict(markets=markets, channels=channels, year=year, month=month)
    o = actual(m, M_ORDERS, **kw)
    rev = actual(m, M_REVENUE, **kw)
    t_o = target_orders(m, markets, year, month)
    t_rev = target_revenue(m, markets, year, month)
    if not cov.days_elapsed:
        return None

    parts = []
    o_pct, r_pct = pct(o, paced(t_o, cov)), pct(rev, paced(t_rev, cov))
    if o_pct is not None and r_pct is not None and abs(r_pct - o_pct) > 8:
        bigger = r_pct > o_pct
        parts.append(
            f"Orders are {o_pct:.0f}% of pace while revenue is {r_pct:.0f}% — "
            f"<b>{'fewer orders, but larger ones' if bigger else 'more orders, but smaller ones'}</b>.")

    g = gap_table(m, markets, channels, year, month, cov, False)
    behind = g[g["Share of gap"] > 0] if not g.empty else g
    if len(behind):
        top = behind.iloc[0]
        if top["Share of gap"] >= 20:
            parts.append(f"<b>{top['Market']} {top['Channel']} is "
                         f"{top['Share of gap']:.0f}% of the gap</b>, "
                         f"{top['Actual']:,.0f} against {top['Paced plan']:,.0f} paced.")

    sp = actual(m, M_SPEND, **kw)
    ceil = plan_budget(m, markets, channels, year, month)
    if sp and ceil:
        if sp > ceil:
            parts.append(f"<b>Spend has passed its ceiling by "
                         f"{fmt(sp-ceil,'AED ')}.</b>")
        elif pct(sp, paced(ceil, cov)) and pct(sp, paced(ceil, cov)) < 95:
            parts.append("Spend is tracking below its ceiling, so budget is not "
                         "the constraint.")
    return " ".join(parts) if parts else None


def where_line(m: Model, markets, channels, year, month, cov: Coverage):
    """A pattern that spans rows — invisible in any single one."""
    g = gap_table(m, markets, channels, year, month, cov, False)
    if g.empty:
        return None
    parts = []
    short = g[(g["vs paced"].notna()) & (g["vs paced"] < 90)]
    ok = g[(g["vs paced"].notna()) & (g["vs paced"] >= 90)]
    if len(short) and len(ok):
        s_ch, o_ch = set(short["Channel"]), set(ok["Channel"])
        if len(s_ch) == 1 and not (s_ch & o_ch):
            parts.append(f"Every cell short of pace is <b>{list(s_ch)[0]}</b>; "
                         f"every {', '.join(sorted(o_ch))} cell is at or ahead of it.")
    never = g[(g["Actual"] == 0) & (g["Spend"] == 0)
              & (g["Paced plan"].fillna(0) > 0)]
    for _, r in never.iterrows():
        parts.append(f"<b>{r['Market']} {r['Channel']} never ran</b> — "
                     f"{r['Paced plan']:,.0f} paced orders with no spend against "
                     f"them, {r['Share of gap']:.0f}% of the gap.")
        break
    return " ".join(parts) if parts else None


def split_line(m: Model, markets, year, month, parent=None) -> Optional[str]:
    """Where a channel's platforms disagree by market — the finding a
    consolidated figure hides entirely.

    The parent is read from the register, never named here: a second channel
    planned as one and reported as several would work with no code change.
    """
    par = m.parent_of()
    if not par:
        return None
    if parent is None:
        parent = max(set(par.values()), key=lambda p: sum(1 for v in par.values()
                                                          if v == p))
    kids = [c for c, p in par.items() if p == parent]
    if len(kids) < 2:
        return None
    cheaper = {}
    for mk in (markets or m.market_list()):
        best, bch = None, None
        for k in kids:
            o = actual(m, M_ORDERS, markets=[mk], channels=[k],
                       year=year, month=month) or 0
            sp = actual(m, M_SPEND, markets=[mk], channels=[k],
                        year=year, month=month) or 0
            c = div(sp, o)
            if c and (best is None or c < best):
                best, bch = c, k
        if bch:
            cheaper[mk] = (bch, best)
    if len(set(v[0] for v in cheaper.values())) < 2:
        return None
    bits = [f"in <b>{mk}, {ch} is cheaper</b> ({fmt(c,'AED ',dec=2)})"
            for mk, (ch, c) in cheaper.items()]
    return (f"The two platforms diverge by market: " + "; ".join(bits)
            + f". A consolidated {parent} figure hides both.")


def compare_line(m: Model, ar, br, markets, channels) -> Optional[str]:
    A, B = cmp_block(m, *ar, markets, channels), cmp_block(m, *br, markets, channels)
    o, r, s = (cmp_change(A, B, "orders"), cmp_change(A, B, "revenue"),
               cmp_change(A, B, "spend"))
    if o["pct"] is None:
        return None
    parts = []
    sp_txt = (f"on {abs(s['pct']):.0f}% {'more' if s['pct'] > 0 else 'less'} spend"
              if s["pct"] is not None and abs(s["pct"]) >= 1 else "on flat spend")
    parts.append(f"Orders {'rose' if o['pct'] > 0 else 'fell'} "
                 f"<b>{abs(o['pct']):.0f}%</b> {sp_txt}, so CPA "
                 f"{'rose' if A['cpa'] > B['cpa'] else 'fell'} from "
                 f"{fmt(B['cpa'],'AED ',dec=2)} to <b>{fmt(A['cpa'],'AED ',dec=2)}</b> "
                 f"and ROAS moved {B['roas']:.1f}x to {A['roas']:.1f}x.")
    # Revenue moving in step with orders means basket size held — a volume
    # problem rather than a mix problem, and neither card says so alone.
    if r["pct"] is not None and abs(r["pct"] - o["pct"]) < 3:
        parts.append("Revenue moved in step with orders, so basket size held — "
                     "this is a volume problem, not a mix problem.")
    elif r["pct"] is not None and abs(r["pct"] - o["pct"]) > 8:
        parts.append(f"Revenue moved {r['pct']:+.0f}% against orders at "
                     f"{o['pct']:+.0f}% — basket size shifted.")
    return " ".join(parts)
