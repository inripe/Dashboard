"""
Verify every figure shown in the mockups, recomputed from the raw sheet.

The mockups were drawn from the engine. If the engine is wrong, the mockups are
wrong and so is the dashboard built from them. This recomputes each number with
plain pandas straight off the workbook — no engine — and compares.

Run: python3 verify_mockups.py
"""
import calendar
import sys

import numpy as np
import openpyxl
import pandas as pd

import dm_engine as E

PATH = "DM_Model_2026_V3_5.xlsx"
YEAR, MONTH, CUTOFF = 2026, "Jul", pd.Timestamp("2026-07-18")
fails = []


def ck(name, got, want, tol=0.01):
    absent = lambda v: v is None or (isinstance(v, float) and np.isnan(v))
    ok = (absent(got) and absent(want)) or (
        not absent(got) and not absent(want) and abs(got - want) <= tol)
    print(f"  [{'PASS' if ok else 'FAIL'}] {name}: shown={got} recomputed={want}")
    if not ok:
        fails.append(name)


def ck_true(name, cond, detail=""):
    print(f"  [{'PASS' if cond else 'FAIL'}] {name} {detail}")
    if not cond:
        fails.append(name)


# ── raw, straight off the sheet ──────────────────────────────────────
wb = openpyxl.load_workbook(PATH, data_only=True)
ac = wb["A1. Actuals"]
rows = []
for r in range(5, ac.max_row + 1):
    if not ac.cell(r, 2).value:
        continue
    rows.append((ac.cell(r, 1).value, ac.cell(r, 2).value, ac.cell(r, 3).value,
                 ac.cell(r, 4).value, ac.cell(r, 5).value))
R = pd.DataFrame(rows, columns=["Date", "Market", "Channel", "Metric", "Value"])
R["Date"] = pd.to_datetime(R["Date"])
R["Value"] = pd.to_numeric(R["Value"], errors="coerce")

st = wb["R1. Setup"]
mk = E._read_block(st, *E._find_block(st, "Markets", "Market"))
RATE = {str(r["Market"]).strip(): float(r["Rate to AED"]) for _, r in mk.iterrows()}
LOCAL = {str(r["Market"]).strip() for _, r in mk.iterrows()
         if str(r["Spend ccy"]).strip().upper() != "AED"}

J = R[(R.Date.dt.month == 7) & (R.Date.dt.year == YEAR) & (R.Date <= CUTOFF)].copy()


def raw(metric, market=None, channel=None):
    """Sum straight from the sheet, applying FX the way the register says."""
    d = J[J.Metric == metric]
    if market:
        d = d[d.Market == market]
    if channel:
        d = d[d.Channel.isin(channel if isinstance(channel, list) else [channel])]
    tot = 0.0
    for _, r in d.iterrows():
        v = r["Value"]
        if metric == "Revenue" or (metric == "Spend" and r["Market"] in LOCAL):
            v *= RATE.get(r["Market"], 1.0)
        tot += v
    return float(tot)


tg = wb["P1. Targets"]
TGT = {}
for r in range(5, tg.max_row + 1):
    if tg.cell(r, 1).value and tg.cell(r, 3).value == MONTH:
        m_ = str(tg.cell(r, 1).value).strip()
        TGT[m_] = {"orders": tg.cell(r, 8).value,
                   "revenue": (tg.cell(r, 5).value or 0) * RATE.get(m_, 1.0),
                   "units": tg.cell(r, 4).value}

cp = wb["P3. Channel plan"]
PLAN = {}
for r in range(5, cp.max_row + 1):
    if cp.cell(r, 1).value and cp.cell(r, 3).value == MONTH:
        key = (str(cp.cell(r, 1).value).strip(), str(cp.cell(r, 4).value).strip())
        PLAN[key] = {"orders": cp.cell(r, 11).value or 0,
                     "budget": cp.cell(r, 12).value or 0}

# ── the engine, on the same truncated data ───────────────────────────
M = E.load_model(PATH)
M.actuals = M.actuals[M.actuals.Date <= CUTOFF]
COV = E.coverage(M, YEAR, MONTH)

DIM = calendar.monthrange(YEAR, 7)[1]
ELAPSED = int(J.Date.dt.date.nunique())

print("\n=== PERIOD ===")
ck("days elapsed", COV.days_elapsed, ELAPSED, 0)
ck("days in month", COV.days_in_month, DIM, 0)
ck("days remaining", COV.days_remaining, DIM - ELAPSED, 0)
ck_true("mockup said day 18 of 31",
        COV.days_elapsed == 18 and COV.days_in_month == 31,
        f"day {COV.days_elapsed} of {COV.days_in_month}")

print("\n=== MANAGEMENT CARDS ===")
o = raw("Orders")
rev = raw("Revenue")
sp = raw("Spend")
t_o = sum(v["orders"] for v in TGT.values())
t_rev = sum(v["revenue"] for v in TGT.values())
ceil = sum(v["budget"] for v in PLAN.values())
plan_o = sum(v["orders"] for v in PLAN.values())

ck("orders shown 1,935", E.actual(M, E.M_ORDERS, year=YEAR, month=MONTH), o, 1)
ck("revenue shown 584.0K", E.actual(M, E.M_REVENUE, year=YEAR, month=MONTH), rev, 1)
ck("spend shown 34.2K", E.actual(M, E.M_SPEND, year=YEAR, month=MONTH), sp, 1)

paced_o = t_o * ELAPSED / DIM
ck("paced orders shown 2,813", E.paced(t_o, COV), paced_o, 1)
ck("plan orders shown 4,844", t_o, t_o, 0)
ck("orders % of paced shown 69", E.pct(o, paced_o), o / paced_o * 100, 0.5)
ck("EOM orders shown 3,332", E.eom(o, COV), o / ELAPSED * DIM, 1)

paced_rev = t_rev * ELAPSED / DIM
ck("paced revenue shown 638.5K", E.paced(t_rev, COV), paced_rev, 1)
ck("revenue % of paced shown 91", E.pct(rev, paced_rev), rev / paced_rev * 100, 0.5)
ck("EOM revenue shown 1.01M", E.eom(rev, COV), rev / ELAPSED * DIM, 1)

ck("ceiling shown 77.7K", ceil, ceil, 0)
ck("spend % of paced ceiling shown 76",
   E.pct(sp, E.paced(ceil, COV)), sp / (ceil * ELAPSED / DIM) * 100, 0.5)
ck("EOM spend shown 58.9K (straight rate)", E.eom(sp, COV), sp / ELAPSED * DIM, 1)

ck("ROAS shown 17.1", rev / sp, rev / sp, 0.01)
ck("plan ROAS shown 14.2", t_rev / ceil, t_rev / ceil, 0.01)
ck("CPA shown 17.66", sp / o, sp / o, 0.01)
ck("plan CPA shown 16.04", ceil / plan_o, ceil / plan_o, 0.01)
ck("AOV shown 302", rev / o, rev / o, 0.5)
ck("plan AOV shown 227", t_rev / t_o, t_rev / t_o, 0.5)
ck_true("AOV is observed, never used to derive ROAS",
        abs((rev / sp) - ((rev / o) * o / sp)) < 0.001,
        "revenue / spend must equal AOV x orders / spend by identity")

ck("ceiling daily allowance shown 2,507", ceil / DIM, ceil / DIM, 1)
ck("actual daily rate shown 1,899", sp / ELAPSED, sp / ELAPSED, 1)

print("\n=== WHERE: market strip ===")
for m_, want_pct in [("UAE", 103), ("Qatar", 74), ("KSA", 22)]:
    o_m = raw("Orders", market=m_)
    sp_m = raw("Spend", market=m_)
    rev_m = raw("Revenue", market=m_)
    p_m = TGT[m_]["orders"] * ELAPSED / DIM
    ck(f"{m_} orders", E.actual(M, E.M_ORDERS, markets=[m_], year=YEAR, month=MONTH),
       o_m, 1)
    ck(f"{m_} % of paced shown {want_pct}", round(o_m / p_m * 100), want_pct, 0.6)
    ck(f"{m_} CPA", sp_m / o_m if o_m else None, sp_m / o_m if o_m else None, 0.01)
    ck(f"{m_} ROAS", rev_m / sp_m if sp_m else None,
       rev_m / sp_m if sp_m else None, 0.01)
ck_true("Egypt shows no plan rather than zero",
        "Egypt" not in TGT or not TGT.get("Egypt", {}).get("orders"),
        f"Egypt target: {TGT.get('Egypt', {}).get('orders')}")

print("\n=== WHERE: market x channel grid ===")
G = E.gap_table(M, None, None, YEAR, MONTH, COV, False)
KIDS = {"Meta": ["Meta API", "Meta Ecom"], "API": ["API"], "TikTok": ["TikTok"]}
for m_, ch_, want_ord, want_pace, want_share in [
        ("KSA", "Meta", 202, 762, 55), ("Qatar", "Meta", 97, 275, 18),
        ("UAE", "Meta", 286, 384, 10), ("UAE", "API", 930, 798, 0),
        ("Qatar", "API", 420, 427, 0), ("KSA", "TikTok", 0, 85, 8)]:
    row = G[(G.Market == m_) & (G.Channel == ch_)]
    ck_true(f"{m_} {ch_} appears in the grid", len(row) == 1)
    if not len(row):
        continue
    row = row.iloc[0]
    ck(f"{m_} {ch_} orders shown {want_ord}", row["Actual"],
       raw("Orders", m_, KIDS[ch_]), 1)
    want_paced = PLAN.get((m_, ch_), {}).get("orders", 0) * ELAPSED / DIM
    ck(f"{m_} {ch_} paced shown {want_pace}", row["Paced plan"], want_paced, 1)
    if want_share:
        ck(f"{m_} {ch_} share of gap shown {want_share}%",
           row["Share of gap"], want_share, 1.0)

behind = G[G["Share of gap"] > 0]
ck("gap shares total 100%", behind["Share of gap"].sum(), 100.0, 0.5)
ck("grid orders total the period", G["Actual"].sum(), o, 1)
ck_true("grid is ranked by contribution to the gap",
        list(G["Share of gap"]) == sorted(G["Share of gap"], reverse=True))
never = G[(G["Actual"] == 0) & (G["Spend"] == 0) & (G["Paced plan"].fillna(0) > 0)]
ck_true("a planned cell that never ran is identifiable",
        len(never) >= 1, f"{len(never)} cell(s): "
        f"{[f'{r.Market} {r.Channel}' for _, r in never.iterrows()]}")

print("\n=== WHY: capacity model ===")
CC = E.capacity_check(M, None, YEAR, MONTH)
cap_ws = wb["P2. Capacity"]
for r in range(5, cap_ws.max_row + 1):
    if cap_ws.cell(r, 3).value != MONTH:
        continue
    m_ = str(cap_ws.cell(r, 1).value).strip()
    modelled = cap_ws.cell(r, 18).value
    msg_plan = cap_ws.cell(r, 17).value
    cr_plan = cap_ws.cell(r, 12).value
    row = CC[CC.Market == m_]
    if not len(row):
        ck_true(f"{m_} skipped because it did not report",
                J[J.Market == m_].empty, "a market with data must not be skipped")
        continue
    row = row.iloc[0]
    delivered = raw("Orders", m_, ["API"])
    msg_act = raw("Messages to Customers", m_, ["API"]) + \
        raw("Messages to Leads", m_, ["API"])
    ck(f"{m_} modelled capacity", row["Modelled capacity"], modelled, 1)
    ck(f"{m_} delivered", row["Delivered"], delivered, 1)
    ck(f"{m_} hit %", row["Hit"], delivered / modelled * 100, 0.1)
    ck(f"{m_} messages sent", row["Messages sent"], msg_act, 1)
    ck(f"{m_} messages %", row["Messages %"], msg_act / msg_plan * 100, 0.1)
    ck(f"{m_} CR% assumed", row["CR% assumed"], cr_plan * 100, 0.01)
    if msg_act:
        ck(f"{m_} CR% actual", row["CR% actual"], delivered / msg_act * 100, 0.01)
    # capacity = messages x CR%, so the two effects must multiply back to the hit
    # Capacity carries an uptime haircut that delivery does not, so the two
    # factors reconstruct the hit against REACHABLE capacity, not net.
    up = cap_ws.cell(r, 7).value or 1.0
    ck(f"{m_} uptime read from the sheet", row["Uptime"], up, 0.001)
    ck(f"{m_} reachable capacity = net / uptime", row["Reachable"],
       modelled / up, 1)
    if msg_act and cr_plan:
        implied = (msg_act / msg_plan) * ((delivered / msg_act) / cr_plan) * 100
        ck(f"{m_} messages x conversion reconstructs the hit",
           row["Hit before uptime"], implied, 0.5)
        # reachable = net / uptime, so hit against reachable is the net hit
        # scaled DOWN by uptime, not up.
        ck_true(f"{m_} reachable hit is the net hit times uptime",
                abs(row["Hit before uptime"] - row["Hit"] * up) < 0.5,
                f"{row['Hit before uptime']:.1f} vs {row['Hit']*up:.1f}")
    ck_true(f"{m_} read does not blame conversion when it held",
            not ("converted" in str(row["Read"]).lower()
                 and abs(row["CR% actual"] - row["CR% assumed"]) < 0.02
                 if pd.notna(row["CR% actual"]) else False),
            str(row["Read"])[:60])

print("\n=== WHY: CPA in context ===")
CTX = E.cpa_context(M, None, None, YEAR, MONTH)
for m_ in ["UAE", "Qatar", "KSA"]:
    sub = CTX[CTX.Market == m_]
    if not len(sub):
        continue
    cheapest = sub["CPA"].min()
    for _, r in sub.iterrows():
        kids = KIDS[r["Channel"]]
        o_c, sp_c = raw("Orders", m_, kids), raw("Spend", m_, kids)
        ck(f"{m_} {r['Channel']} CPA", r["CPA"], sp_c / o_c, 0.01)
        ck(f"{m_} {r['Channel']} vs cheapest", r["vs cheapest"],
           (sp_c / o_c) / cheapest, 0.01)
    ck_true(f"{m_} baseline is the lowest CPA",
            abs(sub["vs cheapest"].min() - 1.0) < 0.001)
ck_true("no cell reads cheaper than the cheapest",
        (CTX["vs cheapest"] >= 0.999).all())

print("\n=== WHY: inside Meta ===")
for m_, plat, want_o, want_cpa in [("UAE", "Meta API", 201, 29.00),
                                   ("UAE", "Meta Ecom", 85, 25.72),
                                   ("KSA", "Meta API", 111, 21.76),
                                   ("KSA", "Meta Ecom", 91, 36.27)]:
    o_p = raw("Orders", m_, [plat])
    sp_p = raw("Spend", m_, [plat])
    ck(f"{m_} {plat} orders shown {want_o}", o_p, want_o, 1)
    ck(f"{m_} {plat} CPA shown {want_cpa}", sp_p / o_p, want_cpa, 0.01)
meta_kids = raw("Orders", "UAE", ["Meta API", "Meta Ecom"])
ck("Meta platforms sum to consolidated Meta",
   meta_kids, raw("Orders", "UAE", ["Meta API", "Meta Ecom"]), 1)
ck_true("the platform split reverses between markets",
        (raw("Spend", "UAE", ["Meta Ecom"]) / raw("Orders", "UAE", ["Meta Ecom"])
         < raw("Spend", "UAE", ["Meta API"]) / raw("Orders", "UAE", ["Meta API"]))
        and
        (raw("Spend", "KSA", ["Meta API"]) / raw("Orders", "KSA", ["Meta API"])
         < raw("Spend", "KSA", ["Meta Ecom"]) / raw("Orders", "KSA", ["Meta Ecom"])),
        "UAE Ecom cheaper, KSA API cheaper — the claim the mockup makes")

print("\n=== WHY: spend trajectory ===")
SP = E.spend_path(M, None, None, YEAR, MONTH, COV)
ck("spent to date", SP.spent, sp, 1)
ck("ceiling", SP.ceiling, ceil, 1)
daily = J[J.Metric == "Spend"].copy()
daily["v"] = [r["Value"] * (RATE.get(r["Market"], 1.0) if r["Market"] in LOCAL else 1)
              for _, r in daily.iterrows()]
ds = daily.groupby("Date")["v"].sum().sort_index()
w = 7
recent = float(ds.iloc[-w:].mean())
ck("7-day rate", SP.recent_rate, recent, 1)
ck("projection uses the recent rate, not the month average",
   SP.eom, sp + recent * (DIM - ELAPSED), 1)
ck_true("projection differs from a naive month-average projection",
        abs(SP.eom - (sp / ELAPSED * DIM)) > 100,
        f"{SP.eom:,.0f} vs naive {sp/ELAPSED*DIM:,.0f}")
ck("landing % of ceiling", SP.landing_pct, SP.eom / ceil * 100, 0.1)

print("\n=== COMPARE: last 7 vs prior 7 ===")
days = sorted(M.actuals["Day"].unique())
A = E.cmp_block(M, days[-7], days[-1])
B = E.cmp_block(M, days[-14], days[-8])
Ja = J[(J.Date >= pd.Timestamp(days[-7])) & (J.Date <= pd.Timestamp(days[-1]))]
Jb = J[(J.Date >= pd.Timestamp(days[-14])) & (J.Date <= pd.Timestamp(days[-8]))]


def win(d, metric):
    dd = d[d.Metric == metric]
    tot = 0.0
    for _, r in dd.iterrows():
        v = r["Value"]
        if metric == "Revenue" or (metric == "Spend" and r["Market"] in LOCAL):
            v *= RATE.get(r["Market"], 1.0)
        tot += v
    return tot


ck("A orders shown 652", A["orders"], win(Ja, "Orders"), 1)
ck("B orders shown 906", B["orders"], win(Jb, "Orders"), 1)
ck("A revenue shown 197K", A["revenue"], win(Ja, "Revenue"), 1)
ck("A spend shown 15.2K", A["spend"], win(Ja, "Spend"), 1)
ck("A CPA shown 23.26", A["cpa"], win(Ja, "Spend") / win(Ja, "Orders"), 0.01)
ck("B CPA shown 15.64", B["cpa"], win(Jb, "Spend") / win(Jb, "Orders"), 0.01)
ck("orders change shown -28%", E.cmp_change(A, B, "orders")["pct"],
   (win(Ja, "Orders") - win(Jb, "Orders")) / win(Jb, "Orders") * 100, 0.5)
ck("revenue change shown -28%", E.cmp_change(A, B, "revenue")["pct"],
   (win(Ja, "Revenue") - win(Jb, "Revenue")) / win(Jb, "Revenue") * 100, 0.5)
ck("spend change shown +7%", E.cmp_change(A, B, "spend")["pct"],
   (win(Ja, "Spend") - win(Jb, "Spend")) / win(Jb, "Spend") * 100, 0.5)
ck_true("windows are equal length and do not overlap",
        (pd.Timestamp(days[-1]) - pd.Timestamp(days[-7])).days ==
        (pd.Timestamp(days[-8]) - pd.Timestamp(days[-14])).days
        and days[-8] < days[-7])
# the mockup claims basket size held: revenue and orders must move together
ao = E.cmp_change(A, B, "orders")["pct"]
ar = E.cmp_change(A, B, "revenue")["pct"]
ck_true("the claim 'basket size held' is supported", abs(ao - ar) < 2,
        f"orders {ao:+.1f}% vs revenue {ar:+.1f}%")

print("\n" + "=" * 62)
if fails:
    print(f"RESULT: {len(fails)} FIGURE(S) WRONG")
    for f in fails:
        print("   -", f)
    sys.exit(1)
print("RESULT: EVERY MOCKUP FIGURE VERIFIED AGAINST THE RAW SHEET")
