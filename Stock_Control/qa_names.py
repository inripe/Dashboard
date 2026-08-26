# -*- coding: utf-8 -*-
"""
No retired movement name may survive anywhere in the code.

This is the suite that was missing when Customs / Loss was renamed: the sheet
changed, the code did not, and a correct entry silently stopped counting.
"""
import sys, glob, re
import engine, entry_ui, labels as L, qa_book
P,F=[],[]
def ck(n,ok,note=""):
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}  {note}")

RETIRED = ["Customs / Loss", "Orders Assigned", "Courier Handover"]
CODE = [f for f in glob.glob("*.py")
        if not f.startswith("qa_") and f not in ("migrate_v4.py",)]

print("=== A. RETIRED NAMES ARE GONE FROM THE CODE ===")
for name in RETIRED:
    hits=[]
    for f in CODE:
        for i, line in enumerate(open(f).read().split("\n"), 1):
            if name in line and not line.strip().startswith("#"):
                hits.append(f"{f}:{i}")
    ck(f"'{name}' appears nowhere", not hits, hits[:3])
ck("Delivered is not a movement the engine looks for",
   "Delivered" not in engine.MV, engine.MV)
ck("nor a movement the form offers",
   "Delivered" not in entry_ui.NEEDS, sorted(entry_ui.NEEDS))
ck("nor a label", "Delivered" not in L.MOVES, sorted(L.MOVES))

print("=== B. THE SHEET AND THE CODE AGREE ===")
import openpyxl
ws = openpyxl.load_workbook(qa_book.book())["MASTER"]
sheet_moves, r = [], 16
while ws.cell(r, 16).value not in (None, ""):
    sheet_moves.append(str(ws.cell(r, 16).value).strip()); r += 1
ck("every movement on MASTER is one the engine counts",
   all(m in engine.MV for m in sheet_moves),
   [m for m in sheet_moves if m not in engine.MV])
ck("every movement on MASTER has a label",
   all(m in L.MOVES for m in sheet_moves),
   [m for m in sheet_moves if m not in L.MOVES])
ck("every movement the form offers is on MASTER",
   set(entry_ui.NEEDS) <= set(sheet_moves),
   sorted(set(entry_ui.NEEDS) - set(sheet_moves)))
ck("no retired movement is on MASTER",
   not (set(RETIRED) & set(sheet_moves)), sorted(set(RETIRED) & set(sheet_moves)))

print("=== C. NOT RECEIVED IS COUNTED ===")
s,m,c,cfg,e = engine.load(qa_book.book())
st = engine.stock_by_item(s, m, cfg["as_of"])
ck("the Customs column exists", "Customs" in st.columns, list(st.columns)[:8])
nr = float(m[m["Movement"]=="Not received"]["Qty"].sum()) if len(m) else 0
ck("what is recorded as Not received reaches the Customs column",
   abs(float(st["Customs"].sum()) - nr) < 0.001,
   f"{st['Customs'].sum()} vs {nr}")
ck("shipped equals received plus not received on every line",
   ((st["Shipped Qty"] - st["Received"] - st["Customs"]).abs() < 0.001).all(),
   int((st["Shipped Qty"]-st["Received"]-st["Customs"]).abs().gt(0.001).sum()))

print("=== D. DELIVERED IS DERIVED, NOT RECORDED ===")
cl = engine.clearance_by_shipment(s, m, cfg["as_of"], cfg)
ck("clearance still reports a delivered figure", "Delivered" in cl.columns)
ck("it is what the courier took less what came back",
   all(abs(r.Delivered - max(r.ToCourier - r.Returned, 0)) < 0.001
       for r in cl.itertuples()),
   cl[["ToCourier","Returned","Delivered"]].head(3).to_dict("records"))
ck("no movement of that name exists in the sheet",
   "Delivered" not in set(m["Movement"].dropna()), sorted(set(m["Movement"].dropna())))

print("=== E. THE GUIDE MATCHES ===")
app = open("app.py").read()
for name in RETIRED + ["OUT Delivered"]:
    ck(f"the guide does not mention '{name}'", name not in app, "")
ck("the guide names Not received", "Not received" in app)

print()
for l in F: print(l)
print(f"\n{len(P)} passed, {len(F)} failed")
sys.exit(1 if F else 0)
