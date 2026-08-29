# -*- coding: utf-8 -*-
"""
Clear the test data and start clean.

    python3 reset_data.py            show what would go, write nothing
    python3 reset_data.py --apply    do it

Empties SHIPMENTS, MOVES and COUNT. MASTER is never touched - your items,
markets, couriers, users, reasons and settings all stay exactly as they are.

A copy of the workbook is saved beside this script first, so nothing is
unrecoverable.
"""
import sys, io, os
import datetime as dt
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter as CL

HEADER, FIRST = 6, 7
CLEAR = {"SHIPMENTS": "tblShipment", "MOVES": "tblMoves",
         "COUNT": "tblCount", "DISPATCH": "tblDispatch"}


def count_rows(wb, sheet):
    ws = wb[sheet]
    return sum(1 for r in range(FIRST, ws.max_row + 1)
               if ws.cell(r, 1).value not in (None, ""))


def clear(data: bytes):
    """Empty the data sheets, keep every heading, table and column."""
    wb = openpyxl.load_workbook(io.BytesIO(data))
    before = {}
    for sheet, tbl in CLEAR.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        before[sheet] = count_rows(wb, sheet)
        last_col = ws.max_column
        # delete the rows rather than blanking them. Blank cells still count
        # towards the sheet's size, so the next entry would land far below the
        # table and fall outside it.
        if ws.max_row >= FIRST:
            ws.delete_rows(FIRST, ws.max_row - FIRST + 1)
        # a table needs at least one row under its header, so it keeps one
        # empty row rather than being deleted and rebuilt differently
        if tbl in ws.tables:
            del ws.tables[tbl]
        t = Table(displayName=tbl,
                  ref=f"A{HEADER}:{CL(last_col)}{FIRST}")
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9",
                                          showRowStripes=True)
        ws.add_table(t)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), before


def main():
    import sharepoint_loader as sp, engine
    apply = "--apply" in sys.argv
    buf, meta = sp.fetch_workbook()
    data = buf.getvalue()
    print(f"{meta['name']} · saved {meta['modified']} · {meta['size_kb']} KB\n")

    s, m, c, cfg, e = engine.load(io.BytesIO(data))
    st = engine.stock_by_item(s, m, cfg["as_of"])
    print("NOW")
    print(f"  {len(s)} shipment lines · {len(m)} movements · {len(c)} counts")
    print(f"  {float(st['Store'].sum()):,.0f} boxes in store")
    print(f"  shipments: {', '.join(sorted(set(s['Shipment ID'])))}")

    new, before = clear(data)
    print("\nWOULD CLEAR")
    for sheet, n in before.items():
        print(f"  {sheet:<12}{n:>5} rows")

    print("\nWOULD KEEP  (MASTER is not touched)")
    print(f"  {len(cfg.get('item_names') or {})} items")
    print(f"  markets   {cfg.get('markets')}")
    print(f"  couriers  {cfg.get('couriers_by_market')}")
    print(f"  users     {sorted((cfg.get('users') or {}))}")
    print(f"  reasons   {len(cfg.get('reasons') or [])}")
    print(f"  settings  clear_target={cfg.get('clear_target')} "
          f"loss_target={cfg.get('loss_target')}")

    s2, m2, c2, cfg2, e2 = engine.load(io.BytesIO(new))
    print("\nAFTER")
    print(f"  {len(s2)} shipment lines · {len(m2)} movements · {len(c2)} counts")
    print(f"  {len(cfg2.get('item_names') or {})} items still on MASTER")
    print(f"  {len(e2)} entry errors")
    if len(e2):
        print("  refusing - the workbook would not be clean")
        return 1
    if len(cfg2.get("item_names") or {}) != len(cfg.get("item_names") or {}):
        print("  refusing - MASTER lost something")
        return 1

    if not apply:
        print("\nNothing was written. Run again with --apply to clear it.")
        return 0

    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M")
    backup = f"backup_{stamp}_{meta['name']}"
    open(backup, "wb").write(data)
    print(f"\nBacked up to {backup} ({len(data)/1024:,.0f} KB)")
    sp.upload_workbook(new, etag=meta.get("etag"))
    print("Cleared and saved to SharePoint.")
    print("\nThe sheet is now empty of data and ready for your history.")
    print("When you have entered it, run:  python3 validate.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
