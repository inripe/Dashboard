# -*- coding: utf-8 -*-
"""
Add the items the history needs but MASTER does not have.

    python3 add_items.py                    show what would be added
    python3 add_items.py --apply            add them
    python3 add_items.py uae.csv qatar.csv  work out the list from the files

With no files given it uses the list below, which came from reading all three
history files. Codes follow your own pattern: EG-FG-<three letters>-<number>-
<01 local, 02 export>-01-01.

Nothing is written without --apply, and the codes are shown first so they can
be checked before they become permanent.
"""
import sys, io, re
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

HEADER, FIRST, COL = 15, 16, 2      # tblItems on MASTER, column B

# name, three-letter stem, local or export. Everything sold in UAE, Qatar and
# KSA is imported, so 02. Only the Yemeni mango is bought locally.
WANTED = [
    ("Watermelon",         "WTM", "02"),
    ("Peach Sugary",       "PCH", "02"),
    ("Apricot Amar",       "APR", "02"),
    ("Peach Baladi",       "PCH", "02"),
    ("Berries Omani",      "BER", "02"),
    ("Grapes White",       "GRP", "02"),
    # Qatar Shopify already carries this one as QA-FG-MNG-16-01-01-01, so it
    # keeps that number rather than getting a new one
    ("Mango Timor Yemeni", "MNG", "01", "EG-FG-MNG-16-01-01-01"),
    ("Plum Hollywood",     "PLM", "02"),
    ("Golden Berry",       "GLD", "02"),
    ("Bashmala",           "BSH", "02"),
    ("Apple Baladi",       "APL", "02"),
]


def existing(wb):
    ms = wb["MASTER"]
    out, r = {}, FIRST
    while ms.cell(r, COL).value not in (None, ""):
        out[str(ms.cell(r, COL).value).strip()] = \
            str(ms.cell(r, COL + 1).value or "").strip()
        r += 1
    return out, r


def next_code(stem, variant, taken):
    """The next free number for that stem, so nothing collides."""
    n = 1
    while f"EG-FG-{stem}-{n:02d}-{variant}-01-01" in taken:
        n += 1
    return f"EG-FG-{stem}-{n:02d}-{variant}-01-01"


def plan(data: bytes, wanted=None):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    have, next_row = existing(wb)
    taken = set(have.values())
    rows = []
    for spec in (wanted or WANTED):
        name, stem, variant = spec[0], spec[1], spec[2]
        fixed = spec[3] if len(spec) > 3 else None
        if name in have:
            continue
        code = fixed if fixed and fixed not in taken \
            else next_code(stem, variant, taken)
        taken.add(code)
        rows.append({"name": name, "code": code,
                     "kind": "local" if variant == "01" else "export"})
    return wb, next_row, have, rows


def apply(wb, next_row, rows):
    ms = wb["MASTER"]
    th = Side("thin", color="BFBFBF"); box = Border(th, th, th, th)
    for i, p in enumerate(rows):
        for j, v in enumerate((p["name"], p["code"], "Yes")):
            c = ms.cell(next_row + i, COL + j)
            c.value = v
            c.font = Font("Arial", 9); c.border = box
            c.fill = PatternFill("solid", fgColor="DDEBF7")
            c.alignment = Alignment("center")
    last = next_row + len(rows) - 1
    if "tblItems" in ms.tables:
        del ms.tables["tblItems"]
    t = Table(displayName="tblItems",
              ref=f"{CL(COL)}{HEADER}:{CL(COL+2)}{last}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9",
                                      showRowStripes=True)
    ms.add_table(t)
    out = io.BytesIO(); wb.save(out)
    return out.getvalue()


def main():
    import sharepoint_loader as sp, engine
    files = [a for a in sys.argv[1:] if a.endswith(".csv")]
    do = "--apply" in sys.argv
    buf, meta = sp.fetch_workbook()
    data = buf.getvalue()
    print(f"{meta['name']} · saved {meta['modified']}\n")

    wanted = None
    if files:
        import import_history as ih
        cfg = engine.load(io.BytesIO(data))[3]
        names = list((cfg.get("item_names") or {}).values())
        seen = {}
        for f in files:
            rows, unknown, _, _d = ih.read(f, names)
            for u in unknown:
                seen.setdefault(u, 0)
                seen[u] += sum(r["received"] for r in rows
                               if r["their_item"] == u)
        known = {n for n, _, _ in WANTED}
        wanted = [w for w in WANTED if w[0] in known]
        for u in seen:
            if u not in known and u not in [w[0] for w in wanted]:
                stem = re.sub(r"[^A-Z]", "", u.upper())[:3] or "ITM"
                wanted.append((u, stem, "02"))
        print(f"  read {len(files)} file(s): {len(seen)} items missing\n")

    wb, next_row, have, rows = plan(data, wanted)
    print(f"MASTER has {len(have)} items")
    if not rows:
        print("  Nothing to add. Every item the history needs is there.")
        return 0
    print(f"\nWOULD ADD  ({len(rows)})")
    print(f"  {'ITEM':<24}{'CODE':<26}KIND")
    for p in rows:
        print(f"  {p['name']:<24}{p['code']:<26}{p['kind']}")
    print("\n  Check these codes read right before applying. They become "
          "permanent.")

    new = apply(wb, next_row, rows)
    s, m, c, cfg, e = engine.load(io.BytesIO(new))
    after = len(cfg.get("item_names") or {})
    print(f"\nAFTER: {after} items on MASTER · {len(e)} entry errors")
    if after != len(have) + len(rows):
        print("  refusing - the count does not add up")
        return 1
    if len(e):
        print("  refusing - it would not be clean")
        return 1
    if not do:
        print("\nNothing was written. Run again with --apply to add them.")
        return 0
    sp.upload_workbook(new, etag=meta.get("etag"))
    print("\nSaved to SharePoint.")
    print("Now run:  python3 import_history.py uae.csv")
    return 0


if __name__ == "__main__":
    sys.exit(main())
