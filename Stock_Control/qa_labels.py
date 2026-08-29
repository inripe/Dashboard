# -*- coding: utf-8 -*-
"""Direction prefix and Arabic on every movement a worker can pick."""
import sys, qa_book, labels as L, entry_ui, engine, qa_book
P,F=[],[]
def ck(n,ok,note=""):
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}  {note}")

print("=== A. EVERY MOVEMENT IS COVERED ===")
s,m,c,cfg,e=engine.load(qa_book.book())
import openpyxl
ws=openpyxl.load_workbook(qa_book.book())["MASTER"]
sheet_moves=[]
r=16
while ws.cell(r,16).value not in (None,""):
    sheet_moves.append(str(ws.cell(r,16).value).strip()); r+=1
missing=[x for x in sheet_moves if x not in L.MOVES]
ck("every movement in the sheet has a label", not missing, missing)
ck("every label has arabic", all(L.MOVES[k][1] for k in L.MOVES),
   [k for k in L.MOVES if not L.MOVES[k][1]])

print("=== B. DIRECTION MATCHES THE SHEET ===")
r=16
bad=[]
while ws.cell(r,16).value not in (None,""):
    name=str(ws.cell(r,16).value).strip(); d=str(ws.cell(r,18).value).strip()
    want = d if d in ("IN","OUT") else ""
    if L.direction(name) != want: bad.append((name, L.direction(name), want))
    r+=1
ck("in / out agrees with the sheet", not bad, bad)
ck("received is IN", L.direction("Received")=="IN")
ck("scrap is OUT", L.direction("Scrap")=="OUT")
ck("returned is IN", L.direction("Returned")=="IN")
ck("orders assigned has no direction", L.direction("Orders Assigned")=="")

print("=== C. THE LABEL LEADS WITH THE DIRECTION ===")
for mv in entry_ui.WORKER_MOVES:
    lab=L.move(mv)
    d=L.direction(mv)
    ck(f"{mv}: direction first", lab.split()[1]==d if d else True, lab)
    ck(f"{mv}: arabic present", L.MOVES[mv][1] in lab, lab)
    ck(f"{mv}: english present", mv in lab)
ck("an arrow shows the direction", "\u2193" in L.move("Received")
   and "\u2191" in L.move("Scrap"))

print("=== D. UI STRINGS ===")
for k in ["How many boxes?","Which shipment?","Save","Void","Today","Why?"]:
    ck(f"{k} is translated", L.UI.get(k) not in (None,""), L.t(k))
ck("an unknown string falls back to english", L.t("Nonexistent")=="Nonexistent")
ck("arabic can be switched off", L.move("Received", arabic=False)=="\u2193 IN  Received",
   L.move("Received", arabic=False))

print("=== E. THE CONFIRMATION SHOWS BOTH ===")
MKT = s["Market"].dropna().iloc[0]
USER = qa_book.entry_user(cfg, MKT) or qa_book.entry_user(cfg) or "manual"
sid=s[s.Market==MKT]["Shipment ID"].iloc[0]
it=s[s["Shipment ID"]==sid]["Item Name"].iloc[0]
t=entry_ui._sentence({"Shipment No":sid,"Movement":"Received","Item Name":it,"In":48},
                     "Received",MKT,USER)
ck("english sentence", "48 boxes" in t and it in t)
ck("arabic line", L.MOVES["Received"][1] in t)
ck("says which way stock moves", "stock goes up" in t, t[-90:])
t2=entry_ui._sentence({"Shipment No":sid,"Movement":"Scrap","Item Name":it,"Out":2,
                       "Reason":"Quality"},"Scrap",MKT,"x")
ck("an out movement says stock goes down", "stock goes down" in t2)
ck("arabic is right to left", 'direction:rtl' in t)

print("=== F. ARABIC ONLY WHERE IT BELONGS ===")
app=open("app.py").read()
ck("no bilingual helper anywhere in the dashboard", app.count("L.t(")==0, app.count("L.t("))
import re
# arabic is allowed on the entry sign-in, in the entry tab, and in the guide
# the whole guide, wherever it sits in the file
guide_from = app.index("def render_guide")
guide_to = app.index("def custom_panel")
entry_from = app.index("# ============================= RECORD")
outside=[]
for m in re.finditer(r"[\u0600-\u06FF]+", app):
    i=m.start(); line=app[:i].count("\n")+1
    ctx=app[max(0,i-260):i+80]
    in_guide = (guide_from <= i < guide_to) or "render_guide" in ctx
    in_entry = i >= entry_from
    in_modes = "MODE ==" in ctx or "What are you doing" in ctx
    if not (in_guide or in_entry or in_modes or 'tab == "entry"' in ctx):
        outside.append(line)
ck("arabic only on the sign-in, the entry tab and the guide", not outside, outside)
ck("the guide explains the flow in both languages",
   "\u0634\u062e\u0635\u0627\u0646" in app or "\u0627\u0644\u0642\u0627\u0639\u062f\u0629" in app)
ck("the dashboard tables stay english",
   'kpi(k[0],"Available to sell"' in app)
ui=open("entry_ui.py").read()
ck("the entry screen does use the translations", ui.count("L.t(")>5, ui.count("L.t("))
ck("and the movement labels", "L.move" in ui)
ck("the tiles are english", 'kpi(k[0],"Available to sell"' in app)
ck("labels.py covers every movement", len(L.MOVES)==9 and len(L.UI)>20,
   f"{len(L.MOVES)} moves, {len(L.UI)} strings")
ck("the retired movements are gone from the labels",
   not ({"Delivered","Orders Assigned","Courier Handover"} & set(L.MOVES)),
   sorted(L.MOVES))

print()
for l in F: print(l)
print(f"\n{len(P)} passed, {len(F)} failed")
sys.exit(1 if F else 0)
