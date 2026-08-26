# -*- coding: utf-8 -*-
"""
Find and void duplicate movements.

    python3 fix_duplicates.py            show them, write nothing
    python3 fix_duplicates.py --apply    void the later ones

Two rows are duplicates when the shipment, movement, item and quantity all
match and neither is already void. The earliest is kept; the rest are marked
Void, which leaves them in the file but out of every calculation.
"""
import sys, io
import openpyxl

HEADER, FIRST = 6, 7


def _cols(ws):
    return {ws.cell(HEADER, c).value: c for c in range(1, ws.max_column + 1)
            if ws.cell(HEADER, c).value}


def find(data: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["MOVES"]
    c = _cols(ws)
    seen, dupes = {}, []
    for r in range(FIRST, ws.max_row + 1):
        if ws.cell(r, c["Date"]).value in (None, ""):
            continue
        if str(ws.cell(r, c["Void"]).value or "").strip().lower() == "yes":
            continue
        key = (str(ws.cell(r, c["Shipment No"]).value).strip(),
               str(ws.cell(r, c["Movement"]).value).strip(),
               str(ws.cell(r, c["Item Name"]).value or "").strip(),
               ws.cell(r, c["In"]).value, ws.cell(r, c["Out"]).value)
        if key in seen:
            dupes.append({"row": r, "kept_row": seen[key], "key": key,
                          "id": ws.cell(r, c["Entry ID"]).value,
                          "by": ws.cell(r, c["Entered by"]).value,
                          "date": ws.cell(r, c["Date"]).value,
                          "reason": ws.cell(r, c["Reason"]).value})
        else:
            seen[key] = r
    return wb, ws, c, dupes


def void(data: bytes, note="duplicate"):
    wb, ws, c, dupes = find(data)
    for d in dupes:
        r = d["row"]
        ws.cell(r, c["Void"]).value = "Yes"
        old = ws.cell(r, c["Note"]).value or ""
        ws.cell(r, c["Note"]).value = (
            f"{old} · voided: {note} of row {d['kept_row']}").strip(" ·")
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), dupes


if __name__ == "__main__":
    import sharepoint_loader as sp, engine
    apply = "--apply" in sys.argv
    buf, meta = sp.fetch_workbook()
    data = buf.getvalue()
    print(f"{meta['name']} · saved {meta['modified']}\n")

    s, m, c0, cfg, e = engine.load(io.BytesIO(data))
    st = engine.stock_by_item(s, m, cfg["as_of"])
    print("BEFORE")
    print(f"  movements {len(m)} · stock {float(st['Store'].sum()):,.0f} boxes "
          f"· entry errors {len(e)}")

    new, dupes = void(data)
    print("\nDUPLICATES FOUND")
    if not dupes:
        print("  none - every movement is unique")
    for d in dupes:
        sid, mv, item, qin, qout = d["key"]
        print(f"  row {d['row']:<4} {sid}  {mv}  {item}  "
              f"{'in ' + str(qin) if qin else 'out ' + str(qout)}"
              f"   (same as row {d['kept_row']})")
        print(f"           entered by {d['by']}, reason {d['reason']}, "
              f"id {d['id'] or 'typed by hand'}")

    s2, m2, c2, cfg2, e2 = engine.load(io.BytesIO(new))
    st2 = engine.stock_by_item(s2, m2, cfg2["as_of"])
    print("\nAFTER")
    print(f"  movements {len(m2)} · stock {float(st2['Store'].sum()):,.0f} boxes "
          f"· entry errors {len(e2)}")
    print(f"  the rows stay in the file, marked Void, so nothing is lost")

    if not dupes:
        sys.exit(0)
    if not apply:
        print("\nNothing was written. Run again with --apply to void them.")
    else:
        sp.upload_workbook(new, etag=meta.get("etag"))
        print("\nSaved to SharePoint.")
