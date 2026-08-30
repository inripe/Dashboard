"""Which workbook the suites test, and who to sign in as. One place."""
import os, io, shutil

_SEEDED = {"path": None}


def book():
    """The file under test. qa_all.py sets QA_BOOK; otherwise the app's own.

    If it is empty, a seeded copy is written once and used from then on, so
    the app under test and the suites both see the same thing."""
    p = _book_path()
    raw = open(p, "rb").read()
    if not _is_empty(raw):
        if len(raw) < 60_000 or os.environ.get("QA_FULL"):
            return p
        # a big sheet is trimmed once, so the suites stay quick
        if not _SEEDED["path"]:
            import tempfile
            f = os.path.join(tempfile.gettempdir(), "qa_trimmed.xlsx")
            open(f, "wb").write(trimmed(raw))
            _SEEDED["path"] = f
            os.environ.setdefault("INRIPE_FILE", f)
        return _SEEDED["path"]
    if not _SEEDED["path"]:
        import tempfile
        f = os.path.join(tempfile.gettempdir(), "qa_seeded.xlsx")
        open(f, "wb").write(_seed(raw))
        _SEEDED["path"] = f
        os.environ.setdefault("INRIPE_FILE", f)
    return _SEEDED["path"]


def _book_path():
    for p in (os.environ.get("QA_BOOK"),
              "INRIPE_Stock_Entry_v1.xlsx",
              "INRIPE_Stock_Entry_v3.xlsx"):
        if p and os.path.exists(p):
            return p
    raise FileNotFoundError("No workbook to test. Put INRIPE_Stock_Entry_v1.xlsx "
                            "in this folder or pass --book.")

def trimmed(raw, keep=3):
    """A copy holding only the last few shipments.

    Once a season of history is in, the workbook is big enough that writing a
    test entry takes a second or more - and the suites write hundreds. They
    are testing the rules, not the volume, so they get the most recent
    shipments and nothing else. MASTER is untouched, so every item, market,
    courier and user is still there."""
    import openpyxl, io as _io
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter as CL
    wb = openpyxl.load_workbook(_io.BytesIO(raw))
    sh = wb["SHIPMENTS"]
    col = {sh.cell(6, i).value: i for i in range(1, sh.max_column + 1)
           if sh.cell(6, i).value}
    codes = []
    for r in range(7, sh.max_row + 1):
        v = sh.cell(r, col["Shipment No"]).value
        if v and str(v).strip() not in codes:
            codes.append(str(v).strip())
    if len(codes) <= keep:
        return raw
    live = set(codes[-keep:])
    for name, tbl, key in (("SHIPMENTS", "tblShipment", "Shipment No"),
                           ("MOVES", "tblMoves", "Shipment No")):
        ws = wb[name]
        c = {ws.cell(6, i).value: i for i in range(1, ws.max_column + 1)
             if ws.cell(6, i).value}
        drop = [r for r in range(7, ws.max_row + 1)
                if ws.cell(r, 1).value not in (None, "")
                and str(ws.cell(r, c[key]).value).strip() not in live]
        for r in reversed(drop):
            ws.delete_rows(r)
        last = max(ws.max_row, 7)
        if tbl in ws.tables:
            del ws.tables[tbl]
        t = Table(displayName=tbl, ref=f"A6:{CL(ws.max_column)}{last}")
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9",
                                          showRowStripes=True)
        ws.add_table(t)
    out = _io.BytesIO()
    wb.save(out)
    return out.getvalue()


def data():
    """The workbook under test, with a little sample data if it is empty.

    A cleared sheet is a perfectly good state for the live file - it is what
    you get before the history goes in - but the rule suites need something to
    push against. So they get a small seeded copy rather than every suite
    collapsing on an empty frame."""
    return open(book(), "rb").read()


def _is_empty(raw):
    import openpyxl, io as _io
    ws = openpyxl.load_workbook(_io.BytesIO(raw))["SHIPMENTS"]
    return all(ws.cell(r, 1).value in (None, "")
               for r in range(7, max(ws.max_row, 7) + 1))


def _seed(raw):
    """One shipment, a few items, a few movements. Enough for every rule to
    have something to refuse."""
    import io as _io, datetime as _dt
    import engine as _e, entry as _en
    s, m, c, cfg, e = _e.load(_io.BytesIO(raw))
    items = list((cfg.get("item_names") or {}).values())[:4]
    market = (cfg.get("markets") or ["Qatar"])[0]
    if not items:
        return raw
    # well back, so a suite writing a movement "yesterday" is never dated
    # before the shipment it belongs to
    arrival = _dt.date(2026, 8, 24)
    sid = _en.next_shipment_no(raw, market)
    rows = [{"Shipment No": sid, "Market": market, "Arrival Date": arrival,
             "Source": "Egypt", "Item Name": it, "Shipped Qty": q}
            for it, q in zip(items, (40, 25, 60, 15))]
    out, made = _en.append_shipment(raw, rows, "qa", market)
    moves = [{"Date": arrival, "Shipment No": made, "Movement": "Received",
              "Item Name": it, "In": q}
             for it, q in zip(items, (40, 25, 58, 15))]
    moves.append({"Date": arrival, "Shipment No": made,
                  "Movement": "Not received", "Item Name": items[2],
                  "Out": 2, "Reason": "Customs"})
    moves.append({"Date": arrival, "Shipment No": made, "Movement": "Scrap",
                  "Item Name": items[0], "Out": 3, "Reason": "Quality"})
    cour = ((cfg.get("couriers_by_market") or {}).get(market) or [None])[0]
    if cour:
        moves.append({"Date": _dt.date.today(), "Shipment No": made,
                      "Movement": "To Courier", "Item Name": items[1],
                      "Out": 10, "Courier": cour})
    for mv in moves:
        try:
            out, _ = _en.append_moves(out, [mv], "qa", market)
        except Exception:
            pass
    return out

def admin_user(cfg):
    """Whatever the admin is called on this sheet."""
    for u, r in (cfg.get("users") or {}).items():
        if str(r.get("role", "")).strip().lower() == "admin":
            return u
    return None

def entry_user(cfg, market=None):
    for u, r in (cfg.get("users") or {}).items():
        if str(r.get("role", "")).strip().lower() == "entry" \
                and (market is None or r.get("market") == market):
            return u
    return None

def old_style_copy(path="/tmp/qa_old.xlsx"):
    """A copy with the audit columns and users table stripped out, to test the
    upgrade against something genuinely old."""
    import openpyxl
    wb = openpyxl.load_workbook(book())
    ws = wb["MOVES"]
    cols = {ws.cell(6, c).value: c for c in range(1, ws.max_column + 1)
            if ws.cell(6, c).value}
    for name in ("Entered at", "Entered by", "Entry ID"):
        if name in cols:
            ws.delete_cols(cols[name])
            cols = {ws.cell(6, c).value: c for c in range(1, ws.max_column + 1)
                    if ws.cell(6, c).value}
    if "tblMoves" in ws.tables:
        del ws.tables["tblMoves"]
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter as CL
    last = ws.max_row
    t = Table(displayName="tblMoves", ref=f"A6:{CL(ws.max_column)}{last}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    ws.add_table(t)
    ms = wb["MASTER"]
    if "tblUsers" in ms.tables:
        del ms.tables["tblUsers"]
        for r in range(15, 30):
            for c in range(25, 29):
                ms.cell(r, c).value = None
    wb.save(path)
    return path


def workbench(raw=None, boxes=200):
    """A workbook the outward-movement suites can actually push against.

    They need a shipment with stock on the shelf. A real sheet at the end of a
    season often has almost none - UAE closed August on ten boxes - so rather
    than picking whatever is there and failing, one shipment is added with
    enough of a few items to test against. Nothing else is touched.

    Returns (bytes, shipment, item, market).
    """
    import io, datetime as dt
    import engine, entry, openpyxl
    raw = raw if raw is not None else data()
    s, m, c, cfg, e = engine.load(io.BytesIO(raw))
    mkt = (s["Market"].dropna().iloc[0] if len(s)
           else (cfg.get("markets") or ["Qatar"])[0])

    # something already on the shelf?
    if len(s):
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        for sid in list(dict.fromkeys(s[s.Market == mkt]["Shipment ID"]))[::-1]:
            store, _, _ = entry.stock_now(wb, sid)
            best = sorted(((v, k) for k, v in store.items() if v >= 20),
                          reverse=True)
            if best:
                return raw, sid, best[0][1], mkt

    items = sorted((cfg.get("item_names") or {}).values())[:4]
    if not items:
        raise RuntimeError("no items on MASTER to build a workbench with")
    # well back, so a suite writing a movement "yesterday" is never dated
    # before the shipment it belongs to
    day = dt.date.today() - dt.timedelta(days=30)
    sid = entry.next_shipment_no(raw, mkt)
    rows = [{"Shipment No": sid, "Market": mkt, "Arrival Date": day,
             "Source": "Egypt", "Item Name": it, "Shipped Qty": boxes}
            for it in items]
    out, made = entry.append_shipment(raw, rows, "qa", mkt)
    out, _ = entry.append_moves(out, [
        {"Date": day, "Shipment No": made, "Movement": "Received",
         "Item Name": it, "In": boxes} for it in items], "qa", mkt)
    return out, made, items[0], mkt
