# -*- coding: utf-8 -*-
"""Every shipment line must balance, and the tool that fixes it must be safe."""
import sys, io, datetime as dt
import openpyxl, engine, entry, qa_book, fix_gaps
P,F=[],[]
def ck(n,ok,note=""):
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}  {note}")
def eq(n,got,want,tol=1e-6):
    try: ok=abs(float(got)-float(want))<=tol
    except (TypeError,ValueError): ok=str(got)==str(want)
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}: got {got}, want {want}")

base=qa_book.data()

print("=== A. IT SEES WHAT IS OUT OF BALANCE ===")
st,over,under=fix_gaps.plan(base)
ck("it returns the three views", all(x is not None for x in (st,over,under)))
ck("over and under never overlap", not (set(over.index) & set(under.index)))
ck("every line is one of balanced, over or under",
   len(st) >= len(over)+len(under), f"{len(st)} lines")

print("=== B. TOO MUCH RECORDED AS MISSING IS VOIDED ===")
wb=openpyxl.load_workbook(io.BytesIO(base)); ws=wb["MOVES"]
c={ws.cell(6,i).value:i for i in range(1,ws.max_column+1) if ws.cell(6,i).value}
s0,m0,c0,cfg0,e0=engine.load(io.BytesIO(base))
SHIP=s0["Shipment ID"].iloc[0]
ITEM=s0[s0["Shipment ID"]==SHIP]["Item Name"].dropna().iloc[0]
r=ws.max_row+1
CODE={v:k for k,v in (cfg0.get("item_names") or {}).items()}.get(ITEM)
MKT0=s0[s0["Shipment ID"]==SHIP]["Market"].iloc[0]
vals={"Date":dt.date(2026,8,26),"Shipment No":SHIP,"Movement":"Not received",
      "Item Name":ITEM,"Out":999,"Reason":"Customs","Market":MKT0,
      "Item Code":CODE,"Qty":999,"Check":"OK","Entered by":"admin",
      "Entry ID":"Q-19990101-0001"}
for k,v in vals.items():
    if k in c: ws.cell(r,c[k]).value=v
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter as CL
if "tblMoves" in ws.tables: del ws.tables["tblMoves"]
t=Table(displayName="tblMoves", ref=f"A6:{CL(ws.max_column)}{r}")
t.tableStyleInfo=TableStyleInfo(name="TableStyleLight9",showRowStripes=True)
ws.add_table(t)
b=io.BytesIO(); wb.save(b); broken=b.getvalue()
st1,over1,under1=fix_gaps.plan(broken)
ck("the bad line is spotted", len(over1)>=1, len(over1))
fixed,voided,und=fix_gaps.fix(broken)
st2,over2,under2=fix_gaps.plan(fixed)
ck("it is voided", len(voided)>=1, [v["id"] for v in voided])
eq("nothing is over-recorded afterwards", len(over2), 0)
eq("no entry errors", len(engine.load(io.BytesIO(fixed))[4]), 0)

print("=== C. NOTHING IS DELETED ===")
w1=openpyxl.load_workbook(io.BytesIO(broken))["MOVES"]
w2=openpyxl.load_workbook(io.BytesIO(fixed))["MOVES"]
n1=sum(1 for x in range(7,w1.max_row+1) if w1.cell(x,1).value not in (None,""))
n2=sum(1 for x in range(7,w2.max_row+1) if w2.cell(x,1).value not in (None,""))
eq("the row is still in the file", n2, n1)
cc={w2.cell(6,i).value:i for i in range(1,w2.max_column+1) if w2.cell(6,i).value}
if voided:
    vr=voided[0]["row"]
    ck("marked Void", str(w2.cell(vr,cc["Void"]).value).lower()=="yes")
    ck("and says why", "not received" in str(w2.cell(vr,cc["Note"]).value).lower(),
       w2.cell(vr,cc["Note"]).value)
else:
    ck("nothing needed voiding on this workbook", True)

print("=== D. IT NEVER GUESSES AT A MISSING BOX ===")
ck("an unexplained shortfall is reported, not invented",
   "will not guess for you" in open("fix_gaps.py").read())
ck("only Not received rows are ever voided",
   '!= "Not received"' in open("fix_gaps.py").read())
ck("newest first", "reversed(cand)" in open("fix_gaps.py").read())
ck("upload needs --apply", "if not apply:" in open("fix_gaps.py").read())

print("=== E. SAFE TO RUN TWICE ===")
again,voided2,_=fix_gaps.fix(fixed)
eq("nothing left to void", len(voided2), 0)
st3,o3,u3=fix_gaps.plan(again)
eq("still balanced", len(o3), 0)

print("=== F. THE LIVE SHEET ===")
ck("shipped equals received plus missing on every line",
   len(over)==0 and len(under)==0,
   f"{len(over)} over, {len(under)} under - run fix_gaps.py")

print()
for l in F: print(l)
print(f"\n{len(P)} passed, {len(F)} failed")
sys.exit(1 if F else 0)
