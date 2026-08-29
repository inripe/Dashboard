# -*- coding: utf-8 -*-
"""
Repair a table that no longer covers its rows.

    python3 fix_tables.py            show what is wrong
    python3 fix_tables.py --apply    fix it

Excel tables have a range. If rows sit outside it, the app appends after the
last used row but the table still ends higher up - so the new entry lands
outside the table and is never read back. This resets each table's range to
cover exactly the rows that are there, and clears any ghost rows left behind
by an earlier clear.

Nothing is deleted except rows that are already completely empty.
"""
import sys, io
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter as CL

HEADER, FIRST = 6, 7
TABLES = {"SHIPMENTS": "tblShipment", "MOVES": "tblMoves",
          "COUNT": "tblCount", "DISPATCH": "tblDispatch"}


def look(wb, sheet, tbl):
    ws = wb[sheet]
    used = [r for r in range(FIRST, ws.max_row + 1)
            if any(ws.cell(r, c).value not in (None, "")
                   for c in range(1, ws.max_column + 1))]
    last_used = used[-1] if used else HEADER
    ref = None
    if tbl in ws.tables:
        t = ws.tables[tbl]
        ref = t if isinstance(t, str) else t.ref
    return ws, used, last_used, ref, ws.max_row


def repair(data: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    report = []
    for sheet, tbl in TABLES.items():
        if sheet not in wb.sheetnames:
            continue
        ws, used, last_used, ref, max_row = look(wb, sheet, tbl)
        # drop rows below the last real one: they are empty but still counted,
        # so the next entry would be written past the end of the table
        ghosts = max_row - last_used
        if ghosts > 0:
            ws.delete_rows(last_used + 1, ghosts)
        want = f"A{HEADER}:{CL(ws.max_column)}{max(last_used, FIRST)}"
        report.append({"sheet": sheet, "rows": len(used), "was": ref,
                       "now": want, "ghosts": max(ghosts, 0)})
        if tbl in ws.tables:
            del ws.tables[tbl]
        t = Table(displayName=tbl, ref=want)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9",
                                          showRowStripes=True)
        ws.add_table(t)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), report


def main():
    import sharepoint_loader as sp, engine
    apply = "--apply" in sys.argv
    buf, meta = sp.fetch_workbook()
    data = buf.getvalue()
    print(f"{meta['name']} · saved {meta['modified']}\n")

    s0, m0, c0, cfg0, e0 = engine.load(io.BytesIO(data))
    new, report = repair(data)
    print(f"  {'SHEET':<12}{'ROWS':>6}  {'TABLE WAS':<14}{'NOW':<14}GHOST ROWS")
    bad = 0
    for r in report:
        wrong = r["was"] != r["now"]
        bad += 1 if (wrong or r["ghosts"]) else 0
        print(f"  {r['sheet']:<12}{r['rows']:>6}  {str(r['was'] or '—'):<14}"
              f"{r['now']:<14}{r['ghosts'] if r['ghosts'] else ''}"
              + ("   <- was wrong" if wrong else ""))
    if not bad:
        print("\n  Every table already covers its rows. Nothing to do.")
        return 0

    s1, m1, c1, cfg1, e1 = engine.load(io.BytesIO(new))
    print(f"\n  before: {len(s0)} shipment lines · {len(m0)} movements · "
          f"{len(e0)} errors")
    print(f"  after : {len(s1)} shipment lines · {len(m1)} movements · "
          f"{len(e1)} errors")
    if len(s1) < len(s0) or len(m1) < len(m0):
        print("  refusing - rows would be lost")
        return 1
    if len(e1) > len(e0):
        print("  refusing - it would not be clean")
        return 1
    if not apply:
        print("\nNothing was written. Run again with --apply to fix it.")
        return 0
    sp.upload_workbook(new, etag=meta.get("etag"))
    print("\nFixed and saved to SharePoint.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
