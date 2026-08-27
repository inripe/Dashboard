# -*- coding: utf-8 -*-
"""
Bring the item list on MASTER into line with what the stores actually sell.

    python3 sync_items.py            show what would change, write nothing
    python3 sync_items.py --apply    add the missing items

It reads every product that appears on an order in any store, matches it
against tblItems, and reports three things: items that match, items on MASTER
that nobody has ordered, and products being sold that MASTER has never heard
of. Only the last group is added, and only with --apply.

A product with no SKU in Shopify cannot be given one here - the code has to be
invented, so the suggestion is printed for you to check before it is used.
"""
import sys, io, re
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter as CL

HEADER_ROW = 15          # tblItems header on MASTER
FIRST = HEADER_ROW + 1
COL = 2                  # column B


def from_shopify():
    """Every product seen on an order, with the SKU the line carried."""
    import shopify_reader as sr
    seen = {}
    for mkt in sr.configured_markets():
        try:
            orders, _ = sr.fetch_orders(mkt, limit_pages=10)
        except Exception as e:
            print(f"  {mkt}: could not read - {e}")
            continue
        for o in orders:
            for ln in o.get("lines", []):
                title = (ln.get("title") or "").strip()
                if not title:
                    continue
                d = seen.setdefault(title, {"skus": set(), "markets": set(),
                                            "lines": 0})
                if ln.get("sku"):
                    d["skus"].add(ln["sku"].strip())
                d["markets"].add(mkt)
                d["lines"] += 1
    return seen


def suggest(title, existing_codes):
    """A code in your own shape: EG-FG-XXX-nn-02-01-01."""
    words = re.findall(r"[A-Za-z]+", title.upper())
    stem = (words[0][:3] if words else "ITM")
    n = 1
    while f"EG-FG-{stem}-{n:02d}-02-01-01" in existing_codes:
        n += 1
    return f"EG-FG-{stem}-{n:02d}-02-01-01"


LOCAL, EXPORT = "-01-01-01", "-02-01-01"

# names Shopify still carries for the sake of old orders, but which nobody
# will order again. Adding them to the item list would only be clutter.
RETIRED = {"prickly pear", "mango timor"}


def variant_of(code):
    """Local and export are the same fruit in a different pack, so they are
    two items, not one. The fifth segment says which: 01 local, 02 export."""
    if code.endswith(LOCAL):
        return "local"
    if code.endswith(EXPORT):
        return "export"
    return None


def plan(data: bytes, seen: dict):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ms = wb["MASTER"]
    rows, r = [], FIRST
    while ms.cell(r, COL).value not in (None, ""):
        rows.append((r, str(ms.cell(r, COL).value).strip(),
                     str(ms.cell(r, COL + 1).value or "").strip(),
                     str(ms.cell(r, COL + 2).value or "").strip()))
        r += 1
    names = {n.lower(): (rr, n, c) for rr, n, c, a in rows}
    codes = {c for _, _, c, _ in rows if c}

    matched, unknown, unused, extra_codes = [], [], [], []
    for title, d in seen.items():
        if title.strip().lower() in RETIRED:
            continue
        hit = names.get(title.lower())
        if hit:
            matched.append((title, hit[2], d))
            # the same fruit sold locally carries its own code, and is its own
            # item: you would never pool local and export stock
            for sku in sorted(d["skus"]):
                if sku and sku not in codes:
                    extra_codes.append({"name": title, "code": sku,
                                        "variant": variant_of(sku),
                                        "markets": sorted(
                                            m for m in d["markets"])})
        else:
            unknown.append((title, d))
    ordered = {t.lower() for t in seen}
    for rr, n, c, a in rows:
        if n.lower() not in ordered:
            unused.append((n, c))

    plan_rows = []
    for e in sorted(extra_codes, key=lambda x: x["name"]):
        # export keeps the plain name; only the local one is marked, because
        # export is what the three importing markets sell every day
        label = " (EG)" if e["variant"] == "local" else ""
        plan_rows.append({"name": f"{e['name']}{label}", "code": e["code"],
                          "from_shopify": True, "markets": e["markets"],
                          "lines": 0, "second_code": True})
        codes.add(e["code"])
    for title, d in sorted(unknown, key=lambda x: -x[1]["lines"]):
        skus = sorted(d["skus"])
        if not skus:
            code = suggest(title, codes)
            codes.add(code)
            plan_rows.append({"name": title, "code": code,
                              "from_shopify": False,
                              "markets": sorted(d["markets"]),
                              "lines": d["lines"], "second_code": False})
            continue
        # a fruit sold both locally and for export is two items, so every
        # distinct code it carries becomes its own row
        many = len(skus) > 1 or any(variant_of(x) for x in skus)
        for sku in skus:
            v = variant_of(sku)
            label = (f"{title} (EG)" if v == "local" and len(skus) > 1
                     else title)
            if sku in codes:
                continue
            codes.add(sku)
            plan_rows.append({"name": label, "code": sku,
                              "from_shopify": True,
                              "markets": sorted(d["markets"]),
                              "lines": d["lines"], "second_code": False})
    return wb, ms, rows, matched, unused, plan_rows


def apply(wb, ms, rows, plan_rows):
    r = FIRST + len(rows)
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    th = Side("thin", color="BFBFBF"); box = Border(th, th, th, th)
    for i, p in enumerate(plan_rows):
        for j, v in enumerate((p["name"], p["code"], "Yes")):
            x = ms.cell(r + i, COL + j); x.value = v
            x.font = Font("Arial", 9); x.border = box
            x.fill = PatternFill("solid", fgColor="DDEBF7")
            x.alignment = Alignment("center")
    last = FIRST + len(rows) + len(plan_rows) - 1
    if "tblItems" in ms.tables:
        del ms.tables["tblItems"]
    t = Table(displayName="tblItems",
              ref=f"{CL(COL)}{HEADER_ROW}:{CL(COL+2)}{last}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9",
                                      showRowStripes=True)
    ms.add_table(t)
    out = io.BytesIO(); wb.save(out)
    return out.getvalue()


def main():
    import sharepoint_loader as sp, engine
    do = "--apply" in sys.argv
    buf, meta = sp.fetch_workbook()
    data = buf.getvalue()
    print(f"{meta['name']} · saved {meta['modified']}\n")

    print("reading what the stores are selling …")
    seen = from_shopify()
    print(f"  {len(seen)} distinct products across the orders read\n")

    wb, ms, rows, matched, unused, plan_rows = plan(data, seen)
    print(f"ALREADY ON MASTER  ({len(matched)})")
    for t, c, d in sorted(matched)[:40]:
        print(f"  {t[:32]:<34}{c}")

    if unused:
        print(f"\nON MASTER BUT NOBODY HAS ORDERED  ({len(unused)})")
        for n, c in unused:
            print(f"  {n[:32]:<34}{c}")
        print("  Left alone. They may simply be out of season.")

    if not plan_rows:
        print("\nNOTHING MISSING. Every product sold is on the item list.")
        return 0

    second = [p for p in plan_rows if p.get("second_code")]
    fresh = [p for p in plan_rows if not p.get("second_code")]
    if second:
        print(f"\nTHE SAME FRUIT, SOLD LOCALLY IN EGYPT  ({len(second)})")
        print( "  Egypt sells locally, the other three import. The fifth part of")
        print( "  the code says which - 01 local, 02 export - so these are their")
        print( "  own items with their own stock, not the same item twice.")
        print( "  Export keeps the plain name; the local one is marked (EG).")
        print(f"\n  {'ITEM':<32}{'CODE':<26}WHERE")
        for p in second:
            print(f"  {p['name'][:31]:<32}{p['code']:<26}"
                  f"{', '.join(p['markets'])}")
    if fresh:
        print(f"\nNOT ON MASTER AT ALL  ({len(fresh)})")
        print(f"  {'ITEM':<30}{'CODE':<26}{'FROM':<10}WHERE")
        for p in fresh:
            print(f"  {p['name'][:29]:<30}{p['code']:<26}"
                  f"{'shopify' if p['from_shopify'] else 'invented':<10}"
                  f"{', '.join(p['markets'])}")
    invented = [p for p in plan_rows if not p["from_shopify"]]
    if invented:
        print(f"\n  {len(invented)} of these have no SKU in Shopify, so the code")
        print( "  above was made up in your pattern. Check them before applying,")
        print( "  then put the same code in Shopify as the product's SKU.")

    if not do:
        print("\nNothing was written. Run again with --apply to add them.")
        return 0

    new = apply(wb, ms, rows, plan_rows)
    s, m, c, cfg, e = engine.load(io.BytesIO(new))
    print(f"\nAFTER: {len(cfg.get('item_names') or {})} items on MASTER, "
          f"{len(e)} entry errors")
    if len(e):
        print("  refusing to save - the workbook would not be clean")
        return 1
    sp.upload_workbook(new, etag=meta.get("etag"))
    print("Saved to SharePoint.")
    print("\nNow put each code above into Shopify as that product's SKU.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
