"""
Stock entry. The only part of this system that writes to the workbook.

Rules it will not break:
  * append only - a row is never edited or deleted once written
  * a correction is a new row, or a Yes in the Void column
  * every row carries an Entry ID, who wrote it and when
  * the Excel table range is widened on every append, or the sheet's own
    formulas and dropdowns would stop applying to new rows
"""
from __future__ import annotations
import io, re, datetime as dt
import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter as CL

HEADER_ROW = 6          # 1-based row holding the MOVES headers
FIRST_DATA = 7
DUP_MINUTES = 10        # an identical entry inside this window is a double tap

MARKET_TZ = {"Qatar": "Asia/Qatar", "UAE": "Asia/Dubai",
             "KSA": "Asia/Riyadh", "Egypt": "Africa/Cairo"}
MARKET_CODE = {"Qatar": "Q", "UAE": "U", "KSA": "K", "Egypt": "E"}


# ----------------------------------------------------------------- helpers
def market_now(market):
    tz = MARKET_TZ.get(market)
    t = pd.Timestamp.now(tz="UTC")
    return t.tz_convert(tz).tz_localize(None) if tz else t.tz_localize(None)


def _cols(ws):
    return {ws.cell(HEADER_ROW, c).value: c
            for c in range(1, ws.max_column + 1)
            if ws.cell(HEADER_ROW, c).value}


def next_entry_id(ws, market, when=None):
    """Q-20260825-0007. Sequential within a market and day, never reused."""
    when = when or market_now(market)
    prefix = f"{MARKET_CODE.get(market, 'X')}-{when:%Y%m%d}-"
    col = _cols(ws).get("Entry ID")
    top = 0
    if col:
        for r in range(FIRST_DATA, ws.max_row + 1):
            v = ws.cell(r, col).value
            if isinstance(v, str) and v.startswith(prefix):
                try:
                    top = max(top, int(v.rsplit("-", 1)[1]))
                except ValueError:
                    pass
    return f"{prefix}{top + 1:04d}"


def find_duplicate(ws, row, market, within_minutes=DUP_MINUTES):
    """A double tap: same movement, shipment, item and quantity, moments ago."""
    c = _cols(ws)
    need = ("Movement", "Shipment No", "Item Name", "In", "Out",
            "Entered at", "Entry ID", "Void")
    if not all(k in c for k in need):
        return None
    cutoff = market_now(market) - pd.Timedelta(minutes=within_minutes)
    for r in range(ws.max_row, FIRST_DATA - 1, -1):
        at = ws.cell(r, c["Entered at"]).value
        if not isinstance(at, dt.datetime) or pd.Timestamp(at) < cutoff:
            continue
        if str(ws.cell(r, c["Void"]).value or "").strip().lower() == "yes":
            continue
        same = (
            ws.cell(r, c["Movement"]).value == row.get("Movement")
            and ws.cell(r, c["Shipment No"]).value == row.get("Shipment No")
            and (ws.cell(r, c["Item Name"]).value or None) == (row.get("Item Name") or None)
            and (ws.cell(r, c["In"]).value or None) == (row.get("In") or None)
            and (ws.cell(r, c["Out"]).value or None) == (row.get("Out") or None))
        if same:
            return {"entry_id": ws.cell(r, c["Entry ID"]).value,
                    "at": pd.Timestamp(at), "row": r}
    return None


def _lookups(wb):
    """Read the reference tables once, so appended rows can be filled with real
    values. openpyxl writes formulas but cannot evaluate them, so a formula
    written here would read as blank until somebody opened the file in Excel."""
    ms = wb["MASTER"]
    def block(first_col, ncols, header_row=15):
        rows, r = [], header_row + 1
        while ms.cell(r, first_col).value not in (None, ""):
            rows.append([ms.cell(r, first_col + k).value for k in range(ncols)])
            r += 1
        return rows
    items = {str(a).strip(): str(b).strip() for a, b, *_ in block(2, 3)}
    moves = {}
    for row in block(16, 8):
        mv = str(row[0]).strip()
        moves[mv] = {"effect": row[1], "dir": str(row[2]).strip(),
                     "item": row[3], "courier": row[4], "orders": row[5],
                     "reason": row[6], "qty": row[7]}
    sh = wb["SHIPMENTS"]
    ship_market, ship_items, ship_arrival = {}, set(), {}
    r = 7
    while sh.cell(r, 1).value not in (None, ""):
        sid = str(sh.cell(r, 1).value).strip()
        ship_market[sid] = sh.cell(r, 2).value
        ship_items.add((sid, str(sh.cell(r, 5).value).strip()))
        d = sh.cell(r, 3).value
        if d is not None:
            ship_arrival[sid] = min(ship_arrival.get(sid, d), d)
        r += 1
    couriers = {str(c[0]).strip(): str(c[1]).strip() for c in block(9, 3)}
    return {"items": items, "moves": moves, "ship_market": ship_market,
            "ship_items": ship_items, "ship_arrival": ship_arrival,
            "couriers": couriers}


def stock_now(wb, shipment, item_name=None):
    """What this shipment holds right now, per item and with the courier.
    Read from MOVES, so it reflects everything recorded so far."""
    ws = wb["MOVES"]
    c = {ws.cell(HEADER_ROW, i).value: i for i in range(1, ws.max_column + 1)
         if ws.cell(HEADER_ROW, i).value}
    IN = {"Received", "Return to Saleable", "Count Adjustment - Add"}
    OUT = {"Scrap", "To Courier", "Return to Scrap", "Count Adjustment - Remove",
           "Not received"}
    store, with_courier, received = {}, 0.0, {}
    for r in range(FIRST_DATA, ws.max_row + 1):
        if ws.cell(r, c["Date"]).value in (None, ""):
            continue
        if str(ws.cell(r, c["Void"]).value or "").strip().lower() == "yes":
            continue
        if str(ws.cell(r, c["Shipment No"]).value).strip() != str(shipment).strip():
            continue
        mv = str(ws.cell(r, c["Movement"]).value or "").strip()
        it = str(ws.cell(r, c["Item Name"]).value or "").strip()
        q = float(ws.cell(r, c["In"]).value or 0) + float(ws.cell(r, c["Out"]).value or 0)
        if mv in IN and it:
            store[it] = store.get(it, 0) + q
            if mv == "Received":
                received[it] = received.get(it, 0) + q
        elif mv in OUT and it:
            store[it] = store.get(it, 0) - q
        if mv == "To Courier":
            with_courier += q
        elif mv == "Returned":
            with_courier -= q
    if item_name is not None:
        return store.get(str(item_name).strip(), 0.0)
    return store, with_courier, received


def _not_received(wb, shipment):
    """How many boxes are already recorded as never having arrived, per item.
    Without this a missing box can be claimed twice."""
    ws = wb["MOVES"]
    c = {ws.cell(HEADER_ROW, i).value: i for i in range(1, ws.max_column + 1)
         if ws.cell(HEADER_ROW, i).value}
    out = {}
    for r in range(FIRST_DATA, ws.max_row + 1):
        if ws.cell(r, c["Date"]).value in (None, ""):
            continue
        if str(ws.cell(r, c["Void"]).value or "").strip().lower() == "yes":
            continue
        if str(ws.cell(r, c["Movement"]).value or "").strip() != "Not received":
            continue
        if str(ws.cell(r, c["Shipment No"]).value).strip() != str(shipment).strip():
            continue
        it = str(ws.cell(r, c["Item Name"]).value or "").strip()
        out[it] = out.get(it, 0.0) + float(ws.cell(r, c["Out"]).value or 0)
    return out


def shipped_qty(lk_or_wb, shipment, item_name):
    """How many boxes the shipment says were sent for this item."""
    wb = lk_or_wb
    ws = wb["SHIPMENTS"]
    c = {ws.cell(HEADER_ROW, i).value: i for i in range(1, ws.max_column + 1)
         if ws.cell(HEADER_ROW, i).value}
    total = 0.0
    for r in range(FIRST_DATA, ws.max_row + 1):
        if str(ws.cell(r, c["Shipment No"]).value or "").strip() == str(shipment).strip() \
                and str(ws.cell(r, c["Item Name"]).value or "").strip() == str(item_name).strip():
            total += float(ws.cell(r, c["Shipped Qty"]).value or 0)
    return total


def check_quantities(row, wb):
    """Stops the mistakes a tired person makes at six in the morning.
    Returns "OK" or a sentence saying what is wrong and what the figure is."""
    mv = str(row.get("Movement") or "").strip()
    sid = str(row.get("Shipment No") or "").strip()
    item = str(row.get("Item Name") or "").strip()
    qty = float(row.get("In") or 0) + float(row.get("Out") or 0)
    if not sid or qty <= 0:
        return "OK"
    store, with_courier, received = stock_now(wb, sid)
    not_recv = _not_received(wb, sid)

    OUT_OF_STORE = {"Scrap", "To Courier", "Return to Scrap",
                    "Count Adjustment - Remove"}
    if mv in OUT_OF_STORE and item:
        have = store.get(item, 0.0)
        if qty > have:
            return (f"Only {have:,.0f} boxes of {item} are in the store for "
                    f"{sid}. You cannot take out {qty:,.0f}.")

    if mv == "Received" and item:
        sent = shipped_qty(wb, sid, item)
        already = received.get(item, 0.0)
        if sent and already + qty > sent:
            room = sent - already
            return (f"{sid} was sent {sent:,.0f} boxes of {item} and "
                    f"{already:,.0f} are already recorded. "
                    f"{'No more can be received' if room <= 0 else f'Only {room:,.0f} left to receive'}.")

    if mv == "Not received" and item:
        sent = shipped_qty(wb, sid, item)
        arrived = received.get(item, 0.0)
        claimed = not_recv.get(item, 0.0)
        room = sent - arrived - claimed
        if sent and qty > room:
            return (f"{sid} was sent {sent:,.0f} boxes of {item}, "
                    f"{arrived:,.0f} arrived"
                    + (f" and {claimed:,.0f} are already recorded as not received"
                       if claimed else "")
                    + ". "
                    + ("Everything is accounted for." if room <= 0
                       else f"At most {room:,.0f} can still be missing."))

    if mv == "Returned":
        if qty > with_courier:
            return (f"The courier is holding {with_courier:,.0f} boxes from "
                    f"{sid}. You cannot record {qty:,.0f}.")

    if mv == "Return to Saleable" or mv == "Return to Scrap":
        back = 0.0
        ws = wb["MOVES"]
        c = {ws.cell(HEADER_ROW, i).value: i for i in range(1, ws.max_column + 1)
             if ws.cell(HEADER_ROW, i).value}
        returned = split = 0.0
        for r in range(FIRST_DATA, ws.max_row + 1):
            if str(ws.cell(r, c["Shipment No"]).value or "").strip() != sid:
                continue
            if str(ws.cell(r, c["Void"]).value or "").strip().lower() == "yes":
                continue
            m2 = str(ws.cell(r, c["Movement"]).value or "").strip()
            q2 = float(ws.cell(r, c["In"]).value or 0) + float(ws.cell(r, c["Out"]).value or 0)
            if m2 == "Returned":
                returned += q2
            elif m2 in ("Return to Saleable", "Return to Scrap"):
                split += q2
        if split + qty > returned:
            return (f"{returned:,.0f} boxes came back on {sid} and "
                    f"{split:,.0f} have already been sorted. "
                    + ("Record the Returned movement first."
                       if returned <= 0 else
                       f"Only {max(returned-split,0):,.0f} are left to sort."))
    return "OK"


def validate(row, lk):
    """The same rules the sheet's Check column applies, run before writing.
    Returns "OK" or the reason. A row that fails is never written."""
    mv = str(row.get("Movement") or "").strip()
    sid = str(row.get("Shipment No") or "").strip()
    item = str(row.get("Item Name") or "").strip()
    spec = lk["moves"].get(mv)
    yes = lambda v: str(v).strip().lower() == "yes"
    qty = (row.get("In") or 0) + (row.get("Out") or 0)
    if not spec:                                    return "Unknown movement"
    if sid not in lk["ship_market"]:                return "Unknown shipment"
    if yes(spec["item"]) and not item:              return "Item needed"
    if not yes(spec["item"]) and item:              return "Leave Item blank"
    if spec["dir"] == "IN" and row.get("Out"):      return "Use the In column, not Out"
    if spec["dir"] == "OUT" and row.get("In"):      return "Use the Out column, not In"
    if yes(spec["qty"]) and qty <= 0:               return "Qty needed"
    if not yes(spec["qty"]) and qty > 0:            return "Leave In and Out blank"
    if yes(spec["courier"]) and not row.get("Courier"):  return "Courier needed"
    if not yes(spec["courier"]) and row.get("Courier"):  return "Leave Courier blank"
    if yes(spec["orders"]) and row.get("Orders") in (None, ""): return "Orders needed"
    if not yes(spec["orders"]) and row.get("Orders") not in (None, ""):
        return "Leave Orders blank"
    if yes(spec["reason"]) and not row.get("Reason"):    return "Reason needed"
    d, arr = row.get("Date"), lk["ship_arrival"].get(sid)
    if d is not None and arr is not None:
        dd = d.date() if hasattr(d, "date") else d
        aa = arr.date() if hasattr(arr, "date") else arr
        if dd < aa:                                 return "Date before arrival"
    if item and (sid, item) not in lk["ship_items"]:
        return "Item not in this shipment"
    cm = lk["couriers"].get(str(row.get("Courier") or "").strip())
    if row.get("Courier") and cm and cm != str(lk["ship_market"].get(sid) or "").strip():
        return "Courier not in this market"
    return "OK"


def _derived(name, row, lk):
    """The columns the sheet would have calculated."""
    mv = str(row.get("Movement") or "").strip()
    item = str(row.get("Item Name") or "").strip()
    sid = str(row.get("Shipment No") or "").strip()
    spec = lk["moves"].get(mv)
    if name == "Item Code":
        return lk["items"].get(item) if item else None
    if name == "Market":
        return lk["ship_market"].get(sid)
    if name == "Qty":
        return (row.get("In") or 0) + (row.get("Out") or 0) or None
    if name == "What to fill":
        if not spec:
            return "unknown movement"
        yes = lambda v: str(v).strip().lower() == "yes"
        head = {"IN": "IN   ", "OUT": "OUT   "}.get(spec["dir"], "--   ")
        need = "".join([
            "Item  " if yes(spec["item"]) else "",
            "In  " if (yes(spec["qty"]) and spec["dir"] == "IN") else "",
            "Out  " if (yes(spec["qty"]) and spec["dir"] == "OUT") else "",
            "Orders  " if yes(spec["orders"]) else "",
            "Courier  " if yes(spec["courier"]) else "",
            "Reason" if yes(spec["reason"]) else ""])
        return head + "fill: " + need
    if name == "Check":
        return "VOID" if str(row.get("Void") or "").lower() == "yes" \
            else validate(row, lk)
    return None


def _last_data_row(ws, key_col):
    r = ws.max_row
    while r >= FIRST_DATA and ws.cell(r, key_col).value in (None, ""):
        r -= 1
    return r


def append_moves(buf, rows, user, market):
    """Append rows to MOVES. Returns (bytes, [entry ids]).

    buf is the workbook as bytes. Nothing else in the file is touched.
    """
    wb = openpyxl.load_workbook(io.BytesIO(buf) if isinstance(buf, bytes) else buf)
    ws = wb["MOVES"]
    c = _cols(ws)
    lk = _lookups(wb)
    bad = []
    for i, r in enumerate(rows):
        why = validate(r, lk)
        if why == "OK":
            why = check_quantities(r, wb)
        if why != "OK":
            bad.append((i, why))
    if bad:
        raise ValueError("Refused: " + "; ".join(f"row {i+1} - {w}" for i, w in bad))
    when = market_now(market)
    start = _last_data_row(ws, c["Date"]) + 1
    ids = []
    for i, row in enumerate(rows):
        r = start + i
        eid = next_entry_id(ws, market, when)
        ids.append(eid)
        vals = dict(row)
        vals["Entry ID"] = eid
        vals["Entered by"] = user
        vals["Entered at"] = when.to_pydatetime()
        for name, col in c.items():
            d = _derived(name, vals, lk)
            if d is not None:
                ws.cell(r, col).value = d
            elif name in vals and vals[name] not in (None, ""):
                ws.cell(r, col).value = vals[name]
        if "Date" in c and isinstance(ws.cell(r, c["Date"]).value, (dt.date, dt.datetime)):
            ws.cell(r, c["Date"]).number_format = "dd-mmm-yy"
        if "Entered at" in c:
            ws.cell(r, c["Entered at"]).number_format = "dd-mmm-yy hh:mm"
    _widen(ws, "tblMoves", start + len(rows) - 1)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), ids


def void_entry(buf, entry_id, user, market):
    """Mark a row voided. The row stays; a note records who voided it."""
    wb = openpyxl.load_workbook(io.BytesIO(buf) if isinstance(buf, bytes) else buf)
    ws = wb["MOVES"]
    c = _cols(ws)
    if "Entry ID" not in c:
        raise ValueError("This workbook has no Entry ID column.")
    when = market_now(market)
    for r in range(FIRST_DATA, ws.max_row + 1):
        if ws.cell(r, c["Entry ID"]).value == entry_id:
            if str(ws.cell(r, c["Void"]).value or "").strip().lower() == "yes":
                raise ValueError(f"{entry_id} is already voided.")
            ws.cell(r, c["Void"]).value = "Yes"
            note = ws.cell(r, c["Note"]).value or ""
            ws.cell(r, c["Note"]).value = (
                f"{note} · voided by {user} {when:%d %b %H:%M}").strip(" ·")
            out = io.BytesIO()
            wb.save(out)
            return out.getvalue()
    raise ValueError(f"{entry_id} not found.")


def _widen(ws, table_name, last_row):
    """Excel tables have a fixed range. Without this, appended rows fall outside
    the table and its formulas and validation stop applying."""
    if table_name not in ws.tables:
        return
    old = ws.tables[table_name]
    ref = old.ref if not isinstance(old, str) else old
    first = ref.split(":")[0]
    last_col = "".join(ch for ch in ref.split(":")[1] if ch.isalpha())
    new_ref = f"{first}:{last_col}{max(last_row, HEADER_ROW + 1)}"
    if new_ref == ref:
        return
    style = getattr(old, "tableStyleInfo", None)
    del ws.tables[table_name]
    t = Table(displayName=table_name, ref=new_ref)
    t.tableStyleInfo = style or TableStyleInfo(name="TableStyleLight9",
                                               showRowStripes=True)
    ws.add_table(t)


# --------------------------------------------------------------- migration
def _ship_check(row, lk, seen):
    sid, item = str(row["Shipment No"] or "").strip(), str(row["Item Name"] or "").strip()
    if (sid, item) in seen:
        return "Item listed twice"
    seen.add((sid, item))
    if row.get("Market") and lk["ship_market"].get(sid) \
            and str(row["Market"]).strip() != str(lk["ship_market"][sid]).strip():
        return "Market differs on same shipment"
    if not row.get("Shipped Qty") or float(row["Shipped Qty"]) <= 0:
        return "Qty must be more than 0"
    return "OK"


def _count_check(row, lk, seen):
    sid, item = str(row["Shipment No"] or "").strip(), str(row["Item Name"] or "").strip()
    if (sid, item) not in lk["ship_items"]:
        return "Item not in this shipment"
    k = (row.get("Date"), sid, item)
    if k in seen:
        return "Counted twice"
    seen.add(k)
    if row.get("Physical Qty") is not None and float(row["Physical Qty"] or 0) < 0:
        return "Cannot be negative"
    return "OK"


def migrate_to_values(buf):
    """Replace every calculated formula with its value.

    openpyxl cannot evaluate formulas, and saving a workbook through it drops
    the values Excel had cached. With the app as the only writer that would
    blank the whole ledger on the first append, so the calculated columns are
    turned into plain values once, here.
    """
    wb = openpyxl.load_workbook(io.BytesIO(buf) if isinstance(buf, bytes) else buf)
    lk = _lookups(wb)
    changed = 0

    ws = wb["MOVES"]
    c = _cols(ws)
    for r in range(FIRST_DATA, ws.max_row + 1):
        if ws.cell(r, c["Date"]).value in (None, ""):
            continue
        row = {n: ws.cell(r, col).value for n, col in c.items()}
        for name in ("What to fill", "Item Code", "Market", "Qty", "Check"):
            if name in c:
                ws.cell(r, c[name]).value = _derived(name, row, lk)
                changed += 1
        if "Entered by" in c and ws.cell(r, c["Entered by"]).value in (None, ""):
            ws.cell(r, c["Entered by"]).value = "manual"

    sh = wb["SHIPMENTS"]
    sc = {sh.cell(HEADER_ROW, i).value: i for i in range(1, sh.max_column + 1)
          if sh.cell(HEADER_ROW, i).value}
    seen = set()
    for r in range(FIRST_DATA, sh.max_row + 1):
        if sh.cell(r, sc["Shipment No"]).value in (None, ""):
            continue
        row = {n: sh.cell(r, i).value for n, i in sc.items()}
        if "Item Code" in sc:
            sh.cell(r, sc["Item Code"]).value = lk["items"].get(
                str(row.get("Item Name") or "").strip())
            changed += 1
        if "Check" in sc:
            sh.cell(r, sc["Check"]).value = _ship_check(row, lk, seen)
            changed += 1

    if "COUNT" in wb.sheetnames:
        cn = wb["COUNT"]
        cc = {cn.cell(HEADER_ROW, i).value: i for i in range(1, cn.max_column + 1)
              if cn.cell(HEADER_ROW, i).value}
        seen2 = set()
        for r in range(FIRST_DATA, cn.max_row + 1):
            if cn.cell(r, cc["Date"]).value in (None, ""):
                continue
            row = {n: cn.cell(r, i).value for n, i in cc.items()}
            if "Item Code" in cc:
                cn.cell(r, cc["Item Code"]).value = lk["items"].get(
                    str(row.get("Item Name") or "").strip())
                changed += 1
            if "Check" in cc:
                cn.cell(r, cc["Check"]).value = _count_check(row, lk, seen2)
                changed += 1

    if "DISPATCH" in wb.sheetnames:
        dp = wb["DISPATCH"]
        dc = {dp.cell(HEADER_ROW, i).value: i for i in range(1, dp.max_column + 1)
              if dp.cell(HEADER_ROW, i).value}
        for r in range(FIRST_DATA, dp.max_row + 1):
            for name in ("Item Code", "Check"):
                if name in dc and isinstance(dp.cell(r, dc[name]).value, str) \
                        and dp.cell(r, dc[name]).value.startswith("="):
                    dp.cell(r, dc[name]).value = None
                    changed += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), changed


SHIPMENT_CODE = re.compile(r"^[A-Z]-\d{2}-\d{3,}$")
MARKET_LETTER = {"Qatar": "Q", "UAE": "U", "KSA": "K", "Egypt": "E"}


def next_shipment_no(buf, market=None, when=None):
    """Q-26-001 - market letter, two-digit year, serial within that pair.
    The market and year are in the code, so two markets can both have 001."""
    letter = MARKET_LETTER.get(str(market or "").strip(), "X")
    when = when or (market_now(market) if market else pd.Timestamp.now())
    yy = when.year % 100
    prefix = f"{letter}-{yy:02d}-"
    wb = openpyxl.load_workbook(io.BytesIO(buf) if isinstance(buf, bytes) else buf)
    ws = wb["SHIPMENTS"]
    top = 0
    for r in range(FIRST_DATA, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().startswith(prefix):
            try:
                top = max(top, int(v.strip().rsplit("-", 1)[1]))
            except (ValueError, IndexError):
                pass
    return f"{prefix}{top + 1:03d}"


def validate_shipment(row, lk, existing):
    """existing is the set of (shipment, item) already on the sheet."""
    sid = str(row.get("Shipment No") or "").strip()
    item = str(row.get("Item Name") or "").strip()
    if not sid:
        return "Shipment number needed"
    if not SHIPMENT_CODE.match(sid):
        return (f"'{sid}' is not a shipment number. They look like Q-26-001: "
                f"market letter, year, then three digits.")
    want = MARKET_LETTER.get(str(row.get("Market") or "").strip())
    if want and not sid.startswith(want + "-"):
        return (f"{sid} does not belong to {row.get('Market')}. "
                f"A {row.get('Market')} shipment starts with {want}-.")
    if not row.get("Market"):
        return "Market needed"
    if not row.get("Arrival Date"):
        return "Arrival date needed"
    if not item:
        return "Item needed"
    if item not in lk["items"]:
        return f"{item} is not on the item list"
    if (sid, item) in existing:
        return "Item listed twice on this shipment"
    q = row.get("Shipped Qty")
    if not q or float(q) <= 0:
        return "Qty must be more than 0"
    known = lk["ship_market"].get(sid)
    if known and str(known).strip() != str(row["Market"]).strip():
        return f"{sid} already belongs to {known}"
    return "OK"


def append_shipment(buf, rows, user, market):
    """Add shipment lines. Same rules as movements: append only, validated
    first, nothing written if any line fails."""
    wb = openpyxl.load_workbook(io.BytesIO(buf) if isinstance(buf, bytes) else buf)
    ws = wb["SHIPMENTS"]
    lk = _lookups(wb)
    c = {ws.cell(HEADER_ROW, i).value: i for i in range(1, ws.max_column + 1)
         if ws.cell(HEADER_ROW, i).value}
    existing = set(lk["ship_items"])
    bad = []
    for i, r in enumerate(rows):
        w = validate_shipment(r, lk, existing)
        if w != "OK":
            bad.append((i, w))
        else:
            existing.add((str(r["Shipment No"]).strip(),
                          str(r["Item Name"]).strip()))
    if bad:
        raise ValueError("Refused: "
                         + "; ".join(f"line {i+1} - {w}" for i, w in bad))
    start = _last_data_row(ws, c["Shipment No"]) + 1
    for i, row in enumerate(rows):
        r = start + i
        vals = dict(row)
        for name, col in c.items():
            if name == "Item Code":
                ws.cell(r, col).value = lk["items"].get(
                    str(vals.get("Item Name") or "").strip())
            elif name == "Check":
                ws.cell(r, col).value = "OK"
            elif name in vals and vals[name] not in (None, ""):
                ws.cell(r, col).value = vals[name]
        if "Arrival Date" in c:
            ws.cell(r, c["Arrival Date"]).number_format = "dd-mmm-yy"
    _widen(ws, "tblShipment", start + len(rows) - 1)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), str(rows[0]["Shipment No"])
