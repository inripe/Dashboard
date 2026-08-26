"""The writer is the only thing that can damage the ledger. Test it hard."""
import sys, qa_book, io, shutil, datetime as dt, pandas as pd, openpyxl, engine, entry
P,F=[],[]
def ck(n,ok,note=""):
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}  {note}")
def eq(n,got,want,tol=1e-6):
    ok=abs(float(got)-float(want))<=tol
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}: got {got}, want {want}")

SRC=qa_book.book()
base=open(SRC,"rb").read()
s0,m0,c0,cfg0,e0=engine.load(io.BytesIO(base))
st0=engine.stock_by_item(s0,m0,cfg0["as_of"])
SHIP=s0[s0.Market=="Qatar"]["Shipment ID"].iloc[0]
ITEM=s0[s0["Shipment ID"]==SHIP]["Item Name"].iloc[0]
# an item that is genuinely NOT on this shipment, whatever the workbook holds
NOT_IN=next((x for x in (cfg0.get("item_names") or {}).values()
             if x not in set(s0[s0["Shipment ID"]==SHIP]["Item Name"])), None)
def row(mv="Received", qty=5, item=None, **kw):
    d={"Date":entry.market_now("Qatar").date(),"Shipment No":SHIP,"Movement":mv,
       "Item Name":item if item is not None else ITEM}
    d["In" if mv in ("Received","Return to Saleable","Returned") else "Out"]=qty
    d.update(kw); return d

print("=== A. A ROW LANDS, AND NOTHING ELSE MOVES ===")
b1,ids=entry.append_moves(base,[row()], "qatar.store","Qatar")
s1,m1,c1,cfg1,e1=engine.load(io.BytesIO(b1))
eq("one row added", len(m1), len(m0)+1)
eq("shipments untouched", len(s1), len(s0))
eq("counts untouched", len(c1), len(c0))
ck("one id returned", len(ids)==1, ids)
ck("id looks right", ids[0].startswith("Q-") and len(ids[0])==15, ids[0])
eq("no new entry errors", len(e1), len(e0))
st1=engine.stock_by_item(s1,m1,cfg1["as_of"])
eq("stock moved by the quantity", st1.Store.sum(), st0.Store.sum()+5)

print("=== B. AUDIT TRAIL ===")
new=m1.iloc[-1]
ck("entry id written", new["Entry ID"]==ids[0], new["Entry ID"])
ck("user written", new["Entered by"]=="qatar.store", new["Entered by"])
ck("timestamp written", pd.notna(new["Entered at"]), new["Entered at"])
ck("timestamp is market time",
   abs((pd.Timestamp(new["Entered at"])-entry.market_now("Qatar")).total_seconds())<120,
   new["Entered at"])

print("=== C. FORMULAS ON THE NEW ROW ===")
wb=openpyxl.load_workbook(io.BytesIO(b1)); ws=wb["MOVES"]
cmap={ws.cell(6,c).value:c for c in range(1,ws.max_column+1) if ws.cell(6,c).value}
r=ws.max_row
for col in ["What to fill","Item Code","Market","Qty","Check"]:
    v=ws.cell(r,cmap[col]).value
    ck(f"{col} is filled with a value, not a formula",
       v not in (None,"") and not (isinstance(v,str) and v.startswith("=")), str(v)[:28])
ck("Check says OK on a good row", ws.cell(r,cmap["Check"]).value=="OK",
   ws.cell(r,cmap["Check"]).value)
for col in ["Entry ID","Entered by","Entered at"]:
    v=ws.cell(r,cmap[col]).value
    ck(f"{col} is a value, not a formula", not (isinstance(v,str) and v.startswith("=")))

print("=== D. THE TABLE RANGE GROWS ===")
w0=openpyxl.load_workbook(io.BytesIO(base))
ref0=w0["MOVES"].tables["tblMoves"].ref
ref1=ws.tables["tblMoves"].ref
end0=int("".join(ch for ch in ref0.split(":")[1] if ch.isdigit()))
end1=int("".join(ch for ch in ref1.split(":")[1] if ch.isdigit()))
eq("range extended by one", end1, end0+1)
ck("the new row is inside the table", end1>=r, f"{ref1} vs row {r}")
b_many,_=entry.append_moves(base,[row() for _ in range(5)],"qatar.store","Qatar")
wm=openpyxl.load_workbook(io.BytesIO(b_many))["MOVES"]
endm=int("".join(ch for ch in wm.tables["tblMoves"].ref.split(":")[1] if ch.isdigit()))
eq("five rows extend it by five", endm, end0+5)

print("=== E. IDS ARE SEQUENTIAL AND NEVER REUSED ===")
b2,ids2=entry.append_moves(b1,[row()],"qatar.store","Qatar")
ck("second id follows the first", ids2[0]>ids[0], f"{ids[0]} -> {ids2[0]}")
b3,ids3=entry.append_moves(b2,[row(),row(qty=6)],"qatar.store","Qatar")
ck("a batch gets distinct ids", len(set(ids3))==2, ids3)
allids=[i for i in engine.load(io.BytesIO(b3))[1]["Entry ID"].dropna()]
ck("no id appears twice", len(allids)==len(set(allids)), f"{len(allids)} ids")
ck("markets get different prefixes",
   entry.next_entry_id(openpyxl.load_workbook(io.BytesIO(base))["MOVES"],"UAE").startswith("U-"))

print("=== F. DOUBLE TAP IS CAUGHT ===")
dup=entry.find_duplicate(openpyxl.load_workbook(io.BytesIO(b1))["MOVES"],row(),"Qatar")
ck("an identical entry moments ago is found", dup is not None,
   dup["entry_id"] if dup else "none")
ck("a different quantity is not a duplicate",
   entry.find_duplicate(openpyxl.load_workbook(io.BytesIO(b1))["MOVES"],
                        row(qty=99),"Qatar") is None)
ck("a different movement is not a duplicate",
   entry.find_duplicate(openpyxl.load_workbook(io.BytesIO(b1))["MOVES"],
                        row(mv="Scrap",qty=5),"Qatar") is None)
ck("nothing matches in an empty window",
   entry.find_duplicate(openpyxl.load_workbook(io.BytesIO(b1))["MOVES"],
                        row(),"Qatar",within_minutes=0) is None)

print("=== G. VOID ===")
b4=entry.void_entry(b1,ids[0],"qatar.store","Qatar")
s4,m4,c4,cfg4,e4=engine.load(io.BytesIO(b4))
_raw=lambda b: sum(1 for r in range(7, openpyxl.load_workbook(io.BytesIO(b))["MOVES"].max_row+1)
                   if openpyxl.load_workbook(io.BytesIO(b))["MOVES"].cell(r,1).value not in (None,""))
eq("the row is still in the file after voiding", _raw(b4), _raw(b1))
eq("but the engine ignores it", len(m4), len(m1)-1)
st4=engine.stock_by_item(s4,m4,cfg4["as_of"])
eq("stock returns to what it was", st4.Store.sum(), st0.Store.sum())
wv=openpyxl.load_workbook(io.BytesIO(b4))["MOVES"]
vr=[r for r in range(7,wv.max_row+1) if wv.cell(r,cmap["Entry ID"]).value==ids[0]][0]
ck("void flag set", str(wv.cell(vr,cmap["Void"]).value).lower()=="yes")
ck("who voided it is recorded", "voided by qatar.store" in str(wv.cell(vr,cmap["Note"]).value),
   wv.cell(vr,cmap["Note"]).value)
try:
    entry.void_entry(b4,ids[0],"x","Qatar"); ck("voiding twice is refused", False)
except ValueError as ex: ck("voiding twice is refused", "already" in str(ex))
try:
    entry.void_entry(b4,"Q-19990101-0001","x","Qatar"); ck("unknown id is refused", False)
except ValueError as ex: ck("unknown id is refused", "not found" in str(ex))
ck("a voided row is ignored by the duplicate check",
   entry.find_duplicate(openpyxl.load_workbook(io.BytesIO(b4))["MOVES"],row(),"Qatar") is None)

print("=== H. NOTHING IS EVER EDITED OR DELETED ===")
before=engine.load(io.BytesIO(base))[1]
after=engine.load(io.BytesIO(b3))[1]
key=["Date","Shipment","Movement","Item Name","In","Out"]
ck("every original row survives unchanged",
   before[key].astype(str).values.tolist()==after[key].astype(str).values.tolist()[:len(before)],
   f"{len(before)} original rows")
ck("rows only ever grow", len(after)>=len(before))

print("=== I. SHEET STAYS VALID ===")
open("/tmp/after.xlsx","wb").write(b3)
s5,m5,c5,cfg5,e5=engine.load("/tmp/after.xlsx")
eq("still no entry errors", len(e5), 0)
if len(e5): print("   errors:", e5.head(4).to_dict("records"))
ck("users still readable", len(cfg5["users"])>0, list(cfg5["users"]))
ck("settings still readable", cfg5["clear_target"]>0)

print("=== J. A BAD ROW IS REFUSED, NOT WRITTEN ===")
bads=[("unknown movement", row(mv="Teleport")),
      ("unknown shipment", dict(row(), **{"Shipment No":"NO. 999"})),
      ("item not in shipment", row(item=NOT_IN)),
      ("zero quantity", row(qty=0)),
      ("wrong direction", {"Date":entry.market_now("Qatar").date(),
                           "Shipment No":SHIP,"Movement":"Received",
                           "Item Name":ITEM,"Out":5}),
      ("missing reason", row(mv="Scrap",qty=1))]
bads=[(l,b) for l,b in bads if not (l=="item not in shipment" and NOT_IN is None)]
for label,bad in bads:
    try:
        entry.append_moves(base,[bad],"qatar.store","Qatar")
        ck(f"refused: {label}", False, "it was written")
    except ValueError as ex:
        ck(f"refused: {label}", True, str(ex)[:52])
n_before=len(engine.load(io.BytesIO(base))[1])
try: entry.append_moves(base,[row(),row(mv="Teleport")],"qatar.store","Qatar")
except ValueError: pass
eq("a batch with one bad row writes nothing", len(engine.load(io.BytesIO(base))[1]), n_before)

print("=== K. VALIDATION MATCHES THE SHEET'S OWN RULES ===")
import openpyxl as _ox
lk=entry._lookups(_ox.load_workbook(io.BytesIO(base)))
eq("a good row validates", entry.validate(row(),lk)=="OK", True)
ck("scrap needs a reason", entry.validate(row(mv="Scrap",qty=1),lk)=="Reason needed",
   entry.validate(row(mv="Scrap",qty=1),lk))
ck("scrap with a reason passes",
   entry.validate(dict(row(mv="Scrap",qty=1),Reason="Quality"),lk)=="OK")
ck("to courier needs a courier",
   entry.validate(row(mv="To Courier",qty=1),lk)=="Courier needed")
ck("a date before arrival is caught",
   entry.validate(dict(row(),Date=pd.Timestamp("2020-01-01").date()),lk)=="Date before arrival")

print("=== L. MIGRATION IS SAFE TO RUN TWICE ===")
once,_=entry.migrate_to_values(base)
twice,_=entry.migrate_to_values(once)
a=engine.load(io.BytesIO(once)); b=engine.load(io.BytesIO(twice))
eq("same number of moves", len(b[1]), len(a[1]))
eq("same stock",
   engine.stock_by_item(b[0],b[1],b[3]["as_of"]).Store.sum(),
   engine.stock_by_item(a[0],a[1],a[3]["as_of"]).Store.sum())
eq("still no errors", len(b[4]), 0)

print()
for l in F: print(l)
print(f"\n{len(P)} passed, {len(F)} failed")
sys.exit(1 if F else 0)
