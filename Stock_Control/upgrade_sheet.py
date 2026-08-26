"""
Upgrade the live workbook in place, keeping every row it already holds.

    python3 upgrade_sheet.py            check what would change
    python3 upgrade_sheet.py --apply    do it

Adds the three audit columns to MOVES, adds the Users table to MASTER if it is
missing, and turns the calculated formulas into values. Nothing is deleted and
no row is touched.
"""
import sys, io
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL
import entry

HEADER, FIRST = 6, 7
AUDIT = ["Entry ID", "Entered by", "Entered at"]
DEFAULT_USERS = [["admin", "All", "Admin", "Yes"],
                 ["qatar.store", "Qatar", "Entry", "Yes"],
                 ["uae.store", "UAE", "Entry", "Yes"],
                 ["ksa.store", "KSA", "Entry", "Yes"],
                 ["egypt.store", "Egypt", "Entry", "Yes"]]
F = "Arial"
th = Side("thin", color="BFBFBF"); BOX = Border(th, th, th, th)


def _cols(ws):
    return {ws.cell(HEADER, c).value: c for c in range(1, ws.max_column + 1)
            if ws.cell(HEADER, c).value}


def _hdr(ws, r, c, text, fill):
    x = ws.cell(r, c); x.value = text
    x.font = Font(F, 9, bold=True, color="FFFFFF")
    x.fill = PatternFill("solid", fgColor=fill)
    x.alignment = Alignment("center", "center", wrap_text=True); x.border = BOX


def upgrade(data: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    notes = []

    ws = wb["MOVES"]
    c = _cols(ws)
    if not all(a in c for a in AUDIT):
        chk = c.get("Check", ws.max_column + 1)
        ws.insert_cols(chk, len(AUDIT))
        last = ws.max_row
        for j, name in enumerate(AUDIT):
            _hdr(ws, HEADER, chk + j, name, "808080")
            ws.column_dimensions[CL(chk + j)].width = [17, 13, 17][j]
        for r in range(FIRST, last + 1):
            if ws.cell(r, 1).value in (None, ""):
                continue
            for j in range(len(AUDIT)):
                x = ws.cell(r, chk + j)
                x.font = Font(F, 10, color="404040")
                x.fill = PatternFill("solid", fgColor="F2F2F2")
                x.border = BOX; x.alignment = Alignment("center")
                if j == 2:
                    x.number_format = "dd-mmm-yy hh:mm"
            ws.cell(r, chk + 1).value = "manual"
        if "tblMoves" in ws.tables:
            old = ws.tables["tblMoves"]
            ref = old.ref if not isinstance(old, str) else old
            end_row = "".join(ch for ch in ref.split(":")[1] if ch.isdigit())
            del ws.tables["tblMoves"]
            t = Table(displayName="tblMoves",
                      ref=f"A{HEADER}:{CL(ws.max_column)}{end_row}")
            t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9",
                                              showRowStripes=True)
            ws.add_table(t)
        notes.append("MOVES: added Entry ID, Entered by, Entered at")
    else:
        notes.append("MOVES: audit columns already there")

    ms = wb["MASTER"]
    if "tblUsers" not in ms.tables:
        col, R0 = 25, 15
        for j, h in enumerate(["User", "Market", "Role", "Active"]):
            _hdr(ms, R0, col + j, h, "2E75B6")
            ms.column_dimensions[CL(col + j)].width = [16, 10, 9, 8][j]
        for i, row in enumerate(DEFAULT_USERS):
            for j, v in enumerate(row):
                x = ms.cell(R0 + 1 + i, col + j); x.value = v
                x.font = Font(F, 9); x.border = BOX
                x.fill = PatternFill("solid", fgColor="DDEBF7")
                x.alignment = Alignment("center")
        t = Table(displayName="tblUsers",
                  ref=f"{CL(col)}{R0}:{CL(col+3)}{R0+len(DEFAULT_USERS)}")
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9",
                                          showRowStripes=True)
        ms.add_table(t)
        notes.append(f"MASTER: added the Users table with "
                     f"{len(DEFAULT_USERS)} rows - edit the names to suit")
    else:
        notes.append("MASTER: Users table already there")

    out = io.BytesIO(); wb.save(out)
    data2, changed = entry.migrate_to_values(out.getvalue())
    notes.append(f"turned {changed} formula cells into values")
    return data2, notes


def summarise(tag, data):
    import engine
    s, m, c, cfg, e = engine.load(io.BytesIO(data))
    print(f"  {tag}")
    print(f"    markets on MASTER : {cfg.get('markets')}")
    print(f"    couriers          : {cfg.get('couriers_by_market')}")
    print(f"    users             : {list(cfg.get('users') or {})}")
    print(f"    shipment lines    : {len(s)}  across "
          f"{sorted(set(s['Market'].dropna()))}")
    print(f"    movements         : {len(m)}")
    print(f"    entry errors      : {len(e)}")


if __name__ == "__main__":
    import sharepoint_loader as sp
    apply = "--apply" in sys.argv
    print("Reading the live workbook…")
    buf, meta = sp.fetch_workbook()
    data = buf.getvalue()
    print(f"  {meta['name']} · saved {meta['modified']} · "
          f"{meta['size_kb']} KB\n")
    print("BEFORE"); summarise("as it is now", data)
    new, notes = upgrade(data)
    print("\nWHAT THIS WOULD CHANGE")
    for n in notes:
        print("   ·", n)
    print("\nAFTER"); summarise("once upgraded", new)
    if not apply:
        print("\nNothing was written. Run again with --apply to save it.")
    else:
        sp.upload_workbook(new, etag=meta.get("etag"))
        print("\nSaved to SharePoint.")
