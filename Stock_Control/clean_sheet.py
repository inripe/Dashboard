# -*- coding: utf-8 -*-
"""
Remove shipment lines that never had a movement recorded against them.

    python3 clean_sheet.py            show what would go, write nothing
    python3 clean_sheet.py --apply    remove them

A shipment line with no movements is a declaration that never happened. It
inflates the shipped total and shows as an unexplained gap. Anything with even
one movement is left alone.
"""
import sys, io
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter as CL

HEADER, FIRST = 6, 7


def _cols(ws):
    return {ws.cell(HEADER, c).value: c for c in range(1, ws.max_column + 1)
            if ws.cell(HEADER, c).value}


def plan(data: bytes):
    """Which shipments have movements, and which do not."""
    wb = openpyxl.load_workbook(io.BytesIO(data))
    mv = wb["MOVES"]
    mc = _cols(mv)
    used = set()
    for r in range(FIRST, mv.max_row + 1):
        if mv.cell(r, mc["Date"]).value in (None, ""):
            continue
        if str(mv.cell(r, mc["Void"]).value or "").strip().lower() == "yes":
            continue
        used.add(str(mv.cell(r, mc["Shipment No"]).value).strip())

    ws = wb["SHIPMENTS"]
    sc = _cols(ws)
    keep, drop = [], []
    for r in range(FIRST, ws.max_row + 1):
        sid = ws.cell(r, sc["Shipment No"]).value
        if sid in (None, ""):
            continue
        row = {n: ws.cell(r, c).value for n, c in sc.items()}
        (keep if str(sid).strip() in used else drop).append(row)

    cnt = wb["COUNT"] if "COUNT" in wb.sheetnames else None
    orphan_counts = []
    if cnt is not None:
        cc = _cols(cnt)
        alive = {str(k["Shipment No"]).strip() for k in keep}
        for r in range(FIRST, cnt.max_row + 1):
            sid = cnt.cell(r, cc["Shipment No"]).value
            if sid in (None, ""):
                continue
            if str(sid).strip() not in alive:
                orphan_counts.append({n: cnt.cell(r, c).value
                                      for n, c in cc.items()})
    return wb, keep, drop, orphan_counts


def _rewrite(ws, table, rows):
    cols = _cols(ws)
    for r in range(ws.max_row, HEADER, -1):
        ws.delete_rows(r)
    for i, row in enumerate(rows):
        for name, c in cols.items():
            ws.cell(FIRST + i, c).value = row.get(name)
    if table in ws.tables:
        del ws.tables[table]
    end = max(FIRST + len(rows) - 1, HEADER + 1)
    t = Table(displayName=table, ref=f"A{HEADER}:{CL(ws.max_column)}{end}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    ws.add_table(t)


def clean(data: bytes):
    wb, keep, drop, orphan_counts = plan(data)
    _rewrite(wb["SHIPMENTS"], "tblShipment", keep)
    if orphan_counts and "COUNT" in wb.sheetnames:
        cnt = wb["COUNT"]
        cc = _cols(cnt)
        alive = {str(k["Shipment No"]).strip() for k in keep}
        rows = []
        for r in range(FIRST, cnt.max_row + 1):
            sid = cnt.cell(r, cc["Shipment No"]).value
            if sid in (None, ""):
                continue
            if str(sid).strip() in alive:
                rows.append({n: cnt.cell(r, c).value for n, c in cc.items()})
        _rewrite(cnt, "tblCount", rows)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), keep, drop, orphan_counts


if __name__ == "__main__":
    import sharepoint_loader as sp, engine
    apply = "--apply" in sys.argv
    buf, meta = sp.fetch_workbook()
    data = buf.getvalue()
    print(f"{meta['name']} · saved {meta['modified']}\n")

    s, m, c, cfg, e = engine.load(io.BytesIO(data))
    print("BEFORE")
    print(f"  shipment lines {len(s)} · movements {len(m)} · counts {len(c)}")
    print(f"  entry errors   {len(e)}")

    new, keep, drop, orphans = clean(data)
    print("\nWOULD REMOVE")
    if not drop:
        print("  nothing - every shipment line has movements against it")
    else:
        by = {}
        for d in drop:
            k = str(d["Shipment No"]).strip()
            by[k] = by.get(k, [0, 0])
            by[k][0] += 1
            by[k][1] += float(d.get("Shipped Qty") or 0)
        for k in sorted(by):
            print(f"  {k:<10} {by[k][0]:>3} lines   "
                  f"{by[k][1]:>8,.0f} boxes declared, none ever received")
        print(f"  {'total':<10} {len(drop):>3} lines   "
              f"{sum(v[1] for v in by.values()):>8,.0f} boxes")
    if orphans:
        print(f"\n  plus {len(orphans)} count rows that referred to them")

    s2, m2, c2, cfg2, e2 = engine.load(io.BytesIO(new))
    print("\nAFTER")
    print(f"  shipment lines {len(s2)} · movements {len(m2)} · counts {len(c2)}")
    print(f"  entry errors   {len(e2)}")
    print(f"  shipments left {sorted(set(s2['Shipment ID']))}")
    st = engine.stock_by_item(s2, m2, cfg2["as_of"])
    print(f"  stock          {float(st['Store'].sum()):,.0f} boxes  "
          f"(unchanged: movements are untouched)")

    if not apply:
        print("\nNothing was written. Run again with --apply to save it.")
    else:
        sp.upload_workbook(new, etag=meta.get("etag"))
        print("\nSaved to SharePoint.")
