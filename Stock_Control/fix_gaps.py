# -*- coding: utf-8 -*-
"""
Make every shipment line balance: shipped = received + not received.

    python3 fix_gaps.py            show what is wrong and what it would do
    python3 fix_gaps.py --apply    do it

Too much recorded as Not received - it voids the extra rows, newest first.
Too little - it tells you, because only you know why a box never arrived.
Nothing is ever deleted.
"""
import sys, io
import openpyxl

HEADER, FIRST = 6, 7


def _cols(ws):
    return {ws.cell(HEADER, c).value: c for c in range(1, ws.max_column + 1)
            if ws.cell(HEADER, c).value}


def plan(data: bytes):
    import engine
    s, m, c, cfg, e = engine.load(io.BytesIO(data))
    st = engine.stock_by_item(s, m, cfg["as_of"])
    nm = cfg.get("item_names") or {}
    st = st.copy()
    st["ItemName"] = st["Item"].map(lambda x: nm.get(x, x)) \
        if "ItemName" not in st.columns else st["ItemName"]
    st["Gap"] = st["Shipped Qty"] - st["Received"] - st["Customs"]
    over  = st[st["Gap"] < -0.001]     # more accounted for than was ever sent
    under = st[st["Gap"] >  0.001]     # boxes nobody has explained
    return st, over, under


def fix(data: bytes):
    """Void surplus Not-received rows, newest first, until each line balances."""
    st, over, under = plan(data)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["MOVES"]
    c = _cols(ws)
    voided = []
    for _, row in over.iterrows():
        surplus = -row["Gap"]
        cand = []
        for r in range(FIRST, ws.max_row + 1):
            if ws.cell(r, c["Date"]).value in (None, ""):
                continue
            if str(ws.cell(r, c["Void"]).value or "").strip().lower() == "yes":
                continue
            if str(ws.cell(r, c["Movement"]).value).strip() != "Not received":
                continue
            if str(ws.cell(r, c["Shipment No"]).value).strip() != str(row["Shipment"]).strip():
                continue
            if str(ws.cell(r, c["Item Name"]).value or "").strip() != str(row["ItemName"]).strip():
                continue
            cand.append((r, float(ws.cell(r, c["Out"]).value or 0)))
        for r, qty in reversed(cand):          # newest first
            if surplus <= 0.001:
                break
            ws.cell(r, c["Void"]).value = "Yes"
            note = ws.cell(r, c["Note"]).value or ""
            ws.cell(r, c["Note"]).value = (
                f"{note} · voided: more was recorded as not received than "
                f"was ever shipped").strip(" ·")
            voided.append({"row": r, "shipment": row["Shipment"],
                           "item": row["ItemName"], "qty": qty,
                           "id": ws.cell(r, c["Entry ID"]).value})
            surplus -= qty
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), voided, under


if __name__ == "__main__":
    import sharepoint_loader as sp, engine
    apply = "--apply" in sys.argv
    buf, meta = sp.fetch_workbook()
    data = buf.getvalue()
    print(f"{meta['name']} · saved {meta['modified']}\n")

    st, over, under = plan(data)
    print(f"{len(st)} shipment lines · {len(over)+len(under)} do not balance\n")
    if len(over):
        print("TOO MUCH RECORDED AS NOT RECEIVED")
        for _, r in over.iterrows():
            print(f"  {r['Shipment']}  {r['ItemName']}: shipped {r['Shipped Qty']:,.0f}, "
                  f"received {r['Received']:,.0f}, not received {r['Customs']:,.0f} "
                  f"- {-r['Gap']:,.0f} too many")
    if len(under):
        print("\nBOXES NOBODY HAS EXPLAINED")
        for _, r in under.iterrows():
            print(f"  {r['Shipment']}  {r['ItemName']}: shipped {r['Shipped Qty']:,.0f}, "
                  f"received {r['Received']:,.0f} - {r['Gap']:,.0f} missing")
        print("  Record these as Not received in the app, with a reason. "
              "This script will not guess for you.")
    if not len(over) and not len(under):
        print("  Everything balances. Nothing to do.")
        sys.exit(0)

    new, voided, _ = fix(data)
    if voided:
        print("\nWOULD VOID")
        for v in voided:
            print(f"  row {v['row']}  {v['shipment']}  {v['item']}  "
                  f"{v['qty']:,.0f} boxes  {v['id'] or 'typed by hand'}")
    st2, over2, under2 = plan(new)
    print(f"\nAFTER: {len(over2)+len(under2)} lines still out of balance"
          + (" - the ones nobody has explained" if len(under2) else ""))

    if not voided:
        sys.exit(0)
    if not apply:
        print("\nNothing was written. Run again with --apply to save it.")
    else:
        sp.upload_workbook(new, etag=meta.get("etag"))
        print("\nSaved to SharePoint.")
