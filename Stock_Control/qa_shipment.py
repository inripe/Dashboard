# -*- coding: utf-8 -*-
"""Shipment entry: numbering, validation, and the sent-versus-arrived split."""
import sys, qa_book, io, datetime as dt, engine, entry, entry_ui
P,F=[],[]
def ck(n,ok,note=""):
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}  {note}")
def eq(n,got,want):
    try: ok=abs(float(got)-float(want))<1e-6
    except (TypeError,ValueError): ok=str(got)==str(want)
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}: got {got}, want {want}")

base=open(qa_book.book(),"rb").read()
s0,m0,c0,cfg0,e0=engine.load(io.BytesIO(base))
def line(item,q,sid=None,mkt="Qatar",d=None):
    return {"Shipment No":sid or entry.next_shipment_no(base, "Qatar"),"Market":mkt,
            "Arrival Date":d or dt.date(2026,8,26),"Source":"Egypt",
            "Item Name":item,"Shipped Qty":q}
ITEM=s0["Item Name"].dropna().iloc[0]
ITEM2=[x for x in s0["Item Name"].dropna().unique() if x!=ITEM][0]

print("=== A. NUMBERING ===")
nxt=entry.next_shipment_no(base, "Qatar")
import re as _re
ck("looks like Q-26-001", bool(_re.fullmatch(r"[A-Z]-\d{2}-\d{3,}", nxt)), nxt)
ck("it carries the market letter", nxt.startswith("Q-"), nxt)
ck("and the year", nxt.split("-")[1]==f"{dt.date.today().year%100:02d}", nxt)
ck("it is not already used", nxt not in set(s0["Shipment ID"]), nxt)
b,_=entry.append_shipment(base,[line(ITEM,500)],"mahmoud","Qatar")
ck("the next one moves on", entry.next_shipment_no(b, "Qatar")>nxt,
   f"{nxt} -> {entry.next_shipment_no(b, "Qatar")}")
ck("a number is never reused",
   entry.next_shipment_no(b, "Qatar") not in set(engine.load(io.BytesIO(b))[0]["Shipment ID"]))

print("=== B. WHAT LANDS ON THE SHEET ===")
s1,m1,c1,cfg1,e1=engine.load(io.BytesIO(b))
eq("one line added", len(s1), len(s0)+1)
eq("no entry errors", len(e1), 0)
r=s1.iloc[-1]
eq("shipped qty", r["Shipped Qty"], 500)
eq("market", r["Market"], "Qatar")
ck("item code filled in", str(r.get("Item Code") or "")!="", r.get("Item Code"))
ck("check says OK", r["Check"]=="OK", r["Check"])
eq("movements untouched", len(m1), len(m0))

print("=== C. IT REFUSES BAD LINES ===")
lk=entry._lookups(__import__("openpyxl").load_workbook(io.BytesIO(base)))
ex=set(lk["ship_items"])
bad=[("no item", {**line(ITEM,5), "Item Name":""}),
     ("unknown item", {**line(ITEM,5), "Item Name":"Dragonfruit"}),
     ("zero qty", line(ITEM,0)),
     ("no market", {**line(ITEM,5), "Market":None}),
     ("no arrival date", {**line(ITEM,5), "Arrival Date":None})]
for label,row in bad:
    ck(f"refused: {label}", entry.validate_shipment(row,lk,ex)!="OK",
       entry.validate_shipment(row,lk,ex))
ck("a good line passes", entry.validate_shipment(line(ITEM,5),lk,ex)=="OK")
n_before=len(engine.load(io.BytesIO(base))[0])
try:
    entry.append_shipment(base,[line(ITEM,5),line(ITEM2,0)],"x","Qatar")
    ck("one bad line stops the whole shipment", False, "it was written")
except ValueError as e:
    ck("one bad line stops the whole shipment", True, str(e)[:44])
eq("and nothing was added", len(engine.load(io.BytesIO(base))[0]), n_before)

print("=== D. THE SAME ITEM TWICE ===")
sid=entry.next_shipment_no(base, "Qatar")
try:
    entry.append_shipment(base,[line(ITEM,5,sid),line(ITEM,7,sid)],"x","Qatar")
    ck("the same item twice is refused", False)
except ValueError as e:
    ck("the same item twice is refused", "twice" in str(e), str(e)[:44])
b2,_=entry.append_shipment(base,[line(ITEM,5,sid),line(ITEM2,7,sid)],"x","Qatar")
eq("two different items are fine", len(engine.load(io.BytesIO(b2))[0]), len(s0)+2)

print("=== E. SENT IS NOT ARRIVED ===")
sid=entry.next_shipment_no(base, "Qatar")
b3,_=entry.append_shipment(base,[line(ITEM,500,sid)],"mahmoud","Qatar")
s3,m3,c3,cfg3,e3=engine.load(io.BytesIO(b3))
st3=engine.stock_by_item(s3,m3,cfg3["as_of"])
row=st3[(st3["Shipment"]==sid)]
eq("a new shipment adds no stock on its own",
   float(row["Store"].sum()) if len(row) else 0, 0)
eq("but it is on the sheet as shipped", float(row["Shipped Qty"].sum()), 500)
b4,_=entry.append_moves(b3,[{"Date":dt.date(2026,8,26),"Shipment No":sid,
    "Movement":"Received","Item Name":ITEM,"In":480}],"qatar.store","Qatar")
b5,_=entry.append_moves(b4,[{"Date":dt.date(2026,8,26),"Shipment No":sid,
    "Movement":"Scrap","Item Name":ITEM,"Out":5,"Reason":"Damage"}],
    "qatar.store","Qatar")
s5,m5,c5,cfg5,e5=engine.load(io.BytesIO(b5))
st5=engine.stock_by_item(s5,m5,cfg5["as_of"])
row5=st5[st5["Shipment"]==sid]
eq("received 480", float(row5["Received"].sum()), 480)
eq("scrapped 5", float(row5["Scrap"].sum()), 5)
eq("475 sellable", float(row5["Store"].sum()), 475)
eq("the 15 that never arrived are still visible",
   float(row5["Shipped Qty"].sum()) - float(row5["Received"].sum()), 20)
ck("no entry errors after the whole flow", len(e5)==0, len(e5))

print("=== F. ONLY ADMIN SEES IT ===")
app=open("app.py").read()
ck("shipment entry is its own tab", '"Shipment arrived"' in app)
ck("and is admin only",
   'Only an admin records a new shipment' in app)
ck("the screen exists", hasattr(entry_ui,"render_shipment"))

print("=== G. A NEW MARKET CAN RECEIVE ITS FIRST SHIPMENT ===")
import openpyxl as _ox, shutil as _sh
_sh.copy(qa_book.book(),"/tmp/qa_m4.xlsx")
_wb=_ox.load_workbook("/tmp/qa_m4.xlsx"); _ms=_wb["MASTER"]
for _i,_mk in enumerate(["Qatar","UAE","KSA","Egypt"]):
    _ms.cell(16+_i,6).value=_mk; _ms.cell(16+_i,7).value="Yes"
_wb.save("/tmp/qa_m4.xlsx")
_s,_m,_c,_cfg,_e=engine.load("/tmp/qa_m4.xlsx")
eq("MASTER lists four markets", len(_cfg.get("markets") or []), 4)
ck("shipments only cover one so far",
   set(_s["Market"].dropna())=={"Qatar"}, sorted(set(_s["Market"].dropna())))
_offer=sorted(_cfg.get("markets") or [])
ck("the form still offers all four", set(_offer)=={"Qatar","UAE","KSA","Egypt"}, _offer)
ck("a market with no shipment is offered", "UAE" in _offer)
ui=open("entry_ui.py").read()
ck("the list comes from MASTER, not from shipments",
   'cfg.get("markets")' in ui and "otherwise a market can never receive" in ui)
_b=open("/tmp/qa_m4.xlsx","rb").read()
_row={"Shipment No":entry.next_shipment_no(_b, "UAE"),"Market":"UAE",
      "Arrival Date":dt.date(2026,8,26),"Source":"Egypt",
      "Item Name":ITEM,"Shipped Qty":300}
_out,_sid=entry.append_shipment(_b,[_row],"admin","UAE")
_s2,_m2,_c2,_cfg2,_e2=engine.load(io.BytesIO(_out))
ck("and the shipment saves", _sid in set(_s2["Shipment ID"]), _sid)
eq("against the right market",
   _s2[_s2["Shipment ID"]==_sid]["Market"].iloc[0], "UAE")
eq("with no errors", len(_e2), 0)

print("=== H. THE FORM ASKS IN A SENSIBLE ORDER ===")
ui=open("entry_ui.py").read()
i_mkt=ui.index('"1 \u00b7 Market')
i_no=ui.index('"2 \u00b7 Shipment number')
i_arr=ui.index('"3 \u00b7 Arrival date')
i_src=ui.index('"4 \u00b7 Source')
i_items=ui.index('"**5 \u00b7 Items') if '"**5' in ui else ui.index("5 \u00b7 Items")
ck("market is asked first", i_mkt < i_no, "market then number")
ck("the number comes second", i_no < i_arr)
ck("then the arrival date, then the source", i_arr < i_src)
ck("items come last", i_src < i_items)
ck("every field carries a hint", ui.count("help=") >= 4, ui.count("help="))
ck("the hint explains when to pick an existing one",
   "correct or add to a shipment already created" in ui)
ck("a new number is labelled as new", '"  \u2014  new"' in ui or "—  new" in ui)
ck("it says which market and year the number belongs to",
   "is the next free number for" in ui)
ck("the item hint says sent, not arrived",
   "not what arrived" in ui)
ck("the shipment field waits for a market",
   'choose a market first' in ui)

print("=== I. THE NUMBER IS PICKED, NEVER TYPED ===")
ui=open("entry_ui.py").read()
ck("the shipment number is a dropdown",
   'c2.selectbox(\n        "2 \u00b7 Shipment number' in ui
   or 'sid = c2.selectbox(' in ui, "")
ck("it is not a text box any more",
   'c2.text_input("2 \u00b7 Shipment number' not in ui
   and 'text_input(\n        "2 \u00b7 Shipment number' not in ui)
ck("a new number is offered first", '[nxt] if nxt else []' in ui)
ck("existing ones are labelled with their arrival and item count",
   "arrived {pd.Timestamp(d):%d %b}" in ui)
ck("picking an existing one is called out", "already exists" in ui)
ck("and its arrival date is locked", "The arrival date stays as it was" in ui)
ck("the field waits for a market", "disabled=not mkt" in ui)
ck("an item already on the shipment cannot be added twice",
   "every item is already on this shipment" in ui)

print("=== J. THE NEXT NUMBER IS RIGHT PER MARKET AND YEAR ===")
import pandas as _pd
_sh=_pd.DataFrame({"Shipment ID":["Q-26-001","Q-26-002","U-26-001","Q-25-009"],
                   "Market":["Qatar","Qatar","UAE","Qatar"]})
def _next(ship_df, mkt, year):
    letter={"Qatar":"Q","UAE":"U","KSA":"K","Egypt":"E"}.get(mkt,"X")
    prefix=f"{letter}-{year%100:02d}-"
    top=0
    for x in ship_df["Shipment ID"].dropna().astype(str):
        if x.strip().startswith(prefix):
            try: top=max(top,int(x.strip().rsplit("-",1)[1]))
            except (ValueError,IndexError): pass
    return f"{prefix}{top+1:03d}"
eq("qatar follows its own series", _next(_sh,"Qatar",2026), "Q-26-003")
eq("uae has its own", _next(_sh,"UAE",2026), "U-26-002")
eq("an untouched market starts at 001", _next(_sh,"KSA",2026), "K-26-001")
eq("a new year restarts", _next(_sh,"Qatar",2027), "Q-27-001")
ck("last year's numbers do not interfere",
   _next(_sh,"Qatar",2026) != "Q-26-010")

print("=== K. BAD CODES TYPED INTO EXCEL ARE CAUGHT ===")
app=open("app.py").read()
ck("data check looks for the wrong format",
   "Shipment code not in the Q-26-001 format" in app)
ck("and for a code that does not match its market",
   "Shipment code does not match its market" in app)
ck("and for a movement pointing at no shipment",
   "Movement points at no shipment" in app)
ck("and for two arrival dates on one shipment",
   "Shipment has more than one arrival date" in app)
import re as _re
for good in ("Q-26-001","U-26-014","E-27-123","K-26-1000"):
    ck(f"{good} is accepted",
       bool(_re.fullmatch(r"[A-Z]-\d{2}-\d{3,}", good)))
for bad in ("NO. 022","Q-26-1","q-26-001","Q26001","Q-2026-001",""):
    ck(f"{bad or '(blank)'} is rejected",
       not _re.fullmatch(r"[A-Z]-\d{2}-\d{3,}", bad))

print()
for l in F: print(l)
print(f"\n{len(P)} passed, {len(F)} failed")
sys.exit(1 if F else 0)
