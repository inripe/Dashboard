# -*- coding: utf-8 -*-
"""
Add a PO Qty column to SHIPMENTS.

    python3 add_po.py            show what would change
    python3 add_po.py --apply    do it

Three numbers, each meaning one thing:
  PO Qty        what was ordered
  Shipped Qty   what actually left Egypt
  Received      what arrived, recorded in MOVES

That gives two variances instead of one confused figure: the supplier short
against the order, and the loss in transit.

Nothing existing moves. The column is added at the end of the table.
"""
import sys, io
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

HEADER, FIRST = 6, 7
COLUMN = "PO Qty"


def add(data: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["SHIPMENTS"]
    heads = [ws.cell(HEADER, i).value for i in range(1, ws.max_column + 1)]
    if COLUMN in heads:
        return None, heads

    # insert it beside Shipped Qty so the two read together
    at = heads.index("Shipped Qty") + 2 if "Shipped Qty" in heads \
        else ws.max_column + 1
    ws.insert_cols(at)
    h = ws.cell(HEADER, at)
    h.value = COLUMN
    src = ws.cell(HEADER, at - 1)
    h.font = Font(name=src.font.name or "Arial", size=src.font.size or 9,
                  bold=True, color="FFFFFF")
    h.fill = PatternFill("solid", fgColor="152947")
    h.alignment = Alignment("center", "center", wrap_text=True)
    th = Side("thin", color="BFBFBF")
    h.border = Border(th, th, th, th)
    last = max(ws.max_row, FIRST)
    for r in range(FIRST, last + 1):
        c = ws.cell(r, at)
        c.font = Font(name="Arial", size=9)
        c.border = Border(th, th, th, th)
        c.alignment = Alignment("center")
    if "tblShipment" in ws.tables:
        del ws.tables["tblShipment"]
    t = Table(displayName="tblShipment",
              ref=f"A{HEADER}:{CL(ws.max_column)}{last}")
    t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9",
                                      showRowStripes=True)
    ws.add_table(t)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), [ws.cell(HEADER, i).value
                            for i in range(1, ws.max_column + 1)]


def main():
    import sharepoint_loader as sp, engine
    apply = "--apply" in sys.argv
    buf, meta = sp.fetch_workbook()
    data = buf.getvalue()
    print(f"{meta['name']} · saved {meta['modified']}\n")

    new, heads = add(data)
    if new is None:
        print(f"  {COLUMN} is already there. Nothing to do.")
        print(f"  SHIPMENTS: {heads}")
        return 0
    print("BEFORE")
    old = openpyxl.load_workbook(io.BytesIO(data))["SHIPMENTS"]
    print(f"  {[old.cell(HEADER, i).value for i in range(1, old.max_column + 1)]}")
    print("\nAFTER")
    print(f"  {heads}")
    s, m, c, cfg, e = engine.load(io.BytesIO(new))
    print(f"\n  {len(s)} shipment lines · {len(m)} movements · "
          f"{len(e)} entry errors")
    if len(e):
        print("  refusing - the workbook would not be clean")
        return 1
    if not apply:
        print("\nNothing was written. Run again with --apply.")
        return 0
    sp.upload_workbook(new, etag=meta.get("etag"))
    print("\nSaved to SharePoint.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
