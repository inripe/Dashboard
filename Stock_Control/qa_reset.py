# -*- coding: utf-8 -*-
"""
Clearing the test data must remove every movement and leave MASTER untouched,
and the app must still work on an empty sheet.
"""
import sys, io, types, os
import openpyxl
import engine, reset_data, qa_book
P,F=[],[]
def ck(n,ok,note=""):
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}  {note}")
def eq(n,got,want):
    ok = (abs(float(got)-float(want)) < 1e-6) if isinstance(want,(int,float)) \
        else got==want
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}: got {got}, want {want}")

data = qa_book.data()
s0,m0,c0,cfg0,e0 = engine.load(io.BytesIO(data))
new, before = reset_data.clear(data)
s1,m1,c1,cfg1,e1 = engine.load(io.BytesIO(new))

print("=== A. THE DATA IS GONE ===")
eq("no shipment lines", len(s1), 0)
eq("no movements", len(m1), 0)
eq("no counts", len(c1), 0)
eq("no stock", float(engine.stock_by_item(s1,m1,cfg1["as_of"])["Store"].sum()
                     if len(s1) else 0), 0)
ck("it reports what it removed", before.get("SHIPMENTS", 0) == len(s0),
   before)

print("=== B. MASTER IS UNTOUCHED ===")
eq("every item is still there",
   len(cfg1.get("item_names") or {}), len(cfg0.get("item_names") or {}))
eq("the markets", cfg1.get("markets"), cfg0.get("markets"))
eq("the couriers", cfg1.get("couriers_by_market"), cfg0.get("couriers_by_market"))
eq("the users", sorted(cfg1.get("users") or {}), sorted(cfg0.get("users") or {}))
eq("the reasons", len(cfg1.get("reasons") or []), len(cfg0.get("reasons") or []))
eq("the clearance target", cfg1.get("clear_target"), cfg0.get("clear_target"))
eq("the loss target", cfg1.get("loss_target"), cfg0.get("loss_target"))

print("=== C. THE SHEETS STILL HAVE THEIR SHAPE ===")
a = openpyxl.load_workbook(io.BytesIO(data))
b = openpyxl.load_workbook(io.BytesIO(new))
eq("same sheets", sorted(b.sheetnames), sorted(a.sheetnames))
for sh in ("SHIPMENTS", "MOVES", "COUNT"):
    ha = [a[sh].cell(6, i).value for i in range(1, a[sh].max_column + 1)]
    hb = [b[sh].cell(6, i).value for i in range(1, b[sh].max_column + 1)]
    eq(f"{sh} keeps every column", hb, ha)
    ck(f"{sh} still has its table", any(t for t in b[sh].tables),
       list(b[sh].tables))
eq("no entry errors on the empty sheet", len(e1), 0)

print("=== D. IT IS SAFE TO RUN TWICE ===")
again, before2 = reset_data.clear(new)
s2,m2,c2,cfg2,e2 = engine.load(io.BytesIO(again))
eq("still empty", len(m2), 0)
eq("MASTER still whole", len(cfg2.get("item_names") or {}),
   len(cfg0.get("item_names") or {}))

print("=== E. NOTHING IS WRITTEN WITHOUT --apply ===")
src = open("reset_data.py").read()
ck("upload needs the flag", 'if not apply:' in src and "upload_workbook" in src)
ck("a backup is written first",
   src.index("open(backup") < src.index("sp.upload_workbook"))
ck("it refuses if MASTER would lose something",
   "MASTER lost something" in src)
ck("and if the result would not be clean", "would not be clean" in src)
ck("MASTER is never in the list of sheets to clear",
   "MASTER" not in reset_data.CLEAR, sorted(reset_data.CLEAR))

print("=== F. THE APP WORKS ON AN EMPTY SHEET ===")
# point the app at the cleared copy before it is imported, or it reads the
# full workbook and this proves nothing
open("/tmp/qa_reset_empty.xlsx", "wb").write(new)
os.environ["QA_BOOK"] = "/tmp/qa_reset_empty.xlsx"
os.environ["INRIPE_FILE"] = "/tmp/qa_reset_empty.xlsx"
os.environ.update({"ENTRY_PASSWORD":"e","DISPATCH_PASSWORD":"d","ADMIN_PASSWORD":"a"})
mock=types.ModuleType("shopify_reader")
mock.configured_markets=lambda: []
mock.is_configured=lambda market=None: False
mock.missing_keys=lambda market=None: ["x"]
mock.market=lambda: None
mock.fetch_orders=lambda *a,**k: ([], False)
mock.MARKETS=("Qatar","UAE","KSA","Egypt")
sys.modules["shopify_reader"]=mock
from streamlit.testing.v1 import AppTest
for mode, tabs in (("Record",3), ("Dispatch",1), ("Review",7)):
    at=AppTest.from_file("app.py",default_timeout=600).run()
    md=[r for r in at.radio if "Review" in r.options]
    if md: md[0].set_value(mode).run()
    ck(f"{mode} renders", not at.exception,
       str(at.exception[0].value)[:70] if at.exception else "")
    eq(f"{mode} has its tabs", len(at.tabs), tabs)

print("=== G. NUMBERING STARTS AGAIN ===")
at=AppTest.from_file("app.py",default_timeout=600).run()
[r for r in at.radio if "Review" in r.options][0].set_value("Record").run()
us=[x for x in at.selectbox if x.label=="User"]
if us:
    us[0].set_value(qa_book.admin_user(cfg1) or list(cfg1["users"])[0]).run()
    at.text_input[0].set_value("a").run()
    [b for b in at.button if "Sign in" in str(b.label)][0].click().run()
    mk=[x for x in at.selectbox if "1 \u00b7 Market" in str(x.label)]
    if mk:
        mk[0].set_value("Qatar").run()
        sn=[x for x in at.selectbox if "Shipment number" in str(x.label)][0]
        ck("the first shipment is 001",
           any("Q-26-001" in str(o) for o in sn.options), sn.options)
        ck("and nothing else is offered", len(sn.options) == 1, sn.options)

print()
for l in F: print(l)
print(f"\n{len(P)} passed, {len(F)} failed")
sys.exit(1 if F else 0)
