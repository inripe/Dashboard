# -*- coding: utf-8 -*-
"""
No movement may take out more than exists. These are the checks that stop a
tired person at six in the morning turning stock negative.
"""
import sys, io, datetime as dt
import openpyxl, engine, entry, entry_ui, qa_book
P,F=[],[]
def ck(n,ok,note=""):
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}  {note}")
def eq(n,got,want,tol=1e-6):
    try: ok=abs(float(got)-float(want))<=tol
    except (TypeError,ValueError): ok=str(got)==str(want)
    (P if ok else F).append(f"{'PASS' if ok else 'FAIL'}  {n}: got {got}, want {want}")

base = qa_book.data()
s0,m0,c0,cfg0,e0 = engine.load(io.BytesIO(base))
MKT  = s0["Market"].dropna().iloc[0]
SHIP = s0[s0.Market==MKT]["Shipment ID"].iloc[0]
ITEM = s0[s0["Shipment ID"]==SHIP]["Item Name"].iloc[0]
COUR = (cfg0.get("couriers_by_market") or {}).get(MKT,[None])[0]
wb   = openpyxl.load_workbook(io.BytesIO(base))
store, with_courier, received = entry.stock_now(wb, SHIP)
HAVE = store.get(ITEM, 0.0)
SENT = entry.shipped_qty(wb, SHIP, ITEM)

def row(mv, **kw):
    r = {"Date": dt.date(2026,8,26), "Shipment No": SHIP, "Movement": mv}
    r.update(kw); return r
def why(mv, **kw):
    return entry.check_quantities(row(mv, **kw), wb)

print("=== A. IT KNOWS WHAT IS THERE ===")
ck("it reads the store", HAVE > 0, f"{ITEM}: {HAVE}")
ck("it reads what was shipped", SENT > 0, SENT)
ck("received is tracked separately", ITEM in received, list(received)[:3])
eq("store never exceeds received for a fresh item",
   min(HAVE, received.get(ITEM,0)), HAVE)

print("=== B. YOU CANNOT TAKE OUT MORE THAN EXISTS ===")
# Return to Scrap is left out here: it also needs returns to exist first,
# which section F covers on its own
for mv, extra in [("Scrap", {"Reason":"Damage"}),
                  ("To Courier", {"Courier": COUR} if COUR else {}),
                  ("Count Adjustment - Remove", {"Reason":"Count Adjustment"})]:
    w = why(mv, **{"Item Name":ITEM, "Out":HAVE+1, **extra})
    ck(f"{mv}: taking {HAVE+1:,.0f} when {HAVE:,.0f} exists is refused",
       w != "OK", w[:58])
    ck(f"{mv}: the message says how many there are",
       f"{HAVE:,.0f}" in w, w[:58])
    w2 = why(mv, **{"Item Name":ITEM, "Out":HAVE, **extra})
    ck(f"{mv}: taking exactly {HAVE:,.0f} is allowed", w2=="OK", w2[:50])
ck("taking one box is fine",
   why("Scrap", **{"Item Name":ITEM,"Out":1,"Reason":"Damage"})=="OK")

print("=== C. RECEIVING MORE THAN WAS SHIPPED IS ALLOWED ===")
# the shipped figure is sometimes typed wrong. Refusing the entry would only
# make the store correct the sheet, and the error would vanish. So it is
# recorded as it happened and reported in Data check instead.
already = received.get(ITEM, 0.0)
w = why("Received", **{"Item Name":ITEM, "In":SENT+1})
ck("more than shipped is accepted", w=="OK", w[:60])
b_over, _ = entry.append_moves(base, [row("Received",
    **{"Item Name":ITEM, "In":SENT+1})], "admin", MKT)
s_o, m_o, c_o, cfg_o, e_o = engine.load(io.BytesIO(b_over))
st_o = engine.stock_by_item(s_o, m_o, cfg_o["as_of"])
eq("and it is written", len(m_o), len(m0)+1)
eq("with no entry errors", len(e_o), 0)
ck("data check spots it",
   int((st_o["Received"] - st_o["Shipped Qty"] > 0.001).sum()) >= 1,
   int((st_o["Received"] - st_o["Shipped Qty"] > 0.001).sum()))
ck("the app reports it by name",
   "More received than was shipped" in open("app.py").read())
ck("and marks it high priority",
   'received against' in open("app.py").read())
ck("the form warns before it is typed",
   "flagged in Data check" in open("entry_ui.py").read())
ck("nothing is capped on the form",
   "return None, note" in open("entry_ui.py").read())

print("=== D. CUSTOMS CANNOT EXCEED THE SHORTFALL ===")
w = why("Not received", **{"Item Name":ITEM, "Out":SENT+1, "Reason":"Customs"})
ck("claiming more lost than was sent is refused", w!="OK", w[:60])
ck("it explains the arithmetic", "lost in transit" in w or "arrived" in w, w[:60])

print("=== E. THE COURIER CANNOT DELIVER WHAT IT DOES NOT HOLD ===")
w = why("Returned", **{"Item Name":ITEM,"In":with_courier+10, "Courier":COUR,
                       "Reason":"Cancelled"})
ck("it says how many the courier has",
   f"{with_courier:,.0f}" in w, w[:60])
ck("returning more than the courier holds is refused", w!="OK", w[:60])

print("=== F. RETURNS MUST COME BACK BEFORE THEY ARE SORTED ===")
for mv, kw in [("Return to Saleable", {"Item Name":ITEM,"In":10}),
               ("Return to Scrap", {"Item Name":ITEM,"Out":10,"Reason":"Damage"})]:
    w = why(mv, **kw)
    ck(f"{mv}: refused when nothing came back", w!="OK", w[:60])
    ck(f"{mv}: it says to record the return first",
       "Returned movement first" in w or "left to sort" in w, w[:60])
# and once something has come back, sorting up to that amount is allowed
if COUR:
    b = base
    b,_ = entry.append_moves(b, [row("To Courier", **{"Item Name":ITEM,"Out":5,
                                                      "Courier":COUR})], "admin", MKT)
    b,_ = entry.append_moves(b, [row("Returned", **{"Item Name":ITEM,"In":5,
                                                    "Courier":COUR,
                                                    "Reason":"Cancelled"})],
                             "admin", MKT)
    wb_r = openpyxl.load_workbook(io.BytesIO(b))
    ok = entry.check_quantities(row("Return to Saleable",
                                    **{"Item Name":ITEM,"In":5}), wb_r)
    ck("sorting exactly what came back is allowed", ok=="OK", ok[:56])
    over = entry.check_quantities(row("Return to Saleable",
                                      **{"Item Name":ITEM,"In":6}), wb_r)
    ck("sorting one more than came back is refused", over!="OK", over[:56])

print("=== G. THE WRITER ENFORCES IT, NOT JUST THE FORM ===")
before = len(engine.load(io.BytesIO(base))[1])
try:
    entry.append_moves(base, [row("Scrap", **{"Item Name":ITEM,
                                              "Out":HAVE+50,"Reason":"Damage"})],
                       "admin", MKT)
    ck("a too-large row cannot be written", False, "it was written")
except ValueError as ex:
    ck("a too-large row cannot be written", True, str(ex)[:52])
eq("and nothing was added", len(engine.load(io.BytesIO(base))[1]), before)
try:
    entry.append_moves(base, [row("Scrap", **{"Item Name":ITEM,"Out":1,
                                              "Reason":"Damage"}),
                              row("Scrap", **{"Item Name":ITEM,"Out":HAVE+50,
                                              "Reason":"Damage"})], "admin", MKT)
    ck("one bad row stops the whole batch", False, "it wrote")
except ValueError:
    ck("one bad row stops the whole batch", True)
eq("still nothing added", len(engine.load(io.BytesIO(base))[1]), before)

print("=== H. STOCK CAN NEVER GO NEGATIVE ===")
b = base
ok_qty = int(HAVE)
if ok_qty > 0:
    b,_ = entry.append_moves(b, [row("Scrap", **{"Item Name":ITEM,
                                                 "Out":ok_qty,"Reason":"Damage"})],
                             "admin", MKT)
    s1,m1,c1,cfg1,e1 = engine.load(io.BytesIO(b))
    st1 = engine.stock_by_item(s1,m1,cfg1["as_of"])
    eq("emptying the item leaves zero, not less",
       float(st1[(st1["Shipment"]==SHIP)]["Store"].sum() if len(st1) else 0) >= 0, True)
    ck("nothing anywhere is negative", (st1["Store"]>=0).all(),
       int((st1["Store"]<0).sum()))
    wb2 = openpyxl.load_workbook(io.BytesIO(b))
    w = entry.check_quantities(row("Scrap", **{"Item Name":ITEM,"Out":1,
                                               "Reason":"Damage"}), wb2)
    ck("and one more box is then refused", w!="OK", w[:56])

print("=== I. THE FORM SHOWS THE LIMIT BEFORE YOU TYPE ===")
for mv in ["Received","Scrap","To Courier","Returned",
           "Return to Saleable","Not received"]:
    cap, note = entry_ui.limits(m0, s0, SHIP, ITEM, mv)
    # Received is deliberately uncapped: more than shipped is allowed
    ck(f"{mv}: a limit is offered where one applies",
       (cap is None) == (mv == "Received"), f"cap={cap}")
    ck(f"{mv}: it is explained", bool(note), note[:44])
    ck(f"{mv}: never negative", cap is None or cap >= 0, cap)
cap,_ = entry_ui.limits(m0, s0, SHIP, ITEM, "Scrap")
eq("the scrap limit matches the store", cap, HAVE)
cap,_ = entry_ui.limits(m0, s0, SHIP, ITEM, "Returned")
eq("the returned limit matches the courier", cap, max(with_courier,0))

print("=== J. VOIDED ROWS DO NOT COUNT ===")
b2,ids = entry.append_moves(base, [row("Scrap", **{"Item Name":ITEM,"Out":1,
                                                   "Reason":"Damage"})],
                            "admin", MKT)
wbv = openpyxl.load_workbook(io.BytesIO(b2))
after_add = entry.stock_now(wbv, SHIP)[0].get(ITEM, 0)
eq("the scrap lowered the store", after_add, HAVE-1)
b3 = entry.void_entry(b2, ids[0], "admin", MKT)
wbv2 = openpyxl.load_workbook(io.BytesIO(b3))
eq("voiding it puts the box back",
   entry.stock_now(wbv2, SHIP)[0].get(ITEM, 0), HAVE)

print("=== K. A MISSING BOX CANNOT BE CLAIMED TWICE ===")
import openpyxl as _ox
_wb=_ox.load_workbook(io.BytesIO(base))
_sent=entry.shipped_qty(_wb,SHIP,ITEM)
_arr=received.get(ITEM,0.0)
_claimed=entry._not_received(_wb,SHIP).get(ITEM,0.0)
ck("the guard knows what is already claimed",
   isinstance(_claimed,float), _claimed)
_room=_sent-_arr-_claimed
w=why("Not received", **{"Item Name":ITEM,"Out":max(_room,0)+1,"Reason":"Customs"})
ck("claiming more than is missing is refused", w!="OK", w[:70])
if _room>0:
    ok=why("Not received", **{"Item Name":ITEM,"Out":_room,"Reason":"Customs"})
    ck("claiming exactly what is missing is allowed", ok=="OK", ok[:60])
else:
    ck("with nothing missing, nothing can be claimed",
       "accounted for" in w or w!="OK", w[:60])
# and once claimed, it cannot be claimed again
if _room>0:
    _b,_=entry.append_moves(base,[row("Not received",
        **{"Item Name":ITEM,"Out":_room,"Reason":"Customs"})],"admin",MKT)
    _wb2=_ox.load_workbook(io.BytesIO(_b))
    _again=entry.check_quantities(row("Not received",
        **{"Item Name":ITEM,"Out":1,"Reason":"Customs"}), _wb2)
    ck("a second claim for the same box is refused", _again!="OK", _again[:70])
    ck("and says everything is accounted for",
       "accounted for" in _again, _again[:70])

print("=== L. THE FORM SHOWS WHAT IS LEFT TO CLAIM ===")
cap,note=entry_ui.limits(m0,s0,SHIP,ITEM,"Not received")
ck("a limit is offered", cap is not None, cap)
ck("it counts what is already recorded as missing",
   "already" in note, note)
ck("the limit never goes below zero", cap>=0, cap)

print()
for l in F: print(l)
print(f"\n{len(P)} passed, {len(F)} failed")
sys.exit(1 if F else 0)
